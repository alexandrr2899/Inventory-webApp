"""admin_ops.py — Backups, importación Excel y gestión de usuarios."""
from .common import *  # noqa: F401,F403
from .stock import *   # noqa: F401,F403

from ..services.backups import (
    backup_root as _backup_root,
    ejecutar_backup,
    listar_backups as _listar_backups,
)


def _puede_gestionar_backups(user):
    return user.is_authenticated and (
        user.is_superuser or user.has_perm(_perm('gestionar_backups'))
    )


def _safe_backup_file(filename):
    if Path(filename).name != filename or not filename.endswith('.sql.gz'):
        raise Http404

    root = _backup_root()
    path = (root / filename).resolve()
    try:
        path.relative_to(root)
    except ValueError:
        raise Http404

    if not path.exists() or not path.is_file():
        raise Http404
    return path


@login_required
@_timed_view('backups_panel')
def backups_panel(request):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    if request.method == 'POST':
        resultado = ejecutar_backup(usuario=request.user, origen='manual')
        if resultado['ok']:
            messages.success(
                request,
                f'Backup creado correctamente: {resultado["backup"]["filename"]}.',
            )
        else:
            messages.error(
                request,
                resultado['mensaje'] or 'No se pudo crear el backup. Revisá los logs del servidor.',
            )
        return redirect('backups_panel')

    return render(request, 'backups/panel.html', {
        'backups': _listar_backups(),
        'jobs': BackupJob.objects.select_related('usuario')[:10],
        'backup_root_label': 'postgres/',
        'webhook_configurado': bool(getattr(settings, 'N8N_WEBHOOK_URL', '')),
    })


@login_required
def backup_descargar(request, filename):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    path = _safe_backup_file(filename)
    return FileResponse(
        open(path, 'rb'),
        as_attachment=True,
        filename=path.name,
        content_type='application/gzip',
    )




# ─── IMPORTAR EXCEL ───────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def importar_items(request):
    resultados = None

    if request.method == 'POST':
        form = ImportarItemsForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            resultados = _procesar_excel_items(archivo)
    else:
        form = ImportarItemsForm()

    return render(request, 'importar/form.html', {
        'form': form,
        'resultados': resultados,
    })


def _procesar_excel_items(archivo):
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl no está instalado.', 'creados': 0, 'actualizados': 0, 'errores': []}

    creados = 0
    actualizados = 0
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {'error': f'No se pudo leer el archivo: {e}', 'creados': 0, 'actualizados': 0, 'errores': []}

    if not rows:
        return {'error': 'El archivo está vacío.', 'creados': 0, 'actualizados': 0, 'errores': []}

    # First row must be headers; find column indices
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    col = {name: headers.index(name) for name in ('codigo', 'nombre', 'tipo', 'unidad_medida') if name in headers}

    required = ['codigo', 'nombre', 'tipo', 'unidad_medida']
    missing = [r for r in required if r not in col]
    if missing:
        return {
            'error': f'Columnas requeridas faltantes: {", ".join(missing)}.',
            'creados': 0, 'actualizados': 0, 'errores': [],
        }

    col_opt = lambda name: headers.index(name) if name in headers else None
    idx_desc = col_opt('descripcion')
    idx_cat = col_opt('categoria')
    idx_min = col_opt('stock_minimo')

    TIPOS_VALIDOS = {'producto', 'repuesto', 'consumible'}

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            codigo = str(row[col['codigo']] or '').strip()
            nombre = str(row[col['nombre']] or '').strip()
            tipo = str(row[col['tipo']] or '').strip().lower()
            unidad = str(row[col['unidad_medida']] or '').strip()

            if not codigo or not nombre:
                errores.append(f'Fila {row_num}: código o nombre vacío, se omite.')
                continue
            if tipo not in TIPOS_VALIDOS:
                errores.append(f'Fila {row_num} ({codigo}): tipo "{tipo}" inválido (usar: producto/repuesto/consumible).')
                continue
            if not unidad:
                errores.append(f'Fila {row_num} ({codigo}): unidad de medida requerida.')
                continue

            descripcion = str(row[idx_desc] or '').strip() if idx_desc is not None else ''

            categoria = None
            if idx_cat is not None and row[idx_cat]:
                cat_nombre = str(row[idx_cat]).strip()
                if cat_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=cat_nombre)

            stock_minimo = Decimal('0')
            if idx_min is not None and row[idx_min] is not None:
                try:
                    stock_minimo = Decimal(str(row[idx_min]))
                except Exception:
                    pass

            item, created = Item.objects.update_or_create(
                codigo=codigo,
                defaults={
                    'nombre': nombre,
                    'tipo': tipo,
                    'unidad_medida': unidad,
                    'descripcion': descripcion,
                    'categoria': categoria,
                    'stock_minimo': stock_minimo,
                    'activo': True,
                },
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        except Exception as e:
            errores.append(f'Fila {row_num}: error inesperado — {e}')

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'error': None}


@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def descargar_plantilla(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado.')
        return redirect('importar_items')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items'

    headers = ['codigo', 'nombre', 'tipo', 'unidad_medida', 'stock_minimo', 'categoria', 'descripcion']
    header_fill = PatternFill('solid', fgColor='003087')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    ejemplos = [
        ['BOLSA-001', 'Bolsa de polietileno 10x15', 'producto', 'unidad', 100, 'Bolsas', ''],
        ['REP-001', 'Rodamiento 6205', 'repuesto', 'pieza', 5, 'Repuestos', 'Para máquina selladora'],
        ['CONS-001', 'Cinta adhesiva', 'consumible', 'rollo', 10, 'Consumibles', ''],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_items.xlsx"'
    return response


# ─── USUARIOS ─────────────────────────────────────────────────────────────────

def _superuser_required(view_func):
    """Restrict view to superusers only; log denied attempts."""
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not request.user.is_superuser:
            ip = _get_client_ip(request)
            send_security_event(
                'forbidden_403',
                title    = '🚫 Acceso denegado a ruta administrativa',
                user     = request.user.username,
                ip       = ip,
                path     = request.path,
                user_agent = request.META.get('HTTP_USER_AGENT', '')[:200],
                message  = f'Usuario "{request.user.username}" sin superuser intentó acceder a {request.path}',
            )
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


def _get_grupo(user):
    return user.groups.first().name if user.groups.exists() else ''


@login_required
@_superuser_required
def usuario_lista(request):
    usuarios = (
        User.objects.prefetch_related('groups')
        .order_by('username')
    )
    data = [
        {'user': u, 'grupo': _get_grupo(u)}
        for u in usuarios
    ]
    return render(request, 'usuarios/lista.html', {'data': data})


@login_required
@_superuser_required
def usuario_crear(request):
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = User.objects.create_user(
                username=cd['username'],
                password=cd['password'],
                first_name=cd.get('first_name', ''),
                last_name=cd.get('last_name', ''),
                email=cd.get('email', ''),
            )
            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    user.groups.set([grupo])
                except Group.DoesNotExist:
                    pass
            messages.success(request, f'Usuario "{user.username}" creado exitosamente.')
            return redirect('usuario_lista')
    else:
        form = UsuarioCrearForm()

    return render(request, 'usuarios/form.html', {'form': form, 'titulo': 'Nuevo Usuario'})


@login_required
@_superuser_required
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            usuario.first_name = cd.get('first_name', '')
            usuario.last_name = cd.get('last_name', '')
            usuario.email = cd.get('email', '')
            usuario.is_active = cd.get('is_active', True)

            nueva_pass = cd.get('nueva_password')
            if nueva_pass:
                usuario.set_password(nueva_pass)

            usuario.save()

            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    usuario.groups.set([grupo])
                except Group.DoesNotExist:
                    usuario.groups.clear()
            else:
                usuario.groups.clear()

            messages.success(request, f'Usuario "{usuario.username}" actualizado.')
            return redirect('usuario_lista')
    else:
        form = UsuarioEditarForm(initial={
            'first_name': usuario.first_name,
            'last_name': usuario.last_name,
            'email': usuario.email,
            'is_active': usuario.is_active,
            'grupo': _get_grupo(usuario),
        })

    return render(request, 'usuarios/form.html', {
        'form': form,
        'titulo': f'Editar: {usuario.username}',
        'usuario': usuario,
    })
