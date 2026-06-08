"""
main.py — Vistas de core aún no extraídas a módulos dedicados.

Refactor incremental: las vistas se van moviendo a módulos temáticos.
Lo que queda aquí usa los helpers compartidos vía estos imports.
"""
from .common import *  # noqa: F401,F403
from .stock import *   # noqa: F401,F403
from .calc import *    # noqa: F401,F403
from .payloads import *  # noqa: F401,F403


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@login_required
@_timed_view('dashboard')
def dashboard(request):
    hoy = timezone.localdate()
    cache_key = f'dashboard:data:{hoy.isoformat()}'
    salidas_del_dia = _calcular_salidas_camiseta_del_dia(hoy)
    produccion_hoy = _calcular_produccion(hoy, salidas_parciales_hasta=timezone.now())
    produccion_hoy['salidas_del_dia'] = salidas_del_dia['total']
    cached = cache.get(cache_key)
    if cached:
        cached = cached.copy()
        cached['hoy'] = hoy
        cached['salidas_del_dia'] = salidas_del_dia
        cached['produccion_hoy'] = produccion_hoy
        return render(request, 'dashboard.html', cached)

    # Single query: annotate stock total, filter bajo_stock in DB (no N+1)
    items_bajo_stock = list(
        Item.objects
        .filter(activo=True)
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .select_related('categoria')
        .order_by('orden', 'nombre')[:20]
    )

    ultimos_detalles = (
        DetalleMovimiento.objects
        .filter(movimiento__eliminado=False)
        .select_related('item', 'movimiento', 'movimiento__usuario')
        .order_by('-movimiento__fecha')[:10]
    )

    hace_30 = timezone.now() - timedelta(days=30)
    repuestos_top = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gte=hace_30,
            item__tipo='repuesto',
        )
        .values('item__nombre', 'item__unidad_medida')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')[:5]
    )

    total_bajo_stock = (
        Item.objects
        .filter(activo=True)
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .count()
    )

    total_pendientes_conciliacion = (
        DetalleMovimiento.objects
        .filter(
            pendiente_conciliacion=True,
            movimiento__anulado=False,
            movimiento__eliminado=False,
        )
        .count()
    )

    context = {
        'items_bajo_stock': items_bajo_stock,
        'ultimos_detalles': ultimos_detalles,
        'produccion_hoy': produccion_hoy,
        'salidas_del_dia': salidas_del_dia,
        'repuestos_top': repuestos_top,
        'hoy': hoy,
        'total_bajo_stock': total_bajo_stock,
        'total_pendientes_conciliacion': total_pendientes_conciliacion,
        'total_alertas': total_bajo_stock + total_pendientes_conciliacion,
    }
    cache.set(cache_key, context, 45)
    return render(request, 'dashboard.html', context)


# ─── ALERTAS OPERATIVAS ──────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('alertas_centro')
def alertas_centro(request):
    cache_key = 'alertas:centro:v1'
    cached = cache.get(cache_key)
    if cached:
        return render(request, 'alertas/centro.html', cached)

    now = timezone.now()
    alertas = []

    items_bajo = list(
        Item.objects
        .filter(activo=True)
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .order_by('stock_calc', 'orden', 'nombre')[:60]
    )
    for item in items_bajo:
        stock = item.stock_calc or Decimal('0')
        severidad = 'critica' if stock <= 0 else 'alta'
        alertas.append({
            'tipo': 'Stock',
            'severidad': severidad,
            'icono': 'bi-exclamation-octagon-fill' if severidad == 'critica' else 'bi-exclamation-triangle-fill',
            'titulo': f'{item.nombre} sin stock' if stock <= 0 else f'{item.nombre} bajo mínimo',
            'detalle': f'{stock} {item.unidad_medida} disponible · mínimo {item.stock_minimo}',
            'fecha': now,
            'referencia': item.codigo,
            'url': reverse('item_detalle', args=[item.pk]),
            'accion': 'Revisar ítem',
        })

    pigmentos_criticos = [
        item for item in items_bajo
        if item.tipo == 'consumible'
        and item.categoria
        and item.categoria.nombre.lower().strip() == 'pigmentos'
    ]
    fecha_inicio_pig = (now.date() - timedelta(days=30))
    pigmentos_qs = list(
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .order_by('orden', 'nombre')[:40]
    )
    consumo_pigmentos = {
        row['item_id']: abs(row['total'])
        for row in (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='ajuste',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__date__gte=fecha_inicio_pig,
                movimiento__fecha_movimiento__date__lte=now.date(),
                item__tipo='consumible',
                item__categoria__nombre__iexact='Pigmentos',
                cantidad__lt=0,
            )
            .values('item_id')
            .annotate(total=Sum('cantidad'))
        )
    }
    pigmentos_panel = []
    for pig in pigmentos_qs:
        consumo_30 = consumo_pigmentos.get(pig.pk, Decimal('0'))
        stock = pig.stock_calc or Decimal('0')
        promedio = consumo_30 / Decimal('30') if consumo_30 > 0 else Decimal('0')
        dias_cobertura = (stock / promedio) if promedio > 0 else None
        pedido = max(Decimal('0'), promedio * Decimal('14') - stock) if promedio > 0 else Decimal('0')
        if dias_cobertura is None:
            estado = 'Sin dato'
        elif dias_cobertura < 3:
            estado = 'Crítico'
        elif dias_cobertura <= 7:
            estado = 'Bajo'
        else:
            estado = 'OK'
        pigmentos_panel.append({
            'item': pig,
            'stock': stock,
            'consumo_30': consumo_30,
            'dias_cobertura': round(dias_cobertura, 1) if dias_cobertura is not None else None,
            'pedido': round(pedido, 2),
            'estado': estado,
        })

    conteos_pendientes = list(
        Conteo.objects
        .filter(anulado=False)
        .exclude(estado='conciliado')
        .select_related('usuario')
        .annotate(num_detalles=Count('detalles'))
        .order_by('-fecha', 'turno')[:25]
    )
    for conteo in conteos_pendientes:
        alertas.append({
            'tipo': 'Conteo',
            'severidad': 'media',
            'icono': 'bi-clipboard-data-fill',
            'titulo': f'Conteo {conteo.get_tipo_conteo_display()} pendiente',
            'detalle': f'{conteo.get_turno_display()} · {conteo.num_detalles} línea(s)',
            'fecha': conteo.fecha_hora_conteo,
            'referencia': conteo.fecha.strftime('%d/%m/%Y'),
            'url': reverse('conteo_detalle', args=[conteo.pk]),
            'accion': 'Revisar conteo',
        })

    pendientes_stock = list(
        DetalleMovimiento.objects
        .filter(
            pendiente_conciliacion=True,
            movimiento__anulado=False,
            movimiento__eliminado=False,
        )
        .select_related('item', 'ubicacion_origen', 'movimiento', 'movimiento__usuario')
        .order_by('-movimiento__fecha_movimiento')[:40]
    )
    for det in pendientes_stock:
        alertas.append({
            'tipo': 'Conciliación',
            'severidad': 'alta',
            'icono': 'bi-arrow-repeat',
            'titulo': f'Salida pendiente de conciliación: {det.item.nombre}',
            'detalle': f'{det.cantidad} {det.item.unidad_medida} · {det.ubicacion_origen.nombre if det.ubicacion_origen else "Sin ubicación"}',
            'fecha': det.movimiento.fecha_movimiento,
            'referencia': f'Op. #{det.movimiento_id}',
            'url': reverse('movimiento_detalle', args=[det.movimiento_id]),
            'accion': 'Ver movimiento',
        })

    movimientos_auditados = list(
        MovimientoInventario.objects
        .filter(Q(anulado=True) | Q(eliminado=True))
        .select_related('usuario', 'usuario_anulacion', 'usuario_eliminacion')
        .prefetch_related('detalles__item')
        .order_by('-fecha_movimiento')[:25]
    )
    for mov in movimientos_auditados:
        if mov.eliminado:
            titulo = f'Movimiento eliminado #{mov.pk}'
            fecha = mov.fecha_eliminacion or mov.fecha_movimiento
            detalle = mov.motivo_eliminacion or 'Eliminación lógica registrada'
            severidad = 'media'
        else:
            titulo = f'Movimiento anulado #{mov.pk}'
            fecha = mov.fecha_anulacion or mov.fecha_movimiento
            detalle = mov.motivo_anulacion or 'Anulación registrada'
            severidad = 'baja'
        alertas.append({
            'tipo': 'Auditoría',
            'severidad': severidad,
            'icono': 'bi-shield-exclamation',
            'titulo': titulo,
            'detalle': detalle,
            'fecha': fecha,
            'referencia': mov.get_tipo_movimiento_display(),
            'url': reverse('movimiento_detalle', args=[mov.pk]),
            'accion': 'Ver auditoría',
        })

    alertas.sort(key=lambda a: a['fecha'] or now, reverse=True)
    context = {
        'alertas': alertas[:120],
        'total_alertas': len(alertas),
        'total_stock': len(items_bajo),
        'total_stock_cero': sum(1 for item in items_bajo if (item.stock_calc or Decimal('0')) <= 0),
        'total_pigmentos': len(pigmentos_criticos),
        'total_conteos': len(conteos_pendientes),
        'total_pendientes_stock': len(pendientes_stock),
        'total_auditados': len(movimientos_auditados),
        'pigmentos_panel': pigmentos_panel[:12],
    }
    cache.set(cache_key, context, 60)
    return render(request, 'alertas/centro.html', context)


# ─── NOTIFICACIONES MANUALES ─────────────────────────────────────────────────

@login_required
@_timed_view('notificaciones_panel')
def notificaciones_panel(request):
    if not _puede_enviar_notificaciones(request.user):
        raise PermissionDenied

    reportes = [
        {'key': key, **{k: v for k, v in cfg.items() if k not in ('builder', 'event_type')}}
        for key, cfg in _REPORTES_MANUALES.items()
    ]

    if request.method == 'POST':
        tipo = request.POST.get('tipo', '').strip()
        cfg = _REPORTES_MANUALES.get(tipo)
        if not cfg:
            messages.error(request, 'Reporte no válido.')
            return redirect('notificaciones_panel')

        payload = cfg['builder']()
        payload['enviado_por'] = request.user.username
        ok = send_event(cfg['event_type'], payload)
        event_log.info('[EVENT] reporte_manual_enviado user=%s tipo=%s ok=%s', request.user.username, tipo, ok)

        if ok:
            messages.success(request, f'Reporte enviado: {cfg["titulo"]}.')
        elif not getattr(settings, 'N8N_WEBHOOK_URL', ''):
            messages.warning(request, 'N8N_WEBHOOK_URL no está configurado. El reporte se generó, pero no se envió.')
        else:
            messages.warning(request, 'No se pudo enviar el reporte al webhook. Revisá logs o n8n.')
        return redirect('notificaciones_panel')

    return render(request, 'notificaciones/panel.html', {
        'reportes': reportes,
        'webhook_configurado': bool(getattr(settings, 'N8N_WEBHOOK_URL', '')),
    })


def _puede_gestionar_backups(user):
    return user.is_authenticated and (
        user.is_superuser or user.has_perm(_perm('gestionar_backups'))
    )


def _backup_root():
    root_env = os.environ.get('BACKUP_ROOT')
    if root_env:
        return Path(root_env).resolve()
    base_dir = Path(os.environ.get('BACKUP_DIR', settings.BASE_DIR / 'backups'))
    return (base_dir / 'postgres').resolve()


def _format_size(size):
    if size >= 1024 * 1024:
        return f'{size / (1024 * 1024):.1f} MB'
    if size >= 1024:
        return f'{size / 1024:.1f} KB'
    return f'{size} B'


def _listar_backups():
    root = _backup_root()
    if not root.exists():
        return []

    backups = []
    for path in sorted(root.glob('*.sql.gz'), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = path.stat()
        backups.append({
            'filename': path.name,
            'relative_path': f'postgres/{path.name}',
            'created_at': timezone.localtime(dt_datetime.fromtimestamp(stat.st_mtime, tz=timezone.get_current_timezone())),
            'size': stat.st_size,
            'size_label': _format_size(stat.st_size),
            'estado': 'disponible' if stat.st_size > 0 else 'vacío',
        })
    return backups


def _safe_backup_file(filename):
    if Path(filename).name != filename or not filename.endswith('.sql.gz'):
        raise Http404

    root = _backup_root()
    path = (root / filename).resolve()
    try:
        path.relative_to(root)
    except ValueError:
        raise Http404

    if not path.exists() or not path.is_file():
        raise Http404
    return path


def _backup_env(root):
    env = os.environ.copy()
    env.update({
        'POSTGRES_HOST': env.get('POSTGRES_HOST') or env.get('DB_HOST') or 'db',
        'POSTGRES_PORT': env.get('POSTGRES_PORT') or env.get('DB_PORT') or '5432',
        'POSTGRES_DB': env.get('POSTGRES_DB') or env.get('DB_NAME') or 'bolsas_inventario',
        'POSTGRES_USER': env.get('POSTGRES_USER') or env.get('DB_USER') or 'bolsas_user',
        'POSTGRES_PASSWORD': env.get('POSTGRES_PASSWORD') or env.get('DB_PASSWORD') or '',
        'BACKUP_ROOT': str(root),
        'BACKUP_RETENTION_DAYS': env.get('BACKUP_RETENTION_DAYS', '14'),
    })
    return env


@login_required
@_timed_view('backups_panel')
def backups_panel(request):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    root = _backup_root()

    if request.method == 'POST':
        job = BackupJob.objects.create(usuario=request.user)
        script_path = (Path(settings.BASE_DIR) / 'scripts' / 'backup_postgres.sh').resolve()
        timeout = int(os.environ.get('BACKUP_TIMEOUT_SECONDS', '300'))
        before = {b['filename'] for b in _listar_backups()}

        try:
            if not script_path.exists():
                raise FileNotFoundError('Script de backup no encontrado.')

            root.mkdir(parents=True, exist_ok=True)
            result = subprocess.run(
                ['sh', str(script_path)],
                cwd=str(settings.BASE_DIR),
                env=_backup_env(root),
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False,
            )

            after = _listar_backups()
            newest = next((b for b in after if b['filename'] not in before), after[0] if after else None)

            if result.returncode == 0 and newest and newest['size'] > 0:
                job.estado = 'exitoso'
                job.archivo = newest['relative_path']
                job.tamano = newest['size']
                messages.success(request, f'Backup creado correctamente: {newest["filename"]}.')
                event_log.info('[EVENT] backup_exitoso user=%s archivo=%s', request.user.username, newest['relative_path'])
                send_event('backup_exitoso', {
                    'archivo': newest['relative_path'],
                    'tamano': newest['size'],
                    'usuario': request.user.username,
                    'fecha': timezone.localtime().strftime('%Y-%m-%d'),
                    'hora': timezone.localtime().strftime('%H:%M:%S'),
                })
            else:
                job.estado = 'fallido'
                job.mensaje_error = 'El backup no se pudo completar. Revisar logs del servidor.'
                event_log.error(
                    'backup_failed user=%s returncode=%s stdout=%s stderr=%s',
                    request.user.username,
                    result.returncode,
                    result.stdout[-2000:],
                    result.stderr[-2000:],
                )
                messages.error(request, 'No se pudo crear el backup. Revisá los logs del servidor.')
                send_event('backup_fallido', {
                    'usuario': request.user.username,
                    'mensaje': 'El proceso de backup finalizó con error.',
                })
        except subprocess.TimeoutExpired:
            job.estado = 'fallido'
            job.mensaje_error = 'El backup excedió el tiempo máximo permitido.'
            event_log.error('backup_timeout user=%s timeout=%s', request.user.username, timeout)
            messages.error(request, 'El backup excedió el tiempo máximo permitido.')
            send_event('backup_fallido', {
                'usuario': request.user.username,
                'mensaje': 'El backup excedió el tiempo máximo permitido.',
            })
        except Exception as exc:
            job.estado = 'fallido'
            job.mensaje_error = 'No se pudo iniciar el backup.'
            event_log.exception('backup_exception user=%s error=%s', request.user.username, exc)
            messages.error(request, 'No se pudo iniciar el backup. Revisá los logs del servidor.')
            send_event('backup_fallido', {
                'usuario': request.user.username,
                'mensaje': 'No se pudo iniciar el backup.',
            })
        finally:
            job.fecha_fin = timezone.now()
            job.save(update_fields=['fecha_fin', 'estado', 'archivo', 'tamano', 'mensaje_error'])

        return redirect('backups_panel')

    return render(request, 'backups/panel.html', {
        'backups': _listar_backups(),
        'jobs': BackupJob.objects.select_related('usuario')[:10],
        'backup_root_label': 'postgres/',
        'webhook_configurado': bool(getattr(settings, 'N8N_WEBHOOK_URL', '')),
    })


@login_required
def backup_descargar(request, filename):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    path = _safe_backup_file(filename)
    return FileResponse(
        open(path, 'rb'),
        as_attachment=True,
        filename=path.name,
        content_type='application/gzip',
    )


# ─── MÁQUINAS ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def maquina_lista(request):
    maquinas = Maquina.objects.order_by('nombre')
    return render(request, 'maquinas/lista.html', {'maquinas': maquinas})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_crear(request):
    if request.method == 'POST':
        form = MaquinaForm(request.POST)
        if form.is_valid():
            m = form.save()
            messages.success(request, f'Máquina "{m.nombre}" creada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm()
    return render(request, 'maquinas/form.html', {'form': form, 'titulo': 'Nueva Máquina'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_editar(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    if request.method == 'POST':
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, f'Máquina "{maquina.nombre}" actualizada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm(instance=maquina)
    return render(request, 'maquinas/form.html', {
        'form': form, 'titulo': f'Editar: {maquina.nombre}', 'maquina': maquina
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def maquina_toggle_activo(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    maquina.activo = not maquina.activo
    maquina.save()
    messages.success(request, f'Máquina "{maquina.nombre}" {"activada" if maquina.activo else "desactivada"}.')
    return redirect('maquina_lista')


# ─── CLIENTES ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def cliente_lista(request):
    q = request.GET.get('q', '')
    clientes = Cliente.objects.order_by('nombre')
    if q:
        clientes = clientes.filter(Q(nombre__icontains=q) | Q(rtn__icontains=q))
    return render(request, 'clientes/lista.html', {'clientes': clientes, 'q': q})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_crear(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            c = form.save()
            messages.success(request, f'Cliente "{c.nombre}" creado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm()
    return render(request, 'clientes/form.html', {'form': form, 'titulo': 'Nuevo Cliente'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente "{cliente.nombre}" actualizado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/form.html', {
        'form': form, 'titulo': f'Editar: {cliente.nombre}', 'cliente': cliente
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def cliente_toggle_activo(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = not cliente.activo
    cliente.save()
    messages.success(request, f'Cliente "{cliente.nombre}" {"activado" if cliente.activo else "desactivado"}.')
    return redirect('cliente_lista')


def _label_grupo_cliente(fecha, agrupar_por):
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if agrupar_por == 'dia':
        return fecha.strftime('%d/%m/%Y'), fecha
    if agrupar_por == 'semana':
        iso = fecha.isocalendar()
        return f'Semana {iso[1]}, {iso[0]}', (iso[0], iso[1])
    return f'{meses[fecha.month]} {fecha.year}', (fecha.year, fecha.month)


def _build_salidas_cliente_context(cliente, request):
    from datetime import datetime as _dt

    hoy = date.today()

    def _parse_date(key, fallback):
        try:
            return _dt.strptime(request.GET.get(key, ''), '%Y-%m-%d').date()
        except ValueError:
            return fallback

    fecha_fin = _parse_date('fecha_fin', hoy)
    fecha_inicio = _parse_date('fecha_inicio', fecha_fin - timedelta(days=30))
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio

    agrupar_por = request.GET.get('agrupar_por', 'dia')
    if agrupar_por not in ('dia', 'semana', 'mes'):
        agrupar_por = 'dia'

    producto_pk = request.GET.get('producto', '').strip()
    try:
        producto_id = int(producto_pk) if producto_pk else None
    except ValueError:
        producto_id = None

    productos = (
        Item.objects
        .filter(activo=True, tipo='producto')
        .select_related('categoria')
        .order_by('orden', 'nombre')
    )

    detalles = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__cliente=cliente,
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__date__gte=fecha_inicio,
            movimiento__fecha_movimiento__date__lte=fecha_fin,
        )
        .select_related('movimiento', 'movimiento__usuario', 'item', 'item__categoria')
        .order_by('movimiento__fecha_movimiento', 'movimiento_id', 'item__orden', 'item__nombre')
    )
    if producto_id:
        detalles = detalles.filter(item_id=producto_id)

    detalles = list(detalles)
    total_general = sum((d.cantidad for d in detalles), Decimal('0'))
    movimiento_ids = {d.movimiento_id for d in detalles}
    producto_ids = {d.item_id for d in detalles}

    por_producto_map = {}
    for det in detalles:
        data = por_producto_map.setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
            'unidad': det.item.unidad_medida,
        })
        data['cantidad'] += det.cantidad
    por_producto = sorted(por_producto_map.values(), key=lambda row: _orden_operativo_producto(row['item']))

    grupos_map = {}
    for det in detalles:
        fecha_local = timezone.localtime(det.movimiento.fecha_movimiento).date()
        label, key = _label_grupo_cliente(fecha_local, agrupar_por)
        grupo = grupos_map.setdefault(key, {
            'label': label,
            'fecha': fecha_local,
            'total': Decimal('0'),
            'items': {},
        })
        grupo['total'] += det.cantidad
        item_row = grupo['items'].setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
        })
        item_row['cantidad'] += det.cantidad

    grupos = []
    for key in sorted(grupos_map):
        grupo = grupos_map[key]
        grupo['items'] = sorted(grupo['items'].values(), key=lambda row: _orden_operativo_producto(row['item']))
        grupos.append(grupo)

    movimientos_map = {}
    for det in detalles:
        mov = det.movimiento
        row = movimientos_map.setdefault(mov.pk, {
            'movimiento': mov,
            'total': Decimal('0'),
            'num_items': 0,
            'items': [],
        })
        row['total'] += det.cantidad
        row['num_items'] += 1
        row['items'].append(det)
    movimientos = sorted(
        movimientos_map.values(),
        key=lambda row: row['movimiento'].fecha_movimiento,
        reverse=True,
    )

    return {
        'cliente': cliente,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'agrupar_por': agrupar_por,
        'producto_id': producto_id,
        'productos': productos,
        'total_general': total_general,
        'num_movimientos': len(movimiento_ids),
        'num_productos': len(producto_ids),
        'por_producto': por_producto,
        'grupos': grupos,
        'movimientos': movimientos,
    }


def _exportar_salidas_cliente_excel(context):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Cliente', context['cliente'].nombre])
    ws.append(['Desde', context['fecha_inicio'].strftime('%d/%m/%Y')])
    ws.append(['Hasta', context['fecha_fin'].strftime('%d/%m/%Y')])
    ws.append(['Total fardos/items', float(context['total_general'])])
    ws.append(['Movimientos', context['num_movimientos']])
    ws.append(['Productos distintos', context['num_productos']])
    for cell in ws['A']:
        cell.font = Font(bold=True)

    ws_prod = wb.create_sheet('Por producto')
    ws_prod.append(['Producto', 'Código', 'Cantidad total', 'Unidad'])
    for row in ws_prod[1]:
        row.font = Font(bold=True)
    for r in context['por_producto']:
        ws_prod.append([r['item'].nombre, r['item'].codigo, float(r['cantidad']), r['unidad']])

    ws_mov = wb.create_sheet('Movimientos detallados')
    ws_mov.append(['Fecha movimiento', 'Movimiento ID', 'Usuario', 'Producto', 'Código', 'Cantidad', 'Unidad', 'Motivo'])
    for row in ws_mov[1]:
        row.font = Font(bold=True)
    for mov_row in context['movimientos']:
        mov = mov_row['movimiento']
        for det in mov_row['items']:
            ws_mov.append([
                timezone.localtime(mov.fecha_movimiento).strftime('%d/%m/%Y %H:%M'),
                mov.pk,
                mov.usuario.username,
                det.item.nombre,
                det.item.codigo,
                float(det.cantidad),
                det.item.unidad_medida,
                mov.motivo,
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'salidas_cliente_{context["cliente"].pk}_{context["fecha_inicio"]}_{context["fecha_fin"]}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('cliente_salidas')
def cliente_salidas(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, activo=True)
    context = _build_salidas_cliente_context(cliente, request)
    if request.GET.get('export') == 'xlsx':
        response = _exportar_salidas_cliente_excel(context)
        if response is not None:
            return response
        messages.warning(request, 'openpyxl no está disponible para exportar Excel.')
    return render(request, 'clientes/salidas.html', context)


# ─── API ──────────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def api_categoria_nueva(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
        cat, created = Categoria.objects.get_or_create(nombre=nombre)
        return JsonResponse({'id': cat.pk, 'nombre': cat.nombre, 'created': created})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_item_info(request, pk):
    item = get_object_or_404(Item, pk=pk)
    stocks = list(Stock.objects.filter(item=item).values('ubicacion__id', 'ubicacion__nombre', 'cantidad_actual'))
    return JsonResponse({
        'tipo': item.tipo,
        'unidad_medida': item.unidad_medida,
        'stock_total': str(item.stock_total()),
        'stocks': [
            {'ubicacion_id': s['ubicacion__id'], 'ubicacion': s['ubicacion__nombre'],
             'cantidad': str(s['cantidad_actual'])}
            for s in stocks
        ]
    })


# ─── PRODUCCIÓN ───────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('registrar_produccion'), raise_exception=True)
def produccion_nueva(request):
    ubicaciones = Ubicacion.objects.all()

    if request.method == 'POST':
        form = ProduccionForm(request.POST)
        if form.is_valid():
            item = form.cleaned_data['item']
            cantidad = form.cleaned_data['cantidad']
            ubicacion = form.cleaned_data['ubicacion_destino']
            fecha_movimiento = form.cleaned_data['fecha_movimiento']
            motivo = form.cleaned_data.get('motivo', '') or f'Producción registrada'

            if timezone.is_naive(fecha_movimiento):
                fecha_movimiento = timezone.make_aware(fecha_movimiento)

            with transaction.atomic():
                mov_prod = MovimientoInventario.objects.create(
                    tipo_movimiento='entrada',
                    motivo=motivo,
                    fecha_movimiento=fecha_movimiento,
                    usuario=request.user,
                )
                det_prod = DetalleMovimiento.objects.create(
                    movimiento=mov_prod,
                    item=item,
                    cantidad=cantidad,
                    ubicacion_destino=ubicacion,
                )
                _aplicar_efecto_detalle(det_prod)

            send_event('production_created', {
                'item': item.nombre, 'codigo': item.codigo,
                'cantidad': str(cantidad), 'ubicacion': ubicacion.nombre,
                'usuario': request.user.username,
                'fecha_movimiento': fecha_movimiento.isoformat(),
            })
            notify_stock(item, movimiento='produccion', usuario=request.user.username)

            messages.success(request, f'Producción registrada: {cantidad} {item.unidad_medida} de {item.nombre}.')
            return redirect('produccion_nueva')
    else:
        form = ProduccionForm(initial={
            'fecha_movimiento': timezone.localtime(timezone.now()).strftime('%Y-%m-%dT%H:%M'),
        })

    return render(request, 'produccion/form.html', {
        'form': form,
        'ubicaciones': ubicaciones,
    })


# ─── IMPORTAR EXCEL ───────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def importar_items(request):
    resultados = None

    if request.method == 'POST':
        form = ImportarItemsForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            resultados = _procesar_excel_items(archivo)
    else:
        form = ImportarItemsForm()

    return render(request, 'importar/form.html', {
        'form': form,
        'resultados': resultados,
    })


def _procesar_excel_items(archivo):
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl no está instalado.', 'creados': 0, 'actualizados': 0, 'errores': []}

    creados = 0
    actualizados = 0
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {'error': f'No se pudo leer el archivo: {e}', 'creados': 0, 'actualizados': 0, 'errores': []}

    if not rows:
        return {'error': 'El archivo está vacío.', 'creados': 0, 'actualizados': 0, 'errores': []}

    # First row must be headers; find column indices
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    col = {name: headers.index(name) for name in ('codigo', 'nombre', 'tipo', 'unidad_medida') if name in headers}

    required = ['codigo', 'nombre', 'tipo', 'unidad_medida']
    missing = [r for r in required if r not in col]
    if missing:
        return {
            'error': f'Columnas requeridas faltantes: {", ".join(missing)}.',
            'creados': 0, 'actualizados': 0, 'errores': [],
        }

    col_opt = lambda name: headers.index(name) if name in headers else None
    idx_desc = col_opt('descripcion')
    idx_cat = col_opt('categoria')
    idx_min = col_opt('stock_minimo')

    TIPOS_VALIDOS = {'producto', 'repuesto', 'consumible'}

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            codigo = str(row[col['codigo']] or '').strip()
            nombre = str(row[col['nombre']] or '').strip()
            tipo = str(row[col['tipo']] or '').strip().lower()
            unidad = str(row[col['unidad_medida']] or '').strip()

            if not codigo or not nombre:
                errores.append(f'Fila {row_num}: código o nombre vacío, se omite.')
                continue
            if tipo not in TIPOS_VALIDOS:
                errores.append(f'Fila {row_num} ({codigo}): tipo "{tipo}" inválido (usar: producto/repuesto/consumible).')
                continue
            if not unidad:
                errores.append(f'Fila {row_num} ({codigo}): unidad de medida requerida.')
                continue

            descripcion = str(row[idx_desc] or '').strip() if idx_desc is not None else ''

            categoria = None
            if idx_cat is not None and row[idx_cat]:
                cat_nombre = str(row[idx_cat]).strip()
                if cat_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=cat_nombre)

            stock_minimo = Decimal('0')
            if idx_min is not None and row[idx_min] is not None:
                try:
                    stock_minimo = Decimal(str(row[idx_min]))
                except Exception:
                    pass

            item, created = Item.objects.update_or_create(
                codigo=codigo,
                defaults={
                    'nombre': nombre,
                    'tipo': tipo,
                    'unidad_medida': unidad,
                    'descripcion': descripcion,
                    'categoria': categoria,
                    'stock_minimo': stock_minimo,
                    'activo': True,
                },
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        except Exception as e:
            errores.append(f'Fila {row_num}: error inesperado — {e}')

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'error': None}


@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def descargar_plantilla(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado.')
        return redirect('importar_items')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items'

    headers = ['codigo', 'nombre', 'tipo', 'unidad_medida', 'stock_minimo', 'categoria', 'descripcion']
    header_fill = PatternFill('solid', fgColor='003087')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    ejemplos = [
        ['BOLSA-001', 'Bolsa de polietileno 10x15', 'producto', 'unidad', 100, 'Bolsas', ''],
        ['REP-001', 'Rodamiento 6205', 'repuesto', 'pieza', 5, 'Repuestos', 'Para máquina selladora'],
        ['CONS-001', 'Cinta adhesiva', 'consumible', 'rollo', 10, 'Consumibles', ''],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_items.xlsx"'
    return response


# ─── USUARIOS ─────────────────────────────────────────────────────────────────

def _superuser_required(view_func):
    """Restrict view to superusers only; log denied attempts."""
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not request.user.is_superuser:
            ip = _get_client_ip(request)
            send_security_event(
                'forbidden_403',
                title    = '🚫 Acceso denegado a ruta administrativa',
                user     = request.user.username,
                ip       = ip,
                path     = request.path,
                user_agent = request.META.get('HTTP_USER_AGENT', '')[:200],
                message  = f'Usuario "{request.user.username}" sin superuser intentó acceder a {request.path}',
            )
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


def _get_grupo(user):
    return user.groups.first().name if user.groups.exists() else ''


@login_required
@_superuser_required
def usuario_lista(request):
    usuarios = (
        User.objects.prefetch_related('groups')
        .order_by('username')
    )
    data = [
        {'user': u, 'grupo': _get_grupo(u)}
        for u in usuarios
    ]
    return render(request, 'usuarios/lista.html', {'data': data})


@login_required
@_superuser_required
def usuario_crear(request):
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = User.objects.create_user(
                username=cd['username'],
                password=cd['password'],
                first_name=cd.get('first_name', ''),
                last_name=cd.get('last_name', ''),
                email=cd.get('email', ''),
            )
            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    user.groups.set([grupo])
                except Group.DoesNotExist:
                    pass
            messages.success(request, f'Usuario "{user.username}" creado exitosamente.')
            return redirect('usuario_lista')
    else:
        form = UsuarioCrearForm()

    return render(request, 'usuarios/form.html', {'form': form, 'titulo': 'Nuevo Usuario'})


@login_required
@_superuser_required
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            usuario.first_name = cd.get('first_name', '')
            usuario.last_name = cd.get('last_name', '')
            usuario.email = cd.get('email', '')
            usuario.is_active = cd.get('is_active', True)

            nueva_pass = cd.get('nueva_password')
            if nueva_pass:
                usuario.set_password(nueva_pass)

            usuario.save()

            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    usuario.groups.set([grupo])
                except Group.DoesNotExist:
                    usuario.groups.clear()
            else:
                usuario.groups.clear()

            messages.success(request, f'Usuario "{usuario.username}" actualizado.')
            return redirect('usuario_lista')
    else:
        form = UsuarioEditarForm(initial={
            'first_name': usuario.first_name,
            'last_name': usuario.last_name,
            'email': usuario.email,
            'is_active': usuario.is_active,
            'grupo': _get_grupo(usuario),
        })

    return render(request, 'usuarios/form.html', {
        'form': form,
        'titulo': f'Editar: {usuario.username}',
        'usuario': usuario,
    })


# ─── REPORTE CONSUMO PIGMENTOS ────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_reportes'), raise_exception=True)
@_timed_view('reporte_consumo_pigmentos')
def reporte_consumo_pigmentos(request):
    """
    Reporte de consumo de pigmentos en un rango de fechas.

    Consumo = ajustes negativos sobre ítems de categoría Pigmentos.
    Muestra: consumo total, promedio diario, stock actual, días de cobertura,
    estado (ok / bajo / crítico) y pedido sugerido.
    """
    hoy = date.today()

    # ── Filtros ───────────────────────────────────────────────────────────────
    def _parse_date(s, fallback):
        try:
            return date.fromisoformat(s) if s else fallback
        except ValueError:
            return fallback

    fecha_inicio = _parse_date(request.GET.get('fecha_inicio', ''), hoy - timedelta(days=30))
    fecha_fin    = _parse_date(request.GET.get('fecha_fin', ''),    hoy)
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio

    pigmento_pk  = request.GET.get('pigmento', '').strip()
    try:
        dias_objetivo = max(1, int(request.GET.get('dias_objetivo', 14)))
    except (ValueError, TypeError):
        dias_objetivo = 14

    dias_rango = max(1, (fecha_fin - fecha_inicio).days + 1)

    # ── Pigmentos activos ─────────────────────────────────────────────────────
    pigmentos_qs = (
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .order_by('orden', 'nombre')
    )
    if pigmento_pk:
        pigmentos_qs = pigmentos_qs.filter(pk=pigmento_pk)

    todos_pigmentos = (
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .order_by('orden', 'nombre')
    )

    # ── Consumo por ítem: ajustes negativos en el rango ───────────────────────
    consumos_base = DetalleMovimiento.objects.filter(
        movimiento__tipo_movimiento='ajuste',
        movimiento__anulado=False,
        movimiento__eliminado=False,
        movimiento__fecha_movimiento__date__gte=fecha_inicio,
        movimiento__fecha_movimiento__date__lte=fecha_fin,
        item__tipo='consumible',
        item__categoria__nombre__iexact='Pigmentos',
        cantidad__lt=0,
    )
    if pigmento_pk:
        consumos_base = consumos_base.filter(item_id=pigmento_pk)

    consumo_por_item = {
        row['item_id']: abs(row['total'])
        for row in consumos_base.values('item_id').annotate(total=Sum('cantidad'))
    }

    # ── Construir tabla de resultados ─────────────────────────────────────────
    ESTADO_LABELS = {
        'ok':          'OK',
        'bajo':        'Bajo',
        'critico':     'Crítico',
        'sin_consumo': 'Sin consumo',
    }

    resultados = []
    total_consumo  = Decimal('0')
    total_criticos = 0
    total_pedido   = Decimal('0')

    for pig in pigmentos_qs:
        consumo = consumo_por_item.get(pig.pk, Decimal('0'))
        stock   = pig.stock_calc or Decimal('0')

        if consumo > 0:
            promedio_diario = consumo / Decimal(str(dias_rango))
            dias_cob = float(stock / promedio_diario) if promedio_diario else None
            pedido   = max(Decimal('0'), promedio_diario * Decimal(str(dias_objetivo)) - stock)
        else:
            promedio_diario = Decimal('0')
            dias_cob = None
            pedido   = Decimal('0')

        if dias_cob is None:
            estado = 'sin_consumo'
        elif dias_cob < 3:
            estado = 'critico'
            total_criticos += 1
        elif dias_cob <= 7:
            estado = 'bajo'
        else:
            estado = 'ok'

        total_consumo += consumo
        total_pedido  += pedido

        resultados.append({
            'item':           pig,
            'consumo':        consumo,
            'promedio_diario': round(promedio_diario, 2),
            'stock':          stock,
            'dias_cobertura': round(dias_cob, 1) if dias_cob is not None else None,
            'pedido':         round(pedido, 2),
            'estado':         estado,
            'estado_label':   ESTADO_LABELS[estado],
        })

    # ── Detalle de movimientos (solo cuando se filtra un pigmento) ─────────────
    detalle_movimientos = []
    if pigmento_pk:
        detalle_movimientos = list(
            consumos_base
            .select_related('movimiento', 'movimiento__usuario', 'item')
            .order_by('-movimiento__fecha_movimiento')
        )

    # ── Exportar CSV ──────────────────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        fname = f'consumo_pigmentos_{fecha_inicio}_{fecha_fin}.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        response.write('﻿')  # BOM para Excel
        writer = csv.writer(response)
        writer.writerow([
            'Pigmento', 'Código', 'Consumo total', 'Promedio diario',
            'Stock actual', 'Cobertura (días)', 'Pedido sugerido', 'Unidad', 'Estado',
        ])
        for r in resultados:
            writer.writerow([
                r['item'].nombre, r['item'].codigo,
                r['consumo'], r['promedio_diario'],
                r['stock'],
                r['dias_cobertura'] if r['dias_cobertura'] is not None else '',
                r['pedido'], r['item'].unidad_medida,
                r['estado_label'],
            ])
        return response

    return render(request, 'reportes/pigmentos.html', {
        'resultados':          resultados,
        'todos_pigmentos':     todos_pigmentos,
        'fecha_inicio':        fecha_inicio,
        'fecha_fin':           fecha_fin,
        'dias_rango':          dias_rango,
        'dias_objetivo':       dias_objetivo,
        'pigmento_pk':         pigmento_pk,
        'total_consumo':       total_consumo,
        'total_criticos':      total_criticos,
        'total_pedido':        total_pedido,
        'detalle_movimientos': detalle_movimientos,
    })
