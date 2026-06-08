"""catalogos.py — Máquinas, clientes y reporte de salidas por cliente."""
from .common import *    # noqa: F401,F403
from .payloads import *  # noqa: F401,F403


# ─── MÁQUINAS ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def maquina_lista(request):
    maquinas = Maquina.objects.order_by('nombre')
    return render(request, 'maquinas/lista.html', {'maquinas': maquinas})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_crear(request):
    if request.method == 'POST':
        form = MaquinaForm(request.POST)
        if form.is_valid():
            m = form.save()
            messages.success(request, f'Máquina "{m.nombre}" creada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm()
    return render(request, 'maquinas/form.html', {'form': form, 'titulo': 'Nueva Máquina'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_editar(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    if request.method == 'POST':
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, f'Máquina "{maquina.nombre}" actualizada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm(instance=maquina)
    return render(request, 'maquinas/form.html', {
        'form': form, 'titulo': f'Editar: {maquina.nombre}', 'maquina': maquina
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def maquina_toggle_activo(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    maquina.activo = not maquina.activo
    maquina.save()
    messages.success(request, f'Máquina "{maquina.nombre}" {"activada" if maquina.activo else "desactivada"}.')
    return redirect('maquina_lista')


# ─── CLIENTES ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def cliente_lista(request):
    q = request.GET.get('q', '')
    clientes = Cliente.objects.order_by('nombre')
    if q:
        clientes = clientes.filter(Q(nombre__icontains=q) | Q(rtn__icontains=q))
    return render(request, 'clientes/lista.html', {'clientes': clientes, 'q': q})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_crear(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            c = form.save()
            messages.success(request, f'Cliente "{c.nombre}" creado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm()
    return render(request, 'clientes/form.html', {'form': form, 'titulo': 'Nuevo Cliente'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente "{cliente.nombre}" actualizado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/form.html', {
        'form': form, 'titulo': f'Editar: {cliente.nombre}', 'cliente': cliente
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def cliente_toggle_activo(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = not cliente.activo
    cliente.save()
    messages.success(request, f'Cliente "{cliente.nombre}" {"activado" if cliente.activo else "desactivado"}.')
    return redirect('cliente_lista')


def _label_grupo_cliente(fecha, agrupar_por):
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if agrupar_por == 'dia':
        return fecha.strftime('%d/%m/%Y'), fecha
    if agrupar_por == 'semana':
        iso = fecha.isocalendar()
        return f'Semana {iso[1]}, {iso[0]}', (iso[0], iso[1])
    return f'{meses[fecha.month]} {fecha.year}', (fecha.year, fecha.month)


def _build_salidas_cliente_context(cliente, request):
    from datetime import datetime as _dt

    hoy = date.today()

    def _parse_date(key, fallback):
        try:
            return _dt.strptime(request.GET.get(key, ''), '%Y-%m-%d').date()
        except ValueError:
            return fallback

    fecha_fin = _parse_date('fecha_fin', hoy)
    fecha_inicio = _parse_date('fecha_inicio', fecha_fin - timedelta(days=30))
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio

    agrupar_por = request.GET.get('agrupar_por', 'dia')
    if agrupar_por not in ('dia', 'semana', 'mes'):
        agrupar_por = 'dia'

    producto_pk = request.GET.get('producto', '').strip()
    try:
        producto_id = int(producto_pk) if producto_pk else None
    except ValueError:
        producto_id = None

    productos = (
        Item.objects
        .filter(activo=True, tipo='producto')
        .select_related('categoria')
        .order_by('orden', 'nombre')
    )

    detalles = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__cliente=cliente,
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__date__gte=fecha_inicio,
            movimiento__fecha_movimiento__date__lte=fecha_fin,
        )
        .select_related('movimiento', 'movimiento__usuario', 'item', 'item__categoria')
        .order_by('movimiento__fecha_movimiento', 'movimiento_id', 'item__orden', 'item__nombre')
    )
    if producto_id:
        detalles = detalles.filter(item_id=producto_id)

    detalles = list(detalles)
    total_general = sum((d.cantidad for d in detalles), Decimal('0'))
    movimiento_ids = {d.movimiento_id for d in detalles}
    producto_ids = {d.item_id for d in detalles}

    por_producto_map = {}
    for det in detalles:
        data = por_producto_map.setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
            'unidad': det.item.unidad_medida,
        })
        data['cantidad'] += det.cantidad
    por_producto = sorted(por_producto_map.values(), key=lambda row: _orden_operativo_producto(row['item']))

    grupos_map = {}
    for det in detalles:
        fecha_local = timezone.localtime(det.movimiento.fecha_movimiento).date()
        label, key = _label_grupo_cliente(fecha_local, agrupar_por)
        grupo = grupos_map.setdefault(key, {
            'label': label,
            'fecha': fecha_local,
            'total': Decimal('0'),
            'items': {},
        })
        grupo['total'] += det.cantidad
        item_row = grupo['items'].setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
        })
        item_row['cantidad'] += det.cantidad

    grupos = []
    for key in sorted(grupos_map):
        grupo = grupos_map[key]
        grupo['items'] = sorted(grupo['items'].values(), key=lambda row: _orden_operativo_producto(row['item']))
        grupos.append(grupo)

    movimientos_map = {}
    for det in detalles:
        mov = det.movimiento
        row = movimientos_map.setdefault(mov.pk, {
            'movimiento': mov,
            'total': Decimal('0'),
            'num_items': 0,
            'items': [],
        })
        row['total'] += det.cantidad
        row['num_items'] += 1
        row['items'].append(det)
    movimientos = sorted(
        movimientos_map.values(),
        key=lambda row: row['movimiento'].fecha_movimiento,
        reverse=True,
    )

    return {
        'cliente': cliente,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'agrupar_por': agrupar_por,
        'producto_id': producto_id,
        'productos': productos,
        'total_general': total_general,
        'num_movimientos': len(movimiento_ids),
        'num_productos': len(producto_ids),
        'por_producto': por_producto,
        'grupos': grupos,
        'movimientos': movimientos,
    }


def _exportar_salidas_cliente_excel(context):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Cliente', context['cliente'].nombre])
    ws.append(['Desde', context['fecha_inicio'].strftime('%d/%m/%Y')])
    ws.append(['Hasta', context['fecha_fin'].strftime('%d/%m/%Y')])
    ws.append(['Total fardos/items', float(context['total_general'])])
    ws.append(['Movimientos', context['num_movimientos']])
    ws.append(['Productos distintos', context['num_productos']])
    for cell in ws['A']:
        cell.font = Font(bold=True)

    ws_prod = wb.create_sheet('Por producto')
    ws_prod.append(['Producto', 'Código', 'Cantidad total', 'Unidad'])
    for row in ws_prod[1]:
        row.font = Font(bold=True)
    for r in context['por_producto']:
        ws_prod.append([r['item'].nombre, r['item'].codigo, float(r['cantidad']), r['unidad']])

    ws_mov = wb.create_sheet('Movimientos detallados')
    ws_mov.append(['Fecha movimiento', 'Movimiento ID', 'Usuario', 'Producto', 'Código', 'Cantidad', 'Unidad', 'Motivo'])
    for row in ws_mov[1]:
        row.font = Font(bold=True)
    for mov_row in context['movimientos']:
        mov = mov_row['movimiento']
        for det in mov_row['items']:
            ws_mov.append([
                timezone.localtime(mov.fecha_movimiento).strftime('%d/%m/%Y %H:%M'),
                mov.pk,
                mov.usuario.username,
                det.item.nombre,
                det.item.codigo,
                float(det.cantidad),
                det.item.unidad_medida,
                mov.motivo,
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'salidas_cliente_{context["cliente"].pk}_{context["fecha_inicio"]}_{context["fecha_fin"]}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('cliente_salidas')
def cliente_salidas(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, activo=True)
    context = _build_salidas_cliente_context(cliente, request)
    if request.GET.get('export') == 'xlsx':
        response = _exportar_salidas_cliente_excel(context)
        if response is not None:
            return response
        messages.warning(request, 'openpyxl no está disponible para exportar Excel.')
    return render(request, 'clientes/salidas.html', context)


