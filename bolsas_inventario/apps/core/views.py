import logging
import os
import subprocess
import time
from functools import wraps
from pathlib import Path

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q, Count, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.db import transaction
from django.http import FileResponse, Http404, HttpResponse, JsonResponse, StreamingHttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.cache import cache
from django.urls import reverse
from datetime import date, timedelta, datetime as dt_datetime, time as dt_time
from decimal import Decimal
import csv
import io
import json

security_log = logging.getLogger('security')
perf_log = logging.getLogger('performance')
event_log = logging.getLogger('events')

# Shared annotation for total stock per item
_STOCK_ANN = Coalesce(
    Sum('stock__cantidad_actual'),
    Value(Decimal('0')),
    output_field=DecimalField(max_digits=12, decimal_places=2),
)

from .models import (
    Item, Categoria, Ubicacion, Stock, Maquina, Cliente,
    MovimientoInventario, DetalleMovimiento, Conteo, ConteoDetalle, BackupJob
)
from django.contrib.auth.models import User, Group
from .forms import (
    ItemForm, CategoriaForm, UbicacionForm, MaquinaForm, ClienteForm,
    MovimientoEntradaForm, MovimientoSalidaForm, MovimientoTransferenciaForm,
    ConteoForm, FiltroMovimientosForm, ProduccionForm, ImportarItemsForm,
    UsuarioCrearForm, UsuarioEditarForm,
)
from .services.notifications import notify_stock, send_event, send_security_event


def _perm(codename):
    """Shorthand permission string for core app."""
    return f'core.{codename}'


def _json_safe(data):
    """
    JSON seguro para incrustar dentro de <script>.
    Evita que valores con </script>, <, > o & rompan el bloque JS.
    """
    return (
        json.dumps(data)
        .replace('&', '\\u0026')
        .replace('<', '\\u003C')
        .replace('>', '\\u003E')
    )


def _get_client_ip(request):
    """Return real client IP, respecting X-Forwarded-For from Cloudflare Tunnel."""
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', 'unknown')


def _timed_view(name):
    """Log elapsed time for high-traffic views without changing responses."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            start = time.monotonic()
            try:
                return view_func(request, *args, **kwargs)
            finally:
                elapsed = time.monotonic() - start
                user = request.user.username if request.user.is_authenticated else 'anon'
                perf_log.info('%s %.3fs user=%s path=%s', name, elapsed, user, request.get_full_path())
        return wrapper
    return decorator


def _rango_local_dia(fecha):
    tz = timezone.get_current_timezone()
    inicio = timezone.make_aware(dt_datetime.combine(fecha, dt_time.min), tz)
    return inicio, inicio + timedelta(days=1)


def _filtro_detalle_camiseta():
    return (
        Q(item__tipo='producto')
        & (
            Q(item__categoria__nombre__icontains='camiseta')
            | Q(item__nombre__icontains='camiseta')
        )
    )


def _calcular_salidas_camiseta_del_dia(fecha):
    inicio, fin = _rango_local_dia(fecha)
    qs = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gte=inicio,
            movimiento__fecha_movimiento__lt=fin,
        )
        .filter(_filtro_detalle_camiseta())
    )
    return {
        'total': qs.aggregate(t=Sum('cantidad'))['t'] or Decimal('0'),
        'movimientos': qs.values('movimiento_id').distinct().count(),
        'productos': qs.values('item_id').distinct().count(),
        'inicio': inicio,
        'fin': fin,
    }


# ─── HELPERS DE STOCK PARA MOVIMIENTOS ────────────────────────────────────────

def _revertir_efecto_detalle(det):
    """
    Deshace el efecto de stock de un DetalleMovimiento.
    Debe llamarse dentro de transaction.atomic().

    Simétrico con _aplicar_efecto_detalle: usa get_or_create en todos los
    casos para permitir reversiones sobre filas inexistentes (queda stock
    negativo o positivo según corresponda).
    """
    tipo = det.movimiento.tipo_movimiento

    if tipo == 'entrada':
        if det.ubicacion_destino:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual -= det.cantidad
            s.save()

    elif tipo == 'salida':
        if det.ubicacion_origen:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_origen,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual += det.cantidad
            s.save()

    elif tipo == 'ajuste':
        if det.ubicacion_destino:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual -= det.cantidad
            s.save()

    elif tipo == 'transferencia':
        if det.ubicacion_origen:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_origen,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual += det.cantidad
            s.save()
        if det.ubicacion_destino:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual -= det.cantidad
            s.save()


def _aplicar_efecto_detalle(det):
    """
    Aplica el efecto de stock de un DetalleMovimiento recién creado/editado.
    Debe llamarse dentro de transaction.atomic().

    IMPORTANTE — stock negativo y pendiente_conciliacion:
        El campo `pendiente_conciliacion` es solo INFORMATIVO: marca que la
        salida se permitió con stock insuficiente. NO afecta este cálculo.
        Toda salida descuenta stock, aunque quede negativo. Si no existe fila
        Stock, se crea con la cantidad negativa correspondiente.

        Una salida pendiente debe aparecer en:
          • Stock actual (cantidad_actual), incluido valor negativo.
          • Conteos (stock sistema mostrado al usuario).
          • Reportes / kardex / historial.
          • _stock_en_momento (cálculo de stock teórico).

        Solo movimientos `anulado=True` o `eliminado=True` deben excluirse.
    """
    tipo = det.movimiento.tipo_movimiento

    if tipo == 'entrada':
        if det.ubicacion_destino:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual += det.cantidad
            s.save()

    elif tipo == 'salida':
        if det.ubicacion_origen:
            # get_or_create: si no hay fila Stock, se crea con 0 y queda negativa
            # al descontar. Salidas con pendiente_conciliacion también pasan por
            # acá — el descuento NUNCA se omite.
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_origen,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual -= det.cantidad
            s.save()

    elif tipo == 'ajuste':
        if det.ubicacion_destino:
            # Ajustes pueden ser positivos o negativos. get_or_create también
            # aquí para permitir ajuste negativo sobre ubicación sin fila previa.
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual += det.cantidad
            s.save()

    elif tipo == 'transferencia':
        if det.ubicacion_origen:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_origen,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual -= det.cantidad
            s.save()
        if det.ubicacion_destino:
            s, _ = Stock.objects.get_or_create(
                item=det.item, ubicacion=det.ubicacion_destino,
                defaults={'cantidad_actual': Decimal('0')},
            )
            s.cantidad_actual += det.cantidad
            s.save()


def _revertir_todos_los_detalles(mov):
    """Revierte el stock de TODOS los detalles de un movimiento cabecera."""
    for det in mov.detalles.select_related(
        'item', 'ubicacion_origen', 'ubicacion_destino'
    ).all():
        _revertir_efecto_detalle(det)


def _stock_en_momento(item, ubicacion, fecha_hora):
    """
    Stock teórico de un ítem/ubicación en un momento dado, usando
    fecha_movimiento como timestamp oficial de cada movimiento.

    Algoritmo:
      1. Parte del stock actual (incluye TODOS los movimientos aplicados,
         incluyendo salidas con pendiente_conciliacion=True).
      2. Resta el efecto neto de los movimientos con fecha_movimiento > fecha_hora
         (ocurrieron después del momento de interés → deben excluirse).

    NO filtra por pendiente_conciliacion: una salida pendiente cuenta para
    el stock igual que cualquier otra. Solo se excluyen movimientos anulados
    o eliminados.

    Esto es correcto independientemente de cuándo se registró cada movimiento
    en el sistema: lo que importa es su fecha_movimiento.
    """
    stock_obj = Stock.objects.filter(item=item, ubicacion=ubicacion).first()
    stock_actual = stock_obj.cantidad_actual if stock_obj else Decimal('0')

    # Movimientos cuya fecha_movimiento es POSTERIOR al momento del conteo.
    # Estos ya están aplicados al stock actual pero no debían estarlo al conteo.
    post_movs = (
        DetalleMovimiento.objects
        .filter(
            item=item,
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gt=fecha_hora,
        )
        .filter(Q(ubicacion_destino=ubicacion) | Q(ubicacion_origen=ubicacion))
        .only('cantidad', 'ubicacion_origen_id', 'ubicacion_destino_id')
    )

    # Efecto neto de esos movimientos post-conteo sobre esta ubicación
    net_post = Decimal('0')
    for pm in post_movs:
        if pm.ubicacion_destino_id == ubicacion.pk:
            net_post += pm.cantidad   # entró a esta ubicación
        if pm.ubicacion_origen_id == ubicacion.pk:
            net_post -= pm.cantidad   # salió de esta ubicación

    # Stock al momento del conteo = stock actual − efecto de post-conteo
    return stock_actual - net_post


def _movimiento_editable(mov):
    """Retorna True si el movimiento puede ser editado/anulado."""
    return not mov.anulado and not mov.eliminado


def _cerrar_pendientes_conciliacion(item, ubicacion):
    """
    Después de aplicar un ajuste, revisa si el stock del ítem en la ubicación
    es ahora ≥ 0 y, de ser así, cierra (marca como conciliadas) todas las líneas
    de DetalleMovimiento que estén pendientes para ese ítem/ubicación.

    Debe llamarse dentro de transaction.atomic().
    """
    stock_obj = Stock.objects.filter(item=item, ubicacion=ubicacion).first()
    if not stock_obj or stock_obj.cantidad_actual < Decimal('0'):
        return 0

    ahora = timezone.now()
    pendientes = DetalleMovimiento.objects.filter(
        item=item,
        ubicacion_origen=ubicacion,
        pendiente_conciliacion=True,
        movimiento__anulado=False,
        movimiento__eliminado=False,
    )
    n = pendientes.update(pendiente_conciliacion=False, fecha_conciliacion=ahora)
    return n


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@login_required
@_timed_view('dashboard')
def dashboard(request):
    hoy = timezone.localdate()
    cache_key = f'dashboard:data:{hoy.isoformat()}'
    salidas_del_dia = _calcular_salidas_camiseta_del_dia(hoy)
    produccion_hoy = _calcular_produccion(hoy, salidas_parciales_hasta=timezone.now())
    produccion_hoy['salidas_del_dia'] = salidas_del_dia['total']
    cached = cache.get(cache_key)
    if cached:
        cached = cached.copy()
        cached['hoy'] = hoy
        cached['salidas_del_dia'] = salidas_del_dia
        cached['produccion_hoy'] = produccion_hoy
        return render(request, 'dashboard.html', cached)

    # Single query: annotate stock total, filter bajo_stock in DB (no N+1)
    items_bajo_stock = list(
        Item.objects
        .filter(activo=True)
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .select_related('categoria')
        .order_by('orden', 'nombre')[:20]
    )

    ultimos_detalles = (
        DetalleMovimiento.objects
        .filter(movimiento__eliminado=False)
        .select_related('item', 'movimiento', 'movimiento__usuario')
        .order_by('-movimiento__fecha')[:10]
    )

    hace_30 = timezone.now() - timedelta(days=30)
    repuestos_top = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gte=hace_30,
            item__tipo='repuesto',
        )
        .values('item__nombre', 'item__unidad_medida')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')[:5]
    )

    total_bajo_stock = (
        Item.objects
        .filter(activo=True)
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .count()
    )

    total_pendientes_conciliacion = (
        DetalleMovimiento.objects
        .filter(
            pendiente_conciliacion=True,
            movimiento__anulado=False,
            movimiento__eliminado=False,
        )
        .count()
    )

    context = {
        'items_bajo_stock': items_bajo_stock,
        'ultimos_detalles': ultimos_detalles,
        'produccion_hoy': produccion_hoy,
        'salidas_del_dia': salidas_del_dia,
        'repuestos_top': repuestos_top,
        'hoy': hoy,
        'total_bajo_stock': total_bajo_stock,
        'total_pendientes_conciliacion': total_pendientes_conciliacion,
        'total_alertas': total_bajo_stock + total_pendientes_conciliacion,
    }
    cache.set(cache_key, context, 45)
    return render(request, 'dashboard.html', context)


def _calcular_produccion(fecha, salidas_parciales_hasta=None):
    """
    Calcula producción de día y noche usando SOLO conteos tipo Camiseta.

    Producción de día:   conteo mañana → conteo tarde  (mismo día)
    Producción de noche: conteo tarde  → conteo mañana del día siguiente

    Fórmulas:
      prod_dia   = total_tarde     - total_manana     + salidas entre ambos conteos
      prod_noche = total_manana_sig - total_tarde     + salidas entre ambos conteos

    Reglas:
    - Solo usa conteos con tipo_conteo='camiseta'.
    - Usa .first()/.last() ordenados por fecha_hora_conteo (nunca .get()).
    - Salidas: solo movimientos activos (no anulados, no eliminados) dentro
      del rango horario exacto entre los dos conteos involucrados.
    - Si falta algún conteo no lanza error — informa qué falta.
    """
    from datetime import timedelta as _td

    qs_camiseta = Conteo.objects.filter(tipo_conteo='camiseta', anulado=False)

    # Conteo mañana: el más temprano del día (fecha_hora_conteo asc)
    conteo_manana = (
        qs_camiseta.filter(fecha=fecha, turno='manana')
        .order_by('fecha_hora_conteo')
        .first()
    )
    # Conteo tarde: el más reciente del día (fecha_hora_conteo desc)
    conteo_tarde = (
        qs_camiseta.filter(fecha=fecha, turno='tarde')
        .order_by('-fecha_hora_conteo')
        .first()
    )
    # Conteo mañana del día siguiente: el más temprano
    conteo_manana_sig = (
        qs_camiseta.filter(fecha=fecha + _td(days=1), turno='manana')
        .order_by('fecha_hora_conteo')
        .first()
    )

    def _total(conteo):
        """Suma cantidad_contada de ítems tipo 'producto' en un conteo."""
        if conteo is None:
            return None
        t = (
            ConteoDetalle.objects
            .filter(conteo=conteo, item__tipo='producto')
            .aggregate(t=Sum('cantidad_contada'))['t']
        )
        return t if t is not None else Decimal('0')

    def _salidas_entre(t_inicio, t_fin):
        """
        Suma de DetalleMovimiento de salidas activas cuya fecha_movimiento
        cae estrictamente entre t_inicio y t_fin.
        Solo ítems tipo 'producto'.
        """
        if not t_inicio or not t_fin or t_inicio >= t_fin:
            return Decimal('0')
        return (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='salida',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__gte=t_inicio,
                movimiento__fecha_movimiento__lt=t_fin,
            )
            .filter(_filtro_detalle_camiseta())
            .aggregate(t=Sum('cantidad'))['t'] or Decimal('0')
        )

    total_manana    = _total(conteo_manana)
    total_tarde     = _total(conteo_tarde)
    total_manana_sig = _total(conteo_manana_sig)

    # ── Producción de día ────────────────────────────────────────────────────
    if conteo_manana and conteo_tarde:
        salidas_dia   = _salidas_entre(
            conteo_manana.fecha_hora_conteo,
            conteo_tarde.fecha_hora_conteo,
        )
        produccion_dia  = total_tarde - total_manana + salidas_dia
        falta_dia       = None
    else:
        if conteo_manana and salidas_parciales_hasta:
            salidas_dia = _salidas_entre(
                conteo_manana.fecha_hora_conteo,
                salidas_parciales_hasta,
            )
        else:
            salidas_dia = Decimal('0')
        produccion_dia = None
        if not conteo_manana and not conteo_tarde:
            falta_dia = 'conteo de mañana y tarde'
        elif not conteo_manana:
            falta_dia = 'conteo de mañana'
        else:
            falta_dia = 'conteo de tarde'

    # ── Producción de noche ──────────────────────────────────────────────────
    if conteo_tarde and conteo_manana_sig:
        salidas_noche   = _salidas_entre(
            conteo_tarde.fecha_hora_conteo,
            conteo_manana_sig.fecha_hora_conteo,
        )
        produccion_noche = total_manana_sig - total_tarde + salidas_noche
        falta_noche      = None
    else:
        if conteo_tarde and salidas_parciales_hasta:
            salidas_noche = _salidas_entre(
                conteo_tarde.fecha_hora_conteo,
                salidas_parciales_hasta,
            )
        else:
            salidas_noche = Decimal('0')
        produccion_noche = None
        if not conteo_tarde and not conteo_manana_sig:
            falta_noche = 'conteo de tarde y mañana del día siguiente'
        elif not conteo_tarde:
            falta_noche = 'conteo de tarde'
        else:
            falta_noche = 'conteo de mañana del día siguiente'

    # ── Total estimado ───────────────────────────────────────────────────────
    if produccion_dia is not None and produccion_noche is not None:
        produccion_total = produccion_dia + produccion_noche
    elif produccion_dia is not None:
        produccion_total = produccion_dia
    else:
        produccion_total = produccion_noche  # None si tampoco hay noche

    return {
        # Producción calculada
        'produccion_dia':    produccion_dia,
        'produccion_noche':  produccion_noche,
        'produccion_total':  produccion_total,
        # Totales brutos de cada conteo
        'total_manana':      total_manana,
        'total_tarde':       total_tarde,
        'total_manana_sig':  total_manana_sig,
        # Salidas por tramo
        'salidas_dia':       salidas_dia,
        'salidas_noche':     salidas_noche,
        # Existencia de conteos
        'tiene_manana':      conteo_manana is not None,
        'tiene_tarde':       conteo_tarde is not None,
        'tiene_manana_sig':  conteo_manana_sig is not None,
        # Horarios usados (para mostrar rango)
        'hora_manana':       conteo_manana.fecha_hora_conteo     if conteo_manana     else None,
        'hora_tarde':        conteo_tarde.fecha_hora_conteo      if conteo_tarde      else None,
        'hora_manana_sig':   conteo_manana_sig.fecha_hora_conteo if conteo_manana_sig else None,
        # Qué falta (string descriptivo)
        'falta_dia':         falta_dia,
        'falta_noche':       falta_noche,
        # ── Compatibilidad con reporte_produccion template ────────────────────
        'produccion':        produccion_dia,
        'conteo_manana':     total_manana,
        'conteo_tarde':      total_tarde,
        'salidas':           salidas_dia,
    }


def _calcular_produccion_rango(fecha_inicio, fecha_fin):
    """
    Calcula producción día/noche para cada fecha del rango [fecha_inicio, fecha_fin]
    usando una sola ronda de consultas DB (batch).

    Devuelve:
        {
            date: {
                'prod_dia':   Decimal,
                'prod_noche': Decimal,
                'tiene_dia':  bool,
                'tiene_noche': bool,
                'por_item':   {item_pk: {'dia': Decimal, 'noche': Decimal}},
            },
            ...
        }
    Misma lógica que _calcular_produccion(fecha) pero vectorizada.
    """
    # Extender un día para capturar el conteo mañana del día siguiente
    fecha_fin_ext = fecha_fin + timedelta(days=1)

    # ── 1. Conteos ──────────────────────────────────────────────────────────────
    conteos = list(
        Conteo.objects
        .filter(tipo_conteo='camiseta', anulado=False,
                fecha__range=[fecha_inicio, fecha_fin_ext])
        .order_by('fecha', 'turno', 'fecha_hora_conteo')
    )

    # Índices: fecha → mejor conteo (mañana=earliest, tarde=latest)
    conteos_manana: dict = {}
    conteos_tarde: dict  = {}
    for c in conteos:
        if c.turno == 'manana':
            if c.fecha not in conteos_manana or c.fecha_hora_conteo < conteos_manana[c.fecha].fecha_hora_conteo:
                conteos_manana[c.fecha] = c
        elif c.turno == 'tarde':
            if c.fecha not in conteos_tarde or c.fecha_hora_conteo > conteos_tarde[c.fecha].fecha_hora_conteo:
                conteos_tarde[c.fecha] = c

    # ── 2. ConteoDetalle batch ──────────────────────────────────────────────────
    conteo_ids = [c.pk for c in conteos]
    detalles_x_conteo: dict = {}   # conteo_pk → {item_pk: cantidad}
    if conteo_ids:
        for cd in (
            ConteoDetalle.objects
            .filter(conteo_id__in=conteo_ids, item__tipo='producto')
            .select_related('item')
        ):
            detalles_x_conteo.setdefault(cd.conteo_id, {})[cd.item_id] = cd.cantidad_contada

    # ── 3. Salidas batch (para la fórmula de producción) ───────────────────────
    all_times = [c.fecha_hora_conteo for c in conteos]
    salidas_prod: list = []   # [(fecha_movimiento, item_pk, cantidad)]
    if all_times:
        t_min, t_max = min(all_times), max(all_times)
        for det in (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='salida',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__gte=t_min,
                movimiento__fecha_movimiento__lte=t_max,
                item__tipo='producto',
            )
            .select_related('movimiento')
            .only('item_id', 'cantidad', 'movimiento__fecha_movimiento')
        ):
            salidas_prod.append(
                (det.movimiento.fecha_movimiento, det.item_id, det.cantidad)
            )

    # ── Funciones auxiliares ────────────────────────────────────────────────────
    def _items(conteo_pk):
        return detalles_x_conteo.get(conteo_pk, {})

    def _salidas_items_entre(t0, t1):
        """Dict {item_pk: qty} para salidas con t0 < fecha_movimiento < t1."""
        r: dict = {}
        for t, pk, qty in salidas_prod:
            if t0 < t < t1:
                r[pk] = r.get(pk, Decimal('0')) + qty
        return r

    # ── 4. Calcular por día ─────────────────────────────────────────────────────
    result: dict = {}
    cur = fecha_inicio
    while cur <= fecha_fin:
        cm  = conteos_manana.get(cur)
        ct  = conteos_tarde.get(cur)
        cms = conteos_manana.get(cur + timedelta(days=1))

        prod_dia   = Decimal('0')
        prod_noche = Decimal('0')
        tiene_dia  = tiene_noche = False
        por_item: dict = {}

        # Producción día
        if cm and ct:
            items_m = _items(cm.pk)
            items_t = _items(ct.pk)
            sal_dia = _salidas_items_entre(cm.fecha_hora_conteo, ct.fecha_hora_conteo)

            prod_dia = (
                sum(items_t.values(), Decimal('0'))
                - sum(items_m.values(), Decimal('0'))
                + sum(sal_dia.values(), Decimal('0'))
            )
            tiene_dia = True

            for pk in set(items_m) | set(items_t):
                por_item.setdefault(pk, {'dia': Decimal('0'), 'noche': Decimal('0')})
                por_item[pk]['dia'] = (
                    items_t.get(pk, Decimal('0'))
                    - items_m.get(pk, Decimal('0'))
                    + sal_dia.get(pk, Decimal('0'))
                )

        # Producción noche
        if ct and cms:
            items_t  = _items(ct.pk)
            items_ms = _items(cms.pk)
            sal_noch = _salidas_items_entre(ct.fecha_hora_conteo, cms.fecha_hora_conteo)

            prod_noche = (
                sum(items_ms.values(), Decimal('0'))
                - sum(items_t.values(), Decimal('0'))
                + sum(sal_noch.values(), Decimal('0'))
            )
            tiene_noche = True

            for pk in set(items_t) | set(items_ms):
                por_item.setdefault(pk, {'dia': Decimal('0'), 'noche': Decimal('0')})
                por_item[pk]['noche'] = (
                    items_ms.get(pk, Decimal('0'))
                    - items_t.get(pk, Decimal('0'))
                    + sal_noch.get(pk, Decimal('0'))
                )

        result[cur] = {
            'prod_dia':    prod_dia,
            'prod_noche':  prod_noche,
            'tiene_dia':   tiene_dia,
            'tiene_noche': tiene_noche,
            'por_item':    por_item,
        }
        cur += timedelta(days=1)

    return result


def _calcular_tramos(fecha_inicio, fecha_fin):
    """
    Calcula producción por tramos entre pares de conteos consecutivos
    tipo Camiseta activos.

    Para cada par (c_ini, c_fin) asignado a una fecha dentro de [fecha_inicio, fecha_fin]:

        produccion = total_fin − total_ini + salidas_entre_conteos

    Tipos de tramo
    ──────────────
    'dia'       mañana → tarde  (mismo día)
    'noche'     tarde  → mañana (día siguiente)
    'extendido' cualquier otro (fines de semana, días sin conteo, etc.)

    Devuelve lista de dicts:
        conteo_ini, conteo_fin, tipo, duracion_h,
        produccion, salidas, por_item,
        fecha_asignada (= c_ini.fecha),
        label_rango   (ej. "Sáb 10/05 11:00 → Lun 12/05 08:00")
    """
    # Buscar 3 días antes para capturar el conteo inicial del primer tramo
    fecha_fetch_ini = fecha_inicio - timedelta(days=3)
    fecha_fetch_fin = fecha_fin + timedelta(days=1)

    conteos = list(
        Conteo.objects
        .filter(tipo_conteo='camiseta', anulado=False,
                fecha__range=[fecha_fetch_ini, fecha_fetch_fin])
        .order_by('fecha_hora_conteo')
    )

    if len(conteos) < 2:
        return []

    # ── Batch: ConteoDetalle ────────────────────────────────────────────────────
    conteo_ids = [c.pk for c in conteos]
    detalles_x_conteo: dict = {}
    for cd in (
        ConteoDetalle.objects
        .filter(conteo_id__in=conteo_ids, item__tipo='producto')
        .select_related('item')
    ):
        detalles_x_conteo.setdefault(cd.conteo_id, {})[cd.item_id] = cd.cantidad_contada

    # ── Batch: salidas en el rango temporal de los conteos ─────────────────────
    t_min = conteos[0].fecha_hora_conteo
    t_max = conteos[-1].fecha_hora_conteo
    salidas_list: list = []  # [(fecha_movimiento, item_pk, qty)]
    for det in (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gt=t_min,
            movimiento__fecha_movimiento__lt=t_max,
            item__tipo='producto',
        )
        .select_related('movimiento')
        .only('item_id', 'cantidad', 'movimiento__fecha_movimiento')
    ):
        salidas_list.append((det.movimiento.fecha_movimiento, det.item_id, det.cantidad))

    # ── Auxiliares ──────────────────────────────────────────────────────────────
    def _items(conteo_pk):
        return detalles_x_conteo.get(conteo_pk, {})

    def _salidas_entre(t0, t1):
        r: dict = {}
        for t, pk, qty in salidas_list:
            if t0 < t < t1:
                r[pk] = r.get(pk, Decimal('0')) + qty
        return r

    DIAS_ES = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    def _fmt_dt(dt):
        dt_local = timezone.localtime(dt)
        return f"{DIAS_ES[dt_local.weekday()]} {dt_local.strftime('%d/%m %H:%M')}"

    def _tipo_tramo(c_ini, c_fin):
        if (c_ini.turno == 'manana' and c_fin.turno == 'tarde'
                and c_ini.fecha == c_fin.fecha):
            return 'dia'
        if (c_ini.turno == 'tarde' and c_fin.turno == 'manana'
                and (c_fin.fecha - c_ini.fecha).days == 1):
            return 'noche'
        return 'extendido'

    # ── Construir tramos ────────────────────────────────────────────────────────
    tramos = []
    for i in range(len(conteos) - 1):
        c_ini = conteos[i]
        c_fin = conteos[i + 1]
        tipo_tramo = _tipo_tramo(c_ini, c_fin)
        fecha_asignada = c_ini.fecha

        # La producción se agrupa por el día operativo donde inicia el tramo:
        # día: mañana → tarde; noche/extendido: tarde o conteo inicial → siguiente cierre.
        if fecha_asignada < fecha_inicio or fecha_asignada > fecha_fin:
            continue

        items_ini = _items(c_ini.pk)
        items_fin = _items(c_fin.pk)
        sal_items = _salidas_entre(c_ini.fecha_hora_conteo, c_fin.fecha_hora_conteo)

        total_ini = sum(items_ini.values(), Decimal('0'))
        total_fin = sum(items_fin.values(), Decimal('0'))
        total_sal = sum(sal_items.values(), Decimal('0'))
        produccion = total_fin - total_ini + total_sal

        por_item: dict = {}
        for pk in set(items_ini) | set(items_fin):
            por_item[pk] = (
                items_fin.get(pk, Decimal('0'))
                - items_ini.get(pk, Decimal('0'))
                + sal_items.get(pk, Decimal('0'))
            )

        duracion_h = round(
            (c_fin.fecha_hora_conteo - c_ini.fecha_hora_conteo).total_seconds() / 3600, 1
        )

        tramos.append({
            'conteo_ini':     c_ini,
            'conteo_fin':     c_fin,
            'tipo':           tipo_tramo,
            'duracion_h':     duracion_h,
            'produccion':     produccion,
            'salidas':        total_sal,
            'diferencia':     produccion - total_sal,
            'por_item':       por_item,
            'fecha_asignada': fecha_asignada,
            'label_rango':    f"{_fmt_dt(c_ini.fecha_hora_conteo)} → {_fmt_dt(c_fin.fecha_hora_conteo)}",
        })

    return tramos


# ─── ALERTAS OPERATIVAS ──────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('alertas_centro')
def alertas_centro(request):
    cache_key = 'alertas:centro:v1'
    cached = cache.get(cache_key)
    if cached:
        return render(request, 'alertas/centro.html', cached)

    now = timezone.now()
    alertas = []

    items_bajo = list(
        Item.objects
        .filter(activo=True)
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .order_by('stock_calc', 'orden', 'nombre')[:60]
    )
    for item in items_bajo:
        stock = item.stock_calc or Decimal('0')
        severidad = 'critica' if stock <= 0 else 'alta'
        alertas.append({
            'tipo': 'Stock',
            'severidad': severidad,
            'icono': 'bi-exclamation-octagon-fill' if severidad == 'critica' else 'bi-exclamation-triangle-fill',
            'titulo': f'{item.nombre} sin stock' if stock <= 0 else f'{item.nombre} bajo mínimo',
            'detalle': f'{stock} {item.unidad_medida} disponible · mínimo {item.stock_minimo}',
            'fecha': now,
            'referencia': item.codigo,
            'url': reverse('item_detalle', args=[item.pk]),
            'accion': 'Revisar ítem',
        })

    pigmentos_criticos = [
        item for item in items_bajo
        if item.tipo == 'consumible'
        and item.categoria
        and item.categoria.nombre.lower().strip() == 'pigmentos'
    ]
    fecha_inicio_pig = (now.date() - timedelta(days=30))
    pigmentos_qs = list(
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .order_by('orden', 'nombre')[:40]
    )
    consumo_pigmentos = {
        row['item_id']: abs(row['total'])
        for row in (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='ajuste',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__date__gte=fecha_inicio_pig,
                movimiento__fecha_movimiento__date__lte=now.date(),
                item__tipo='consumible',
                item__categoria__nombre__iexact='Pigmentos',
                cantidad__lt=0,
            )
            .values('item_id')
            .annotate(total=Sum('cantidad'))
        )
    }
    pigmentos_panel = []
    for pig in pigmentos_qs:
        consumo_30 = consumo_pigmentos.get(pig.pk, Decimal('0'))
        stock = pig.stock_calc or Decimal('0')
        promedio = consumo_30 / Decimal('30') if consumo_30 > 0 else Decimal('0')
        dias_cobertura = (stock / promedio) if promedio > 0 else None
        pedido = max(Decimal('0'), promedio * Decimal('14') - stock) if promedio > 0 else Decimal('0')
        if dias_cobertura is None:
            estado = 'Sin dato'
        elif dias_cobertura < 3:
            estado = 'Crítico'
        elif dias_cobertura <= 7:
            estado = 'Bajo'
        else:
            estado = 'OK'
        pigmentos_panel.append({
            'item': pig,
            'stock': stock,
            'consumo_30': consumo_30,
            'dias_cobertura': round(dias_cobertura, 1) if dias_cobertura is not None else None,
            'pedido': round(pedido, 2),
            'estado': estado,
        })

    conteos_pendientes = list(
        Conteo.objects
        .filter(anulado=False)
        .exclude(estado='conciliado')
        .select_related('usuario')
        .annotate(num_detalles=Count('detalles'))
        .order_by('-fecha', 'turno')[:25]
    )
    for conteo in conteos_pendientes:
        alertas.append({
            'tipo': 'Conteo',
            'severidad': 'media',
            'icono': 'bi-clipboard-data-fill',
            'titulo': f'Conteo {conteo.get_tipo_conteo_display()} pendiente',
            'detalle': f'{conteo.get_turno_display()} · {conteo.num_detalles} línea(s)',
            'fecha': conteo.fecha_hora_conteo,
            'referencia': conteo.fecha.strftime('%d/%m/%Y'),
            'url': reverse('conteo_detalle', args=[conteo.pk]),
            'accion': 'Revisar conteo',
        })

    pendientes_stock = list(
        DetalleMovimiento.objects
        .filter(
            pendiente_conciliacion=True,
            movimiento__anulado=False,
            movimiento__eliminado=False,
        )
        .select_related('item', 'ubicacion_origen', 'movimiento', 'movimiento__usuario')
        .order_by('-movimiento__fecha_movimiento')[:40]
    )
    for det in pendientes_stock:
        alertas.append({
            'tipo': 'Conciliación',
            'severidad': 'alta',
            'icono': 'bi-arrow-repeat',
            'titulo': f'Salida pendiente de conciliación: {det.item.nombre}',
            'detalle': f'{det.cantidad} {det.item.unidad_medida} · {det.ubicacion_origen.nombre if det.ubicacion_origen else "Sin ubicación"}',
            'fecha': det.movimiento.fecha_movimiento,
            'referencia': f'Op. #{det.movimiento_id}',
            'url': reverse('movimiento_detalle', args=[det.movimiento_id]),
            'accion': 'Ver movimiento',
        })

    movimientos_auditados = list(
        MovimientoInventario.objects
        .filter(Q(anulado=True) | Q(eliminado=True))
        .select_related('usuario', 'usuario_anulacion', 'usuario_eliminacion')
        .prefetch_related('detalles__item')
        .order_by('-fecha_movimiento')[:25]
    )
    for mov in movimientos_auditados:
        if mov.eliminado:
            titulo = f'Movimiento eliminado #{mov.pk}'
            fecha = mov.fecha_eliminacion or mov.fecha_movimiento
            detalle = mov.motivo_eliminacion or 'Eliminación lógica registrada'
            severidad = 'media'
        else:
            titulo = f'Movimiento anulado #{mov.pk}'
            fecha = mov.fecha_anulacion or mov.fecha_movimiento
            detalle = mov.motivo_anulacion or 'Anulación registrada'
            severidad = 'baja'
        alertas.append({
            'tipo': 'Auditoría',
            'severidad': severidad,
            'icono': 'bi-shield-exclamation',
            'titulo': titulo,
            'detalle': detalle,
            'fecha': fecha,
            'referencia': mov.get_tipo_movimiento_display(),
            'url': reverse('movimiento_detalle', args=[mov.pk]),
            'accion': 'Ver auditoría',
        })

    alertas.sort(key=lambda a: a['fecha'] or now, reverse=True)
    context = {
        'alertas': alertas[:120],
        'total_alertas': len(alertas),
        'total_stock': len(items_bajo),
        'total_stock_cero': sum(1 for item in items_bajo if (item.stock_calc or Decimal('0')) <= 0),
        'total_pigmentos': len(pigmentos_criticos),
        'total_conteos': len(conteos_pendientes),
        'total_pendientes_stock': len(pendientes_stock),
        'total_auditados': len(movimientos_auditados),
        'pigmentos_panel': pigmentos_panel[:12],
    }
    cache.set(cache_key, context, 60)
    return render(request, 'alertas/centro.html', context)


# ─── NOTIFICACIONES MANUALES ─────────────────────────────────────────────────

def _puede_enviar_notificaciones(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['Administrador', 'Supervisor']).exists()


def _decimal_payload(value):
    if value is None:
        return None
    return float(value)


def _fecha_hora_payload():
    ahora = timezone.localtime(timezone.now())
    return ahora.strftime('%d/%m/%Y'), ahora.strftime('%H:%M')


def _inventario_camiseta_actual():
    orden_operativo = [
        'Bolsa Camiseta Grande',
        'Bolsa Camiseta Mediana',
        'Bolsa Camiseta Pequeña',
        'Bolsa Camiseta Grande Negra',
        'Bolsa Camiseta Mediana Negra',
        'Bolsa Camiseta Pequeña Negra',
    ]
    orden_map = {nombre.lower(): idx for idx, nombre in enumerate(orden_operativo)}
    items = list(
        Item.objects
        .filter(activo=True, tipo='producto')
        .filter(Q(nombre__icontains='camiseta') | Q(categoria__nombre__icontains='camiseta'))
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
    )
    items.sort(key=lambda item: (orden_map.get(item.nombre.lower(), 999), item.orden, item.nombre))
    return [
        {
            'nombre': item.nombre,
            'codigo': item.codigo,
            'stock_actual': _decimal_payload(item.stock_calc),
            'unidad': item.unidad_medida,
        }
        for item in items
    ]


def _payload_inventario_camiseta():
    fecha, hora = _fecha_hora_payload()

    return {
        'titulo': 'Inventario actual de Camiseta',
        'fecha': fecha,
        'hora': hora,
        'items': _inventario_camiseta_actual(),
    }


def _enviar_inventario_camiseta_post_conciliacion(conteo_pk, conteo_tipo, usuario=''):
    """
    Reenvía el inventario actual de Camiseta a n8n (mismo evento y payload que
    el envío manual: event_type='inventario_camiseta_actual').

    Debe invocarse SOLO vía transaction.on_commit(), es decir, después de que
    los ajustes de conciliación quedaron persistidos. Si el conteo no es de
    tipo 'camiseta', no hace nada.

    El fallo del webhook NUNCA debe afectar la conciliación: se captura y se
    registra como warning/error, pero no se propaga.
    """
    if conteo_tipo != 'camiseta':
        return
    try:
        payload = _payload_inventario_camiseta()
        payload['origen'] = 'conciliacion_automatica'
        payload['conteo_id'] = conteo_pk
        if usuario:
            payload['enviado_por'] = usuario
        ok = send_event('inventario_camiseta_actual', payload)
        if ok:
            event_log.info(
                'Inventario camiseta enviado automáticamente tras conciliación #%s',
                conteo_pk,
            )
        else:
            event_log.warning(
                'No se pudo enviar inventario camiseta tras conciliación #%s '
                '(webhook sin configurar o n8n no respondió). La conciliación NO se afecta.',
                conteo_pk,
            )
    except Exception:
        # Nunca romper la conciliación por un fallo de notificación
        event_log.exception(
            'Error al enviar inventario camiseta tras conciliación #%s. '
            'La conciliación se completó correctamente de todas formas.',
            conteo_pk,
        )


def _notificar_si_conciliacion_completa(conteo, estado_antes, usuario=''):
    """
    Programa (vía transaction.on_commit) el reenvío del inventario camiseta
    SOLO cuando el conteo recién transiciona a 'conciliado'.

    Garantiza exactamente UN envío por conteo: si ya estaba 'conciliado' antes
    de esta operación, no hace nada. Debe llamarse dentro de transaction.atomic()
    después de que el estado del conteo quedó actualizado.
    """
    if conteo.estado != 'conciliado' or estado_antes == 'conciliado':
        return
    _conteo_pk, _conteo_tipo, _usuario = conteo.pk, conteo.tipo_conteo, usuario
    transaction.on_commit(
        lambda: _enviar_inventario_camiseta_post_conciliacion(_conteo_pk, _conteo_tipo, _usuario)
    )


def _payload_inventario_pigmentos():
    from .services.notifications import generar_resumen_pigmentos
    payload = generar_resumen_pigmentos()
    payload['titulo'] = 'Inventario actual de Pigmentos'
    return payload


def _payload_stock_bajo():
    fecha, hora = _fecha_hora_payload()
    items = (
        Item.objects
        .filter(activo=True)
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .filter(stock_calc__lte=F('stock_minimo'))
        .order_by('stock_calc', 'orden', 'nombre')
    )
    bajo = []
    cero = []
    for item in items:
        fila = {
            'nombre': item.nombre,
            'codigo': item.codigo,
            'tipo': item.tipo,
            'categoria': item.categoria.nombre if item.categoria else '',
            'stock_actual': _decimal_payload(item.stock_calc),
            'stock_minimo': _decimal_payload(item.stock_minimo),
            'unidad': item.unidad_medida,
        }
        if item.stock_calc <= 0:
            cero.append(fila)
        else:
            bajo.append(fila)
    return {
        'titulo': 'Reporte de stock bajo',
        'fecha': fecha,
        'hora': hora,
        'total': len(bajo) + len(cero),
        'total_bajo': len(bajo),
        'total_cero': len(cero),
        'stock_bajo': bajo,
        'stock_cero': cero,
    }


def _salidas_camiseta_detalle_dia(fecha_operativa):
    inicio, fin = _rango_local_dia(fecha_operativa)
    detalles = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gte=inicio,
            movimiento__fecha_movimiento__lt=fin,
        )
        .filter(_filtro_detalle_camiseta())
        .select_related('movimiento', 'movimiento__cliente', 'item', 'item__categoria')
        .order_by('movimiento__fecha_movimiento', 'movimiento_id', 'item__orden', 'item__nombre')
    )

    movimientos = {}
    total = Decimal('0')
    por_producto = {}

    for det in detalles:
        mov = det.movimiento
        cliente = mov.cliente or det.cliente
        grupo = movimientos.setdefault(mov.pk, {
            'movimiento': mov,
            'cliente': cliente.nombre if cliente else 'Sin cliente',
            'items': [],
            'total': Decimal('0'),
        })
        grupo['items'].append(det)
        grupo['total'] += det.cantidad
        total += det.cantidad

        producto = por_producto.setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
        })
        producto['cantidad'] += det.cantidad

    salidas = []
    for grupo in movimientos.values():
        mov = grupo['movimiento']
        fecha_local = timezone.localtime(mov.fecha_movimiento)
        items = []
        for det in sorted(grupo['items'], key=lambda d: _orden_operativo_producto(d.item)):
            items.append({
                'nombre': det.item.nombre,
                'codigo': det.item.codigo,
                'cantidad': _decimal_payload(det.cantidad),
                'unidad': det.item.unidad_medida,
            })
        salidas.append({
            'movimiento_id': mov.pk,
            'fecha': fecha_local.strftime('%d/%m/%Y'),
            'hora': fecha_local.strftime('%H:%M'),
            'cliente': grupo['cliente'],
            'items': items,
            'total_movimiento': _decimal_payload(grupo['total']),
        })

    total_por_producto = []
    for data in sorted(por_producto.values(), key=lambda d: _orden_operativo_producto(d['item'])):
        item = data['item']
        total_por_producto.append({
            'nombre': item.nombre,
            'codigo': item.codigo,
            'cantidad': _decimal_payload(data['cantidad']),
            'unidad': item.unidad_medida,
        })

    return {
        'total': total,
        'detalle': salidas,
        'por_producto': total_por_producto,
        'inicio': inicio,
        'fin': fin,
    }


def _texto_salidas_dia_telegram(salidas_detalle, total):
    if not salidas_detalle:
        return '📤 Salidas del día\nSin salidas de producto terminado registradas.'

    bloques = ['📤 Salidas del día']
    for salida in salidas_detalle:
        bloques.append(f'👤 Cliente: {salida["cliente"]}')
        for item in salida['items']:
            bloques.append(f'• {item["nombre"]}: {item["cantidad"]:g} {item["unidad"]}')
        bloques.append(f'Total: {salida["total_movimiento"]:g} Fardo')
        bloques.append('')
    bloques.append(f'Total salidas del día: {_decimal_payload(total):g} Fardo')
    return '\n'.join(bloques).strip()


def _payload_produccion_dia():
    fecha_operativa = timezone.localdate()
    fecha, hora = _fecha_hora_payload()
    produccion = _calcular_produccion(fecha_operativa)
    salidas_completas = _salidas_camiseta_detalle_dia(fecha_operativa)

    salidas_calculo_total = (produccion['salidas_dia'] or Decimal('0')) + (produccion['salidas_noche'] or Decimal('0'))

    return {
        'titulo': 'Reporte de producción del día',
        'fecha': fecha,
        'hora': hora,
        'fecha_operativa': fecha_operativa.isoformat(),
        'produccion_dia': _decimal_payload(produccion['produccion_dia']),
        'produccion_noche': _decimal_payload(produccion['produccion_noche']),
        'produccion_total': _decimal_payload(produccion['produccion_total']),
        'salidas_dia': _decimal_payload(produccion['salidas_dia']),
        'salidas_noche': _decimal_payload(produccion['salidas_noche']),
        'salidas_calculo': {
            'dia': _decimal_payload(produccion['salidas_dia']),
            'noche': _decimal_payload(produccion['salidas_noche']),
            'total': _decimal_payload(salidas_calculo_total),
        },
        'inventario_actual': _inventario_camiseta_actual(),
        'salidas_dia_total': _decimal_payload(salidas_completas['total']),
        'salidas_del_dia_detalle': salidas_completas['detalle'],
        'salidas_del_dia_por_producto': salidas_completas['por_producto'],
        'salidas_del_dia_texto': _texto_salidas_dia_telegram(
            salidas_completas['detalle'],
            salidas_completas['total'],
        ),
    }


def _payload_salidas_dia():
    fecha_operativa = timezone.localdate()
    fecha, hora = _fecha_hora_payload()
    inicio, fin = _rango_local_dia(fecha_operativa)
    detalles = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__gte=inicio,
            movimiento__fecha_movimiento__lt=fin,
            item__tipo='producto',
        )
        .select_related('movimiento', 'item', 'cliente', 'movimiento__cliente')
        .order_by('cliente__nombre', 'movimiento__cliente__nombre', 'item__orden', 'item__nombre')
    )

    grupos = {}
    total = Decimal('0')
    for det in detalles:
        cliente = det.cliente or det.movimiento.cliente
        cliente_nombre = cliente.nombre if cliente else 'Sin cliente'
        grupo = grupos.setdefault(cliente_nombre, {'cliente': cliente_nombre, 'total': Decimal('0'), 'items': []})
        grupo['total'] += det.cantidad
        total += det.cantidad
        grupo['items'].append({
            'movimiento_id': det.movimiento_id,
            'hora': timezone.localtime(det.movimiento.fecha_movimiento).strftime('%H:%M'),
            'nombre': det.item.nombre,
            'codigo': det.item.codigo,
            'cantidad': _decimal_payload(det.cantidad),
            'unidad': det.item.unidad_medida,
        })

    salidas = []
    for grupo in grupos.values():
        salidas.append({
            'cliente': grupo['cliente'],
            'total': _decimal_payload(grupo['total']),
            'items': grupo['items'],
        })

    return {
        'titulo': 'Reporte de salidas del día',
        'fecha': fecha,
        'hora': hora,
        'fecha_operativa': fecha_operativa.isoformat(),
        'total_salidas': _decimal_payload(total),
        'clientes': salidas,
    }


_REPORTES_MANUALES = {
    'inventario_camiseta': {
        'titulo': 'Inventario Camiseta',
        'descripcion': 'Stock actual de productos Camiseta en orden operativo.',
        'event_type': 'inventario_camiseta_actual',
        'builder': _payload_inventario_camiseta,
        'icono': 'bi-bag-check-fill',
        'color': 'primary',
    },
    'inventario_pigmentos': {
        'titulo': 'Inventario Pigmentos',
        'descripcion': 'Stock, mínimo y estado de pigmentos.',
        'event_type': 'inventario_pigmentos_actual',
        'builder': _payload_inventario_pigmentos,
        'icono': 'bi-droplet-half',
        'color': 'info',
    },
    'stock_bajo': {
        'titulo': 'Stock bajo',
        'descripcion': 'Ítems activos separados entre bajo y cero.',
        'event_type': 'reporte_stock_bajo',
        'builder': _payload_stock_bajo,
        'icono': 'bi-exclamation-triangle-fill',
        'color': 'warning',
    },
    'produccion_dia': {
        'titulo': 'Producción del día',
        'descripcion': 'Producción día/noche con conteos usados y salidas incluidas.',
        'event_type': 'reporte_produccion_dia',
        'builder': _payload_produccion_dia,
        'icono': 'bi-graph-up-arrow',
        'color': 'success',
    },
    'salidas_dia': {
        'titulo': 'Salidas del día',
        'descripcion': 'Salidas activas de producto terminado agrupadas por cliente.',
        'event_type': 'reporte_salidas_dia',
        'builder': _payload_salidas_dia,
        'icono': 'bi-arrow-up-circle-fill',
        'color': 'danger',
    },
}


@login_required
@_timed_view('notificaciones_panel')
def notificaciones_panel(request):
    if not _puede_enviar_notificaciones(request.user):
        raise PermissionDenied

    reportes = [
        {'key': key, **{k: v for k, v in cfg.items() if k not in ('builder', 'event_type')}}
        for key, cfg in _REPORTES_MANUALES.items()
    ]

    if request.method == 'POST':
        tipo = request.POST.get('tipo', '').strip()
        cfg = _REPORTES_MANUALES.get(tipo)
        if not cfg:
            messages.error(request, 'Reporte no válido.')
            return redirect('notificaciones_panel')

        payload = cfg['builder']()
        payload['enviado_por'] = request.user.username
        ok = send_event(cfg['event_type'], payload)
        event_log.info('[EVENT] reporte_manual_enviado user=%s tipo=%s ok=%s', request.user.username, tipo, ok)

        if ok:
            messages.success(request, f'Reporte enviado: {cfg["titulo"]}.')
        elif not getattr(settings, 'N8N_WEBHOOK_URL', ''):
            messages.warning(request, 'N8N_WEBHOOK_URL no está configurado. El reporte se generó, pero no se envió.')
        else:
            messages.warning(request, 'No se pudo enviar el reporte al webhook. Revisá logs o n8n.')
        return redirect('notificaciones_panel')

    return render(request, 'notificaciones/panel.html', {
        'reportes': reportes,
        'webhook_configurado': bool(getattr(settings, 'N8N_WEBHOOK_URL', '')),
    })


def _puede_gestionar_backups(user):
    return user.is_authenticated and (
        user.is_superuser or user.has_perm(_perm('gestionar_backups'))
    )


def _backup_root():
    root_env = os.environ.get('BACKUP_ROOT')
    if root_env:
        return Path(root_env).resolve()
    base_dir = Path(os.environ.get('BACKUP_DIR', settings.BASE_DIR / 'backups'))
    return (base_dir / 'postgres').resolve()


def _format_size(size):
    if size >= 1024 * 1024:
        return f'{size / (1024 * 1024):.1f} MB'
    if size >= 1024:
        return f'{size / 1024:.1f} KB'
    return f'{size} B'


def _listar_backups():
    root = _backup_root()
    if not root.exists():
        return []

    backups = []
    for path in sorted(root.glob('*.sql.gz'), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = path.stat()
        backups.append({
            'filename': path.name,
            'relative_path': f'postgres/{path.name}',
            'created_at': timezone.localtime(dt_datetime.fromtimestamp(stat.st_mtime, tz=timezone.get_current_timezone())),
            'size': stat.st_size,
            'size_label': _format_size(stat.st_size),
            'estado': 'disponible' if stat.st_size > 0 else 'vacío',
        })
    return backups


def _safe_backup_file(filename):
    if Path(filename).name != filename or not filename.endswith('.sql.gz'):
        raise Http404

    root = _backup_root()
    path = (root / filename).resolve()
    try:
        path.relative_to(root)
    except ValueError:
        raise Http404

    if not path.exists() or not path.is_file():
        raise Http404
    return path


def _backup_env(root):
    env = os.environ.copy()
    env.update({
        'POSTGRES_HOST': env.get('POSTGRES_HOST') or env.get('DB_HOST') or 'db',
        'POSTGRES_PORT': env.get('POSTGRES_PORT') or env.get('DB_PORT') or '5432',
        'POSTGRES_DB': env.get('POSTGRES_DB') or env.get('DB_NAME') or 'bolsas_inventario',
        'POSTGRES_USER': env.get('POSTGRES_USER') or env.get('DB_USER') or 'bolsas_user',
        'POSTGRES_PASSWORD': env.get('POSTGRES_PASSWORD') or env.get('DB_PASSWORD') or '',
        'BACKUP_ROOT': str(root),
        'BACKUP_RETENTION_DAYS': env.get('BACKUP_RETENTION_DAYS', '14'),
    })
    return env


@login_required
@_timed_view('backups_panel')
def backups_panel(request):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    root = _backup_root()

    if request.method == 'POST':
        job = BackupJob.objects.create(usuario=request.user)
        script_path = (Path(settings.BASE_DIR) / 'scripts' / 'backup_postgres.sh').resolve()
        timeout = int(os.environ.get('BACKUP_TIMEOUT_SECONDS', '300'))
        before = {b['filename'] for b in _listar_backups()}

        try:
            if not script_path.exists():
                raise FileNotFoundError('Script de backup no encontrado.')

            root.mkdir(parents=True, exist_ok=True)
            result = subprocess.run(
                ['sh', str(script_path)],
                cwd=str(settings.BASE_DIR),
                env=_backup_env(root),
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False,
            )

            after = _listar_backups()
            newest = next((b for b in after if b['filename'] not in before), after[0] if after else None)

            if result.returncode == 0 and newest and newest['size'] > 0:
                job.estado = 'exitoso'
                job.archivo = newest['relative_path']
                job.tamano = newest['size']
                messages.success(request, f'Backup creado correctamente: {newest["filename"]}.')
                event_log.info('[EVENT] backup_exitoso user=%s archivo=%s', request.user.username, newest['relative_path'])
                send_event('backup_exitoso', {
                    'archivo': newest['relative_path'],
                    'tamano': newest['size'],
                    'usuario': request.user.username,
                    'fecha': timezone.localtime().strftime('%Y-%m-%d'),
                    'hora': timezone.localtime().strftime('%H:%M:%S'),
                })
            else:
                job.estado = 'fallido'
                job.mensaje_error = 'El backup no se pudo completar. Revisar logs del servidor.'
                event_log.error(
                    'backup_failed user=%s returncode=%s stdout=%s stderr=%s',
                    request.user.username,
                    result.returncode,
                    result.stdout[-2000:],
                    result.stderr[-2000:],
                )
                messages.error(request, 'No se pudo crear el backup. Revisá los logs del servidor.')
                send_event('backup_fallido', {
                    'usuario': request.user.username,
                    'mensaje': 'El proceso de backup finalizó con error.',
                })
        except subprocess.TimeoutExpired:
            job.estado = 'fallido'
            job.mensaje_error = 'El backup excedió el tiempo máximo permitido.'
            event_log.error('backup_timeout user=%s timeout=%s', request.user.username, timeout)
            messages.error(request, 'El backup excedió el tiempo máximo permitido.')
            send_event('backup_fallido', {
                'usuario': request.user.username,
                'mensaje': 'El backup excedió el tiempo máximo permitido.',
            })
        except Exception as exc:
            job.estado = 'fallido'
            job.mensaje_error = 'No se pudo iniciar el backup.'
            event_log.exception('backup_exception user=%s error=%s', request.user.username, exc)
            messages.error(request, 'No se pudo iniciar el backup. Revisá los logs del servidor.')
            send_event('backup_fallido', {
                'usuario': request.user.username,
                'mensaje': 'No se pudo iniciar el backup.',
            })
        finally:
            job.fecha_fin = timezone.now()
            job.save(update_fields=['fecha_fin', 'estado', 'archivo', 'tamano', 'mensaje_error'])

        return redirect('backups_panel')

    return render(request, 'backups/panel.html', {
        'backups': _listar_backups(),
        'jobs': BackupJob.objects.select_related('usuario')[:10],
        'backup_root_label': 'postgres/',
        'webhook_configurado': bool(getattr(settings, 'N8N_WEBHOOK_URL', '')),
    })


@login_required
def backup_descargar(request, filename):
    if not _puede_gestionar_backups(request.user):
        raise PermissionDenied

    path = _safe_backup_file(filename)
    return FileResponse(
        open(path, 'rb'),
        as_attachment=True,
        filename=path.name,
        content_type='application/gzip',
    )


# ─── INVENTARIO ───────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('inventario_lista')
def inventario_lista(request):
    q = request.GET.get('q', '').strip()

    qs = (
        Item.objects
        .filter(activo=True)
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .order_by('orden', 'nombre')
    )
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(codigo__icontains=q))

    paginator = Paginator(qs, 100)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_items = list(page_obj.object_list)

    # Second query: principal location per item (max stock)
    stocks_raw = (
        Stock.objects
        .filter(item__in=page_items)
        .values('item_id', 'ubicacion__nombre', 'cantidad_actual')
    )
    ub_map: dict = {}
    for s in stocks_raw:
        iid = s['item_id']
        if iid not in ub_map or s['cantidad_actual'] > ub_map[iid][0]:
            ub_map[iid] = (s['cantidad_actual'], s['ubicacion__nombre'])

    # Pendientes de conciliación por ítem (salidas con stock insuficiente activas)
    pendientes_map: dict = {}
    pend_qs = (
        DetalleMovimiento.objects
        .filter(
            item__in=page_items,
            pendiente_conciliacion=True,
            movimiento__anulado=False,
            movimiento__eliminado=False,
        )
        .values('item_id')
        .annotate(n=Count('pk'))
    )
    for row in pend_qs:
        pendientes_map[row['item_id']] = row['n']

    items_data = [
        {
            'item': item,
            'stock': item.stock_calc,
            'bajo': item.stock_calc <= item.stock_minimo,
            'negativo': item.stock_calc < 0,
            'pendientes': pendientes_map.get(item.pk, 0),
            'ub_principal': ub_map.get(item.pk, (None, '–'))[1],
        }
        for item in page_items
    ]

    context = {'items_data': items_data, 'q': q, 'page_obj': page_obj}
    return render(request, 'inventario/lista.html', context)


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def item_detalle(request, pk):
    item = get_object_or_404(Item, pk=pk)
    stocks = Stock.objects.filter(item=item).select_related('ubicacion')
    # Últimas líneas que afectan este ítem (sin movimientos eliminados)
    detalles_recientes = (
        DetalleMovimiento.objects
        .filter(item=item, movimiento__eliminado=False)
        .select_related(
            'movimiento', 'movimiento__usuario',
            'ubicacion_origen', 'ubicacion_destino', 'cliente', 'maquina',
        )
        .order_by('-movimiento__fecha_movimiento')[:20]
    )
    context = {'item': item, 'stocks': stocks, 'detalles_recientes': detalles_recientes}
    return render(request, 'inventario/detalle.html', context)


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def item_historial(request, pk):
    item = get_object_or_404(Item, pk=pk)
    tipo_filtro   = request.GET.get('tipo',   '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # Líneas que afectan este ítem (kardex por detalle, con FK al movimiento cabecera)
    todos = list(
        DetalleMovimiento.objects
        .filter(item=item)
        .select_related(
            'movimiento',
            'movimiento__usuario',
            'movimiento__usuario_edicion',
            'movimiento__usuario_anulacion',
            'movimiento__usuario_eliminacion',
            'ubicacion_origen', 'ubicacion_destino',
            'cliente', 'maquina',
        )
        .order_by('movimiento__fecha_movimiento', 'movimiento__fecha', 'movimiento__pk', 'pk')
    )

    def _delta_activo(det):
        mov = det.movimiento
        if mov.anulado or mov.eliminado:
            return Decimal('0')
        t = mov.tipo_movimiento
        if t == 'entrada':
            return det.cantidad
        if t == 'salida':
            return -det.cantidad
        if t == 'ajuste':
            return det.cantidad
        return Decimal('0')

    kardex = []
    acum = Decimal('0')
    for det in todos:
        d = _delta_activo(det)
        stock_antes   = acum
        stock_despues = acum + d
        acum = stock_despues
        kardex.append({
            'det':          det,
            'mov':          det.movimiento,   # alias conveniente para la plantilla
            'delta':        d,
            'stock_antes':  stock_antes,
            'stock_despues': stock_despues,
            'afecta_stock': not (det.movimiento.anulado or det.movimiento.eliminado),
        })

    kardex.reverse()

    tipos_validos = ('entrada', 'salida', 'ajuste', 'transferencia')
    tipo_filtro_clean = tipo_filtro if tipo_filtro in tipos_validos else ''

    kardex_filtrado = kardex
    if tipo_filtro_clean:
        kardex_filtrado = [
            k for k in kardex_filtrado
            if k['mov'].tipo_movimiento == tipo_filtro_clean
        ]

    if estado_filtro == 'activos':
        kardex_filtrado = [
            k for k in kardex_filtrado
            if not k['mov'].anulado and not k['mov'].eliminado
        ]
    elif estado_filtro == 'anulados':
        kardex_filtrado = [k for k in kardex_filtrado if k['mov'].anulado]
    elif estado_filtro == 'eliminados':
        kardex_filtrado = [k for k in kardex_filtrado if k['mov'].eliminado]

    total_general    = len(kardex)
    total_activos    = sum(1 for k in kardex if not k['mov'].anulado and not k['mov'].eliminado)
    total_anulados   = sum(1 for k in kardex if k['mov'].anulado)
    total_eliminados = sum(1 for k in kardex if k['mov'].eliminado)

    paginator = Paginator(kardex_filtrado, 20)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'inventario/historial.html', {
        'item':             item,
        'page_obj':         page_obj,
        'tipo_filtro':      tipo_filtro_clean,
        'estado_filtro':    estado_filtro,
        'total':            len(kardex_filtrado),
        'total_general':    total_general,
        'total_activos':    total_activos,
        'total_anulados':   total_anulados,
        'total_eliminados': total_eliminados,
    })


@login_required
@permission_required(_perm('crear_item'), raise_exception=True)
def item_crear(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f'Ítem "{item.nombre}" creado exitosamente.')
            return redirect('item_detalle', pk=item.pk)
    else:
        form = ItemForm()

    categorias = Categoria.objects.all()
    return render(request, 'inventario/form.html', {
        'form': form, 'titulo': 'Nuevo Ítem', 'categorias': categorias
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def item_editar(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ítem "{item.nombre}" actualizado.')
            return redirect('item_detalle', pk=item.pk)
    else:
        form = ItemForm(instance=item)

    return render(request, 'inventario/form.html', {
        'form': form, 'titulo': f'Editar: {item.nombre}', 'item': item
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def item_toggle_activo(request, pk):
    item = get_object_or_404(Item, pk=pk)
    item.activo = not item.activo
    item.save()
    estado = 'activado' if item.activo else 'desactivado'
    messages.success(request, f'Ítem "{item.nombre}" {estado}.')
    return redirect('inventario_lista')


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def ubicacion_lista(request):
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'inventario/ubicaciones.html', {'ubicaciones': ubicaciones})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def ubicacion_crear(request):
    if request.method == 'POST':
        form = UbicacionForm(request.POST)
        if form.is_valid():
            u = form.save()
            messages.success(request, f'Ubicación "{u.nombre}" creada.')
            return redirect('ubicacion_lista')
    else:
        form = UbicacionForm()
    return render(request, 'inventario/ubicacion_form.html', {
        'form': form, 'titulo': 'Nueva Ubicación'
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def ubicacion_editar(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicacion)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ubicación "{ubicacion.nombre}" actualizada.')
            return redirect('ubicacion_lista')
    else:
        form = UbicacionForm(instance=ubicacion)
    return render(request, 'inventario/ubicacion_form.html', {
        'form': form, 'titulo': f'Editar: {ubicacion.nombre}', 'ubicacion': ubicacion
    })


# ─── MOVIMIENTOS ──────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('movimiento_lista')
def movimiento_lista(request):
    form = FiltroMovimientosForm(request.GET or None)
    movimientos = (
        MovimientoInventario.objects
        .prefetch_related('detalles__item', 'detalles__ubicacion_origen',
                          'detalles__ubicacion_destino', 'detalles__cliente',
                          'detalles__maquina')
        .select_related('usuario', 'usuario_anulacion', 'usuario_edicion',
                        'usuario_eliminacion', 'cliente')
        .annotate(
            num_detalles=Count('detalles', distinct=True),
            num_pendientes=Count(
                'detalles',
                filter=Q(detalles__pendiente_conciliacion=True),
                distinct=True,
            ),
        )
        .order_by('-fecha_movimiento')
    )

    if form.is_valid():
        if form.cleaned_data.get('fecha_inicio'):
            movimientos = movimientos.filter(
                fecha_movimiento__date__gte=form.cleaned_data['fecha_inicio']
            )
        if form.cleaned_data.get('fecha_fin'):
            movimientos = movimientos.filter(
                fecha_movimiento__date__lte=form.cleaned_data['fecha_fin']
            )
        if form.cleaned_data.get('tipo_movimiento'):
            movimientos = movimientos.filter(
                tipo_movimiento=form.cleaned_data['tipo_movimiento']
            )
        if form.cleaned_data.get('item'):
            movimientos = movimientos.filter(
                detalles__item=form.cleaned_data['item']
            ).distinct()

    if request.GET.get('export') == 'csv':
        return _exportar_movimientos_csv(movimientos)

    paginator = Paginator(movimientos, 50)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'movimientos/lista.html', {
        'movimientos': page_obj,
        'page_obj': page_obj,
        'form': form,
    })


def _exportar_movimientos_csv(movimientos):
    """Exporta cada LÍNEA (DetalleMovimiento) como una fila CSV, con datos del cabecera."""
    class Echo:
        def write(self, value):
            return value

    writer = csv.writer(Echo())

    def rows():
        yield '﻿'
        yield writer.writerow([
            'Movimiento #', 'Fecha Movimiento', 'Fecha Registro', 'Tipo',
            'Ítem', 'Código', 'Cantidad', 'Unidad',
            'Origen', 'Destino', 'Cliente', 'Máquina', 'Motivo', 'Usuario',
            'Estado',
        ])
        detalles = (
            DetalleMovimiento.objects
            .filter(movimiento__in=movimientos.values('pk'))
            .select_related(
                'movimiento', 'movimiento__usuario',
                'item', 'ubicacion_origen', 'ubicacion_destino', 'cliente', 'maquina',
            )
            .order_by('-movimiento__fecha_movimiento', 'movimiento_id', 'id')
            .iterator(chunk_size=500)
        )
        for det in detalles:
            mov = det.movimiento
            estado = ('Anulado' if mov.anulado else
                      'Eliminado' if mov.eliminado else
                      'Editado' if mov.editado else 'Activo')
            yield writer.writerow([
                mov.pk,
                timezone.localtime(mov.fecha_movimiento).strftime('%Y-%m-%d %H:%M'),
                timezone.localtime(mov.fecha).strftime('%Y-%m-%d %H:%M'),
                mov.get_tipo_movimiento_display(),
                det.item.nombre,
                det.item.codigo,
                det.cantidad,
                det.item.unidad_medida,
                det.ubicacion_origen.nombre if det.ubicacion_origen else '',
                det.ubicacion_destino.nombre if det.ubicacion_destino else '',
                det.cliente.nombre if det.cliente else '',
                det.maquina.nombre if det.maquina else '',
                mov.motivo,
                mov.usuario.get_full_name() or mov.usuario.username,
                estado,
            ])

    response = StreamingHttpResponse(rows(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="movimientos.csv"'
    return response


@login_required
@permission_required(_perm('registrar_entrada'), raise_exception=True)
def movimiento_entrada(request):
    items = Item.objects.filter(activo=True).order_by('orden', 'nombre')
    ubicaciones = Ubicacion.objects.all()
    item_id_inicial = request.GET.get('item', '')

    items_json = _json_safe([
        {'pk': it.pk, 'nombre': it.nombre, 'codigo': it.codigo, 'unidad': it.unidad_medida}
        for it in items
    ])
    ubicaciones_json = _json_safe([
        {'pk': u.pk, 'nombre': u.nombre, 'tipo': u.get_tipo_display()}
        for u in ubicaciones
    ])

    if request.method == 'POST':
        ubicacion_destino_id = request.POST.get('ubicacion_destino', '').strip()
        motivo = request.POST.get('motivo', '')
        fecha_mov_str = request.POST.get('fecha_movimiento', '').strip()
        item_ids = request.POST.getlist('item[]')
        cantidades = request.POST.getlist('cantidad[]')

        errores = []
        ubicacion_destino = None

        try:
            ubicacion_destino = Ubicacion.objects.get(pk=ubicacion_destino_id)
        except (Ubicacion.DoesNotExist, ValueError):
            errores.append('Debes seleccionar una ubicación de destino.')

        fecha_movimiento = timezone.now()
        if fecha_mov_str:
            try:
                from django.utils.dateparse import parse_datetime
                parsed = parse_datetime(fecha_mov_str)
                if parsed:
                    fecha_movimiento = timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed
            except Exception:
                pass

        filas_validas = []
        for i, (item_id, cant_str) in enumerate(zip(item_ids, cantidades), 1):
            cant_str = cant_str.strip()
            if not item_id and not cant_str:
                continue
            if not item_id:
                errores.append(f'Fila {i}: selecciona un ítem.')
                continue
            if not cant_str:
                errores.append(f'Fila {i}: ingresa una cantidad.')
                continue
            try:
                cantidad = Decimal(cant_str)
                if cantidad <= 0:
                    raise ValueError
            except (ValueError, Exception):
                errores.append(f'Fila {i}: cantidad inválida.')
                continue
            try:
                item = Item.objects.get(pk=item_id, activo=True)
            except Item.DoesNotExist:
                errores.append(f'Fila {i}: ítem no encontrado.')
                continue
            filas_validas.append((item, cantidad))

        if not filas_validas and not errores:
            errores.append('Agrega al menos un ítem con cantidad.')

        if errores:
            for e in errores:
                messages.error(request, e)
            filas_previas = [
                {'item_id': iid, 'cantidad': cant}
                for iid, cant in zip(item_ids, cantidades)
                if iid or cant.strip()
            ]
            return render(request, 'movimientos/entrada.html', {
                'items_json': items_json,
                'ubicaciones_json': ubicaciones_json,
                'item_id_inicial': item_id_inicial,
                'ub_destino_previo': ubicacion_destino_id,
                'motivo_previo': motivo,
                'fecha_mov_previo': fecha_mov_str,
                'filas_previas_json': _json_safe(filas_previas),
            })

        with transaction.atomic():
            mov = MovimientoInventario.objects.create(
                tipo_movimiento='entrada',
                motivo=motivo,
                fecha_movimiento=fecha_movimiento,
                usuario=request.user,
            )
            for item, cantidad in filas_validas:
                det = DetalleMovimiento.objects.create(
                    movimiento=mov,
                    item=item,
                    cantidad=cantidad,
                    ubicacion_destino=ubicacion_destino,
                )
                _aplicar_efecto_detalle(det)
                send_event('movement_created', {
                    'tipo': 'entrada', 'item': item.nombre, 'codigo': item.codigo,
                    'cantidad': str(cantidad), 'ubicacion': ubicacion_destino.nombre,
                    'usuario': request.user.username,
                })
                notify_stock(item, movimiento='entrada', usuario=request.user.username)
        messages.success(
            request,
            f'Movimiento #{mov.pk} registrado con {len(filas_validas)} ítem(s).'
        )
        return redirect('movimiento_detalle', pk=mov.pk)

    return render(request, 'movimientos/entrada.html', {
        'items_json': items_json,
        'ubicaciones_json': ubicaciones_json,
        'item_id_inicial': item_id_inicial,
        'filas_previas_json': '[]',
    })


@login_required
@permission_required(_perm('registrar_salida'), raise_exception=True)
def movimiento_salida(request):
    """
    Registro de salidas con cuatro tabs:

    • Producto Terminado  — filas dinámicas (usuario agrega solo los necesarios),
                           dropdown filtrado a productos, orden por item.orden,
                           cliente a nivel de cabecera, permite stock negativo
                           (marca pendiente_conciliacion=True en cada línea).
                           Usa campos item_pt[] y cantidad_pt[] para evitar
                           colisión con los tabs de filas dinámicas (item[], cantidad[]).
    • Repuestos           — filas dinámicas, bloquea si stock insuficiente.
    • Consumibles         — filas dinámicas, bloquea si stock insuficiente.
    • Otros               — filas dinámicas, bloquea si stock insuficiente.
    """
    from django.utils.dateparse import parse_datetime

    # ── Datos maestros ────────────────────────────────────────────────────────
    items_producto   = list(Item.objects.filter(activo=True, tipo='producto')
                            .order_by('orden', 'nombre'))
    items_repuesto   = list(Item.objects.filter(activo=True, tipo='repuesto')
                            .order_by('orden', 'nombre'))
    items_consumible = list(Item.objects.filter(activo=True, tipo='consumible')
                            .order_by('orden', 'nombre'))
    items_otros      = list(Item.objects.filter(activo=True)
                            .exclude(tipo__in=['producto', 'repuesto', 'consumible'])
                            .order_by('orden', 'nombre'))

    ubicaciones = list(Ubicacion.objects.all().order_by('nombre'))
    clientes    = list(Cliente.objects.filter(activo=True).order_by('nombre'))
    maquinas    = list(Maquina.objects.filter(activo=True).order_by('nombre'))

    # ── Stock por item {item_pk: {ub_pk: stock_actual}} ───────────────────────
    stocks_qs = Stock.objects.select_related('item', 'ubicacion').all()
    stocks_por_item = {}
    for s in stocks_qs:
        stocks_por_item.setdefault(s.item_id, {})[s.ubicacion_id] = float(s.cantidad_actual)

    # ── JSON para JavaScript ──────────────────────────────────────────────────
    def _items_json(lst):
        return _json_safe([
            {'pk': it.pk, 'nombre': it.nombre, 'codigo': it.codigo,
             'tipo': it.tipo, 'unidad': it.unidad_medida}
            for it in lst
        ])

    ubicaciones_json = _json_safe([
        {'pk': u.pk, 'nombre': u.nombre, 'tipo': u.get_tipo_display()}
        for u in ubicaciones
    ])
    clientes_json = _json_safe([
        {'pk': c.pk, 'nombre': c.nombre} for c in clientes
    ])
    maquinas_json = _json_safe([
        {'pk': m.pk, 'nombre': m.nombre} for m in maquinas
    ])
    stocks_json = _json_safe(stocks_por_item)

    def _parse_fecha(fecha_str):
        if not fecha_str:
            return timezone.now()
        try:
            parsed = parse_datetime(fecha_str.strip())
            if parsed:
                return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed
        except Exception:
            pass
        return timezone.now()

    # ── Contexto base (GET y re-render en error) ───────────────────────────────
    def _ctx(extra=None):
        ctx = {
            'items_producto_json':   _items_json(items_producto),
            'items_repuesto_json':   _items_json(items_repuesto),
            'items_consumible_json': _items_json(items_consumible),
            'items_otros_json':      _items_json(items_otros),
            'ubicaciones_json':      ubicaciones_json,
            'clientes_json':         clientes_json,
            'maquinas_json':         maquinas_json,
            'stocks_json':           stocks_json,
            'ubicaciones':           ubicaciones,
            'clientes':              clientes,
        }
        if extra:
            ctx.update(extra)
        return ctx

    if request.method != 'POST':
        tab_inicial = request.GET.get('tab', 'producto_terminado')
        item_id_inicial = request.GET.get('item', '')
        return render(request, 'movimientos/salida.html',
                      _ctx({'tab_inicial': tab_inicial,
                            'item_id_inicial': item_id_inicial,
                            'filas_pt_previas_json': '[]',
                            'filas_previas_json': '[]'}))

    # ═══════════════════════════════════════════════════════════════════════════
    # POST
    # ═══════════════════════════════════════════════════════════════════════════
    tipo_salida   = request.POST.get('tipo_salida', 'producto_terminado')
    motivo        = request.POST.get('motivo', '').strip()
    fecha_mov_str = request.POST.get('fecha_movimiento', '')
    fecha_movimiento = _parse_fecha(fecha_mov_str)

    errores      = []
    filas_validas = []   # (item, cantidad, ubicacion, pendiente, maquina)

    # ── TAB: Producto Terminado ────────────────────────────────────────────────
    if tipo_salida == 'producto_terminado':
        cliente_id  = request.POST.get('cliente_header', '').strip()
        ub_pt_id    = request.POST.get('ubicacion_origen_pt', '').strip()

        # Filas dinámicas PT usan nombres distintos (item_pt[], cantidad_pt[])
        # para evitar colisión con los tabs de rep/con/otros (item[], cantidad[])
        item_ids_pt  = request.POST.getlist('item_pt[]')
        cantidades_pt = request.POST.getlist('cantidad_pt[]')

        cliente = None
        if not cliente_id:
            errores.append('Selecciona un cliente.')
        else:
            try:
                cliente = Cliente.objects.get(pk=cliente_id)
            except Cliente.DoesNotExist:
                errores.append('Cliente no encontrado.')

        ubicacion_pt = None
        if not ub_pt_id:
            errores.append('Selecciona la ubicación de origen.')
        else:
            try:
                ubicacion_pt = Ubicacion.objects.get(pk=ub_pt_id)
            except Ubicacion.DoesNotExist:
                errores.append('Ubicación no encontrada.')

        items_producto_map = {str(it.pk): it for it in items_producto}
        for i, (item_id, cant_str) in enumerate(zip(item_ids_pt, cantidades_pt), 1):
            cant_str = cant_str.strip()
            if not item_id and not cant_str:
                continue
            if not item_id:
                errores.append(f'Fila {i}: selecciona un producto.')
                continue
            if not cant_str:
                errores.append(f'Fila {i}: ingresa una cantidad.')
                continue
            try:
                cantidad = Decimal(cant_str)
                if cantidad <= 0:
                    raise ValueError
            except (ValueError, Exception):
                errores.append(f'Fila {i}: cantidad inválida.')
                continue
            it = items_producto_map.get(str(item_id))
            if not it:
                errores.append(f'Fila {i}: producto no encontrado.')
                continue
            filas_validas.append((it, cantidad, None, True, None))

        if not filas_validas and not errores:
            errores.append('Agrega al menos un producto con cantidad.')

        if errores:
            for e in errores:
                messages.error(request, e)
            filas_pt_previas = [
                {'item_id': iid, 'cantidad': cant}
                for iid, cant in zip(item_ids_pt, cantidades_pt)
                if iid or cant.strip()
            ]
            return render(request, 'movimientos/salida.html',
                          _ctx({'tab_inicial': 'producto_terminado',
                                'motivo_previo': motivo,
                                'fecha_mov_previo': fecha_mov_str,
                                'cliente_previo': cliente_id,
                                'ub_pt_previo': ub_pt_id,
                                'filas_pt_previas_json': _json_safe(filas_pt_previas),
                                'filas_previas_json': '[]'}))

        # Todo OK → guardar
        with transaction.atomic():
            mov = MovimientoInventario.objects.create(
                tipo_movimiento='salida',
                tipo_salida='producto_terminado',
                motivo=motivo,
                fecha_movimiento=fecha_movimiento,
                usuario=request.user,
                cliente=cliente,
            )
            pendientes_creados = []
            for it, cantidad, _, _pendiente, _maq in filas_validas:
                # Calcular si habrá stock negativo
                stock_ub = Stock.objects.filter(item=it, ubicacion=ubicacion_pt).first()
                stock_actual = stock_ub.cantidad_actual if stock_ub else Decimal('0')
                pendiente = stock_actual < cantidad

                det = DetalleMovimiento.objects.create(
                    movimiento=mov,
                    item=it,
                    cantidad=cantidad,
                    ubicacion_origen=ubicacion_pt,
                    cliente=cliente,
                    pendiente_conciliacion=pendiente,
                )
                _aplicar_efecto_detalle(det)
                if pendiente:
                    pendientes_creados.append(det)
                send_event('movement_created', {
                    'tipo': 'salida', 'item': it.nombre, 'codigo': it.codigo,
                    'cantidad': str(cantidad), 'ubicacion': ubicacion_pt.nombre,
                    'cliente': cliente.nombre if cliente else None,
                    'pendiente_conciliacion': pendiente,
                    'usuario': request.user.username,
                })
                notify_stock(it, movimiento='salida', usuario=request.user.username)

            # Notificar pendientes
            for det in pendientes_creados:
                send_event('salida_pendiente_conciliacion', {
                    'movimiento_pk': mov.pk,
                    'item': det.item.nombre, 'codigo': det.item.codigo,
                    'cantidad': str(det.cantidad),
                    'ubicacion': ubicacion_pt.nombre,
                    'cliente': cliente.nombre,
                    'usuario': request.user.username,
                })

        n_pendientes = len(pendientes_creados)
        msg = f'Movimiento #{mov.pk} registrado con {len(filas_validas)} ítem(s).'
        if n_pendientes:
            msg += f' ⚠️ {n_pendientes} línea(s) con stock insuficiente marcada(s) como pendiente(s) de conciliación.'
        messages.success(request, msg)
        return redirect('movimiento_detalle', pk=mov.pk)

    # ── TABS: Repuestos / Consumibles / Otros (filas dinámicas) ───────────────
    item_ids      = request.POST.getlist('item[]')
    cantidades    = request.POST.getlist('cantidad[]')
    ubicacion_ids = request.POST.getlist('ubicacion_origen[]')
    maquina_ids   = request.POST.getlist('maquina[]')

    all_items_qs = {str(it.pk): it for it in
                    Item.objects.filter(activo=True)
                    .exclude(tipo='producto')}

    for i, (item_id, cant_str, ub_id, maq_id) in enumerate(
        zip(item_ids, cantidades, ubicacion_ids, maquina_ids), 1
    ):
        cant_str = cant_str.strip()
        if not item_id and not cant_str:
            continue
        if not item_id:
            errores.append(f'Fila {i}: selecciona un ítem.')
            continue
        if not cant_str:
            errores.append(f'Fila {i}: ingresa una cantidad.')
            continue
        try:
            cantidad = Decimal(cant_str)
            if cantidad <= 0:
                raise ValueError
        except (ValueError, Exception):
            errores.append(f'Fila {i}: cantidad inválida.')
            continue

        item = all_items_qs.get(str(item_id))
        if not item:
            errores.append(f'Fila {i}: ítem no encontrado.')
            continue

        if not ub_id:
            errores.append(f'Fila {i} ({item.nombre}): selecciona ubicación de origen.')
            continue
        try:
            ubicacion = Ubicacion.objects.get(pk=ub_id)
        except Ubicacion.DoesNotExist:
            errores.append(f'Fila {i}: ubicación no encontrada.')
            continue

        # Bloquear si stock insuficiente (repuestos/consumibles/otros)
        try:
            stock = Stock.objects.get(item=item, ubicacion=ubicacion)
            if stock.cantidad_actual < cantidad:
                errores.append(
                    f'Fila {i} ({item.nombre}): stock insuficiente. '
                    f'Disponible: {stock.cantidad_actual} {item.unidad_medida}.'
                )
                continue
        except Stock.DoesNotExist:
            errores.append(f'Fila {i} ({item.nombre}): no hay stock en esa ubicación.')
            continue

        maquina = None
        if maq_id:
            try:
                maquina = Maquina.objects.get(pk=maq_id)
            except Maquina.DoesNotExist:
                pass

        # Repuesto requiere máquina
        if item.tipo == 'repuesto' and not maquina:
            errores.append(f'Fila {i} ({item.nombre}): selecciona una máquina.')
            continue

        filas_validas.append((item, cantidad, ubicacion, False, maquina))

    if not filas_validas and not errores:
        errores.append('Agrega al menos un ítem con cantidad.')

    if errores:
        for e in errores:
            messages.error(request, e)
        filas_previas = [
            {'item_id': iid, 'cantidad': cant, 'ub_id': ub, 'maq_id': maq}
            for iid, cant, ub, maq
            in zip(item_ids, cantidades, ubicacion_ids, maquina_ids)
            if iid or cant.strip()
        ]
        return render(request, 'movimientos/salida.html',
                      _ctx({'tab_inicial': tipo_salida,
                            'motivo_previo': motivo,
                            'fecha_mov_previo': fecha_mov_str,
                            'filas_previas_json': _json_safe(filas_previas)}))

    with transaction.atomic():
        mov = MovimientoInventario.objects.create(
            tipo_movimiento='salida',
            tipo_salida=tipo_salida,
            motivo=motivo,
            fecha_movimiento=fecha_movimiento,
            usuario=request.user,
        )
        for it, cantidad, ubicacion, _pendiente, maquina in filas_validas:
            det = DetalleMovimiento.objects.create(
                movimiento=mov,
                item=it,
                cantidad=cantidad,
                ubicacion_origen=ubicacion,
                maquina=maquina,
            )
            _aplicar_efecto_detalle(det)
            send_event('movement_created', {
                'tipo': 'salida', 'item': it.nombre, 'codigo': it.codigo,
                'cantidad': str(cantidad), 'ubicacion': ubicacion.nombre,
                'usuario': request.user.username,
            })
            notify_stock(it, movimiento='salida', usuario=request.user.username)

    messages.success(
        request,
        f'Movimiento #{mov.pk} registrado con {len(filas_validas)} ítem(s).'
    )
    return redirect('movimiento_detalle', pk=mov.pk)


@login_required
@permission_required(_perm('registrar_entrada'), raise_exception=True)
def movimiento_transferencia(request):
    if request.method == 'POST':
        form = MovimientoTransferenciaForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                mov = MovimientoInventario.objects.create(
                    tipo_movimiento='transferencia',
                    motivo=form.cleaned_data.get('motivo', ''),
                    usuario=request.user,
                )
                det = DetalleMovimiento.objects.create(
                    movimiento=mov,
                    item=form.cleaned_data['item'],
                    cantidad=form.cleaned_data['cantidad'],
                    ubicacion_origen=form.cleaned_data['ubicacion_origen'],
                    ubicacion_destino=form.cleaned_data['ubicacion_destino'],
                )
                _aplicar_efecto_detalle(det)
            messages.success(request, f'Transferencia registrada (Movimiento #{mov.pk}).')
            return redirect('movimiento_detalle', pk=mov.pk)
    else:
        form = MovimientoTransferenciaForm()

    return render(request, 'movimientos/transferencia.html', {'form': form})


# ─── DETALLE / GESTIÓN DE MOVIMIENTOS ────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def movimiento_detalle(request, pk):
    """Vista de detalle de un Movimiento (cabecera + líneas)."""
    mov = get_object_or_404(
        MovimientoInventario.objects.select_related(
            'usuario', 'usuario_edicion', 'usuario_anulacion', 'usuario_eliminacion'
        ).prefetch_related(
            'detalles__item', 'detalles__ubicacion_origen',
            'detalles__ubicacion_destino', 'detalles__cliente', 'detalles__maquina',
        ),
        pk=pk,
    )
    return render(request, 'movimientos/detalle.html', {'mov': mov})


@login_required
@permission_required(_perm('editar_movimiento'), raise_exception=True)
def movimiento_editar(request, pk):
    """
    Edita un movimiento existente.
    Permite cambiar la fecha, motivo y las cantidades/ubicaciones por línea.
    Flujo: revertir todos los detalles → guardar cambios → re-aplicar detalles.
    Todo en transaction.atomic().
    """
    mov = get_object_or_404(
        MovimientoInventario.objects.prefetch_related(
            'detalles__item', 'detalles__ubicacion_origen',
            'detalles__ubicacion_destino', 'detalles__cliente', 'detalles__maquina',
        ).select_related('usuario'),
        pk=pk,
    )

    if not _movimiento_editable(mov):
        messages.error(request, 'Este movimiento está anulado o eliminado y no puede editarse.')
        return redirect('movimiento_detalle', pk=pk)

    ubicaciones = Ubicacion.objects.all()
    clientes    = Cliente.objects.filter(activo=True).order_by('nombre')
    maquinas    = Maquina.objects.filter(activo=True).order_by('nombre')

    if request.method == 'POST':
        motivo_edicion = request.POST.get('motivo_edicion', '').strip()
        nuevo_motivo   = request.POST.get('motivo', mov.motivo)
        nueva_fecha_str = request.POST.get('fecha_movimiento', '').strip()

        if not motivo_edicion:
            messages.error(request, 'El motivo de edición es obligatorio.')
        else:
            nueva_fecha = mov.fecha_movimiento
            if nueva_fecha_str:
                from django.utils.dateparse import parse_datetime
                parsed = parse_datetime(nueva_fecha_str)
                if parsed:
                    nueva_fecha = timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed

            # Recoger nuevas cantidades por detalle
            det_cantidades   = request.POST.getlist('det_cantidad[]')
            det_ids          = request.POST.getlist('det_id[]')
            det_ub_origen    = request.POST.getlist('det_ub_origen[]')
            det_ub_destino   = request.POST.getlist('det_ub_destino[]')

            errores = []
            nuevos_valores = []
            detalles = list(mov.detalles.all())

            for i, det in enumerate(detalles):
                cant_str = det_cantidades[i].strip() if i < len(det_cantidades) else ''
                try:
                    nueva_cant = Decimal(cant_str)
                    if nueva_cant <= 0:
                        raise ValueError
                except Exception:
                    errores.append(f'Línea {i+1} ({det.item.nombre}): cantidad inválida.')
                    continue

                ub_or_id  = det_ub_origen[i]  if i < len(det_ub_origen)  else ''
                ub_dst_id = det_ub_destino[i] if i < len(det_ub_destino) else ''
                nueva_ub_origen  = Ubicacion.objects.filter(pk=ub_or_id).first()  if ub_or_id  else None
                nueva_ub_destino = Ubicacion.objects.filter(pk=ub_dst_id).first() if ub_dst_id else None
                nuevos_valores.append((det, nueva_cant, nueva_ub_origen, nueva_ub_destino))

            if errores:
                for e in errores:
                    messages.error(request, e)
            else:
                with transaction.atomic():
                    # 1. Revertir todos los detalles originales
                    _revertir_todos_los_detalles(mov)
                    # 2. Actualizar cabecera
                    mov.motivo           = nuevo_motivo
                    mov.fecha_movimiento = nueva_fecha
                    mov.editado          = True
                    mov.fecha_edicion    = timezone.now()
                    mov.usuario_edicion  = request.user
                    mov.motivo_edicion   = motivo_edicion
                    mov.save(update_fields=[
                        'motivo', 'fecha_movimiento',
                        'editado', 'fecha_edicion', 'usuario_edicion', 'motivo_edicion',
                    ])
                    # 3. Actualizar detalles y re-aplicar stock
                    for det, nueva_cant, nueva_ub_origen, nueva_ub_destino in nuevos_valores:
                        det.cantidad          = nueva_cant
                        det.ubicacion_origen  = nueva_ub_origen
                        det.ubicacion_destino = nueva_ub_destino
                        det.save(update_fields=['cantidad', 'ubicacion_origen', 'ubicacion_destino'])
                        _aplicar_efecto_detalle(det)
                        notify_stock(det.item, movimiento='edicion', usuario=request.user.username)

                security_log.info(
                    'Movimiento #%s editado por %s — %s',
                    mov.pk, request.user.username, motivo_edicion,
                )
                messages.success(request, f'Movimiento #{mov.pk} editado correctamente.')
                return redirect('movimiento_detalle', pk=mov.pk)

    return render(request, 'movimientos/editar.html', {
        'mov': mov,
        'ubicaciones': ubicaciones,
        'clientes': clientes,
        'maquinas': maquinas,
    })


@login_required
@permission_required(_perm('anular_movimiento'), raise_exception=True)
def movimiento_anular(request, pk):
    """
    Anula un movimiento completo: revierte el stock de TODOS sus detalles y
    marca la cabecera como anulada. No lo borra — queda visible con badge «Anulado».
    """
    mov = get_object_or_404(
        MovimientoInventario.objects.prefetch_related(
            'detalles__item', 'detalles__ubicacion_origen', 'detalles__ubicacion_destino'
        ).select_related('usuario'),
        pk=pk,
    )

    if mov.anulado:
        messages.warning(request, 'Este movimiento ya estaba anulado.')
        return redirect('movimiento_detalle', pk=pk)
    if mov.eliminado:
        messages.error(request, 'No se puede anular un movimiento eliminado.')
        return redirect('movimiento_detalle', pk=pk)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_anulacion', '').strip()
        if not motivo:
            messages.error(request, 'El motivo de anulación es obligatorio.')
        else:
            with transaction.atomic():
                _revertir_todos_los_detalles(mov)
                mov.anulado           = True
                mov.fecha_anulacion   = timezone.now()
                mov.usuario_anulacion = request.user
                mov.motivo_anulacion  = motivo
                mov.save(update_fields=[
                    'anulado', 'fecha_anulacion', 'usuario_anulacion', 'motivo_anulacion'
                ])
                for det in mov.detalles.select_related('item').all():
                    notify_stock(det.item, movimiento='anulacion', usuario=request.user.username)

            security_log.info(
                'Movimiento #%s ANULADO por %s — %s',
                mov.pk, request.user.username, motivo,
            )
            messages.success(
                request,
                f'Movimiento #{mov.pk} anulado. Stock revertido en {mov.detalles.count()} ítem(s).'
            )
            return redirect('movimiento_lista')

    return render(request, 'movimientos/anular.html', {'mov': mov})


@login_required
@permission_required(_perm('eliminar_movimiento'), raise_exception=True)
def movimiento_eliminar(request, pk):
    """
    Eliminación lógica de un movimiento.
    Si no estaba anulado, revierte el stock de todos sus detalles.
    Requiere doble confirmación y motivo.
    """
    mov = get_object_or_404(
        MovimientoInventario.objects.prefetch_related(
            'detalles__item', 'detalles__ubicacion_origen', 'detalles__ubicacion_destino'
        ).select_related('usuario'),
        pk=pk,
    )

    if mov.eliminado:
        messages.warning(request, 'Este movimiento ya estaba eliminado.')
        return redirect('movimiento_lista')

    if mov.anulado and not request.user.is_superuser:
        messages.error(request, 'Solo un superusuario puede eliminar un movimiento ya anulado.')
        return redirect('movimiento_detalle', pk=pk)

    if request.method == 'POST':
        confirmacion = request.POST.get('confirmacion', '').strip()
        motivo = request.POST.get('motivo_eliminacion', '').strip()

        if confirmacion != 'ELIMINAR':
            messages.error(request, 'Escribe ELIMINAR en el campo de confirmación.')
        elif not motivo:
            messages.error(request, 'El motivo de eliminación es obligatorio.')
        else:
            with transaction.atomic():
                # Solo revertir si no estaba ya anulado (el anulado ya lo revirtió)
                if not mov.anulado:
                    _revertir_todos_los_detalles(mov)
                mov.eliminado           = True
                mov.fecha_eliminacion   = timezone.now()
                mov.usuario_eliminacion = request.user
                mov.motivo_eliminacion  = motivo
                mov.save(update_fields=[
                    'eliminado', 'fecha_eliminacion', 'usuario_eliminacion', 'motivo_eliminacion'
                ])
                if not mov.anulado:
                    for det in mov.detalles.select_related('item').all():
                        notify_stock(det.item, movimiento='eliminacion',
                                     usuario=request.user.username)

            security_log.warning(
                'Movimiento #%s ELIMINADO por %s — %s',
                mov.pk, request.user.username, motivo,
            )
            messages.success(request, f'Movimiento #{mov.pk} eliminado lógicamente.')
            return redirect('movimiento_lista')

    return render(request, 'movimientos/eliminar.html', {'mov': mov})


# ─── CONTEOS ──────────────────────────────────────────────────────────────────

@login_required
def conteo_anular(request, pk):
    """
    Anulación lógica de un conteo: revierte en stock todos los ajustes de
    conciliación generados por este conteo, marca cada movimiento de ajuste
    como anulado, y finalmente marca el conteo como anulado.
    No elimina ningún registro.
    """
    if not (request.user.has_perm(_perm('anular_conteo')) or request.user.is_superuser):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied

    conteo = get_object_or_404(
        Conteo.objects.select_related('usuario'),
        pk=pk,
    )

    if conteo.anulado:
        messages.warning(request, 'Este conteo ya estaba anulado.')
        return redirect('conteo_detalle', pk=pk)

    # Movimientos de ajuste vinculados a este conteo (por motivo)
    ajustes_qs = (
        MovimientoInventario.objects
        .filter(
            tipo_movimiento='ajuste',
            motivo__contains=f'Conteo #{conteo.pk}',
            anulado=False,
            eliminado=False,
        )
        .prefetch_related(
            'detalles__item',
            'detalles__ubicacion_origen',
            'detalles__ubicacion_destino',
        )
    )
    ajustes = list(ajustes_qs)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_anulacion', '').strip()
        if not motivo:
            messages.error(request, 'El motivo de anulación es obligatorio.')
        else:
            ahora = timezone.now()
            with transaction.atomic():
                for mov_ajuste in ajustes:
                    _revertir_todos_los_detalles(mov_ajuste)
                    mov_ajuste.anulado           = True
                    mov_ajuste.fecha_anulacion   = ahora
                    mov_ajuste.usuario_anulacion = request.user
                    mov_ajuste.motivo_anulacion  = f'Anulación de Conteo #{conteo.pk} — {motivo}'
                    mov_ajuste.save(update_fields=[
                        'anulado', 'fecha_anulacion', 'usuario_anulacion', 'motivo_anulacion',
                    ])
                    for det in mov_ajuste.detalles.select_related('item').all():
                        notify_stock(det.item, movimiento='anulacion', usuario=request.user.username)

                conteo.anulado           = True
                conteo.fecha_anulacion   = ahora
                conteo.usuario_anulacion = request.user
                conteo.motivo_anulacion  = motivo
                conteo.save(update_fields=[
                    'anulado', 'fecha_anulacion', 'usuario_anulacion', 'motivo_anulacion',
                ])

            security_log.info(
                'Conteo #%s ANULADO por %s — %s ajuste(s) revertido(s) — %s',
                conteo.pk, request.user.username, len(ajustes), motivo,
            )
            messages.success(
                request,
                f'Conteo #{conteo.pk} anulado. {len(ajustes)} ajuste(s) de conciliación revertido(s).'
            )
            return redirect('conteo_lista')

    return render(request, 'conteos/anular.html', {
        'conteo': conteo,
        'ajustes': ajustes,
    })


@login_required
@permission_required(_perm('registrar_conteo'), raise_exception=True)
@_timed_view('conteo_lista')
def conteo_lista(request):
    qs = (
        Conteo.objects
        .select_related('usuario')
        .annotate(num_detalles=Count('detalles'))
        .order_by('-fecha', 'turno')
    )
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'conteos/lista.html', {'conteos': page_obj, 'page_obj': page_obj})


@login_required
@permission_required(_perm('registrar_conteo'), raise_exception=True)
def conteo_nuevo(request):
    hoy = date.today()
    ubicaciones = list(Ubicacion.objects.all())

    stocks_map = {}
    for s in Stock.objects.select_related('item', 'ubicacion').filter(item__activo=True):
        stocks_map[(s.item_id, s.ubicacion_id)] = s.cantidad_actual

    stocks_totales = {}
    for s in Stock.objects.filter(item__activo=True).values('item_id').annotate(t=Sum('cantidad_actual')):
        stocks_totales[s['item_id']] = s['t'] or Decimal('0')

    all_items = list(
        Item.objects.filter(activo=True)
        .select_related('categoria')
        .order_by('orden', 'nombre')
    )

    def _clasificar(item):
        cat = (item.categoria.nombre if item.categoria else '').lower()
        nom = item.nombre.lower()
        if item.tipo == 'producto' and ('camiseta' in cat or 'camiseta' in nom):
            return 'camiseta'
        if item.tipo == 'consumible' and ('pigment' in cat or 'pigment' in nom):
            return 'pigmentos'
        if item.tipo == 'producto' and ('lisa' in cat or 'lisa' in nom):
            return 'lisa'
        return 'otros'

    def _build_item_dict(item):
        stocks_by_ub = {
            str(upk): str(qty)
            for (ipk, upk), qty in stocks_map.items()
            if ipk == item.pk
        }
        best_ub = max(stocks_by_ub, key=lambda k: Decimal(stocks_by_ub[k]), default=None)
        return {
            'pk': item.pk,
            'nombre': item.nombre,
            'codigo': item.codigo,
            'categoria': item.categoria.nombre if item.categoria else '',
            'unidad': item.unidad_medida,
            'stock_total': str(stocks_totales.get(item.pk, Decimal('0'))),
            'default_ub': int(best_ub) if best_ub else (ubicaciones[0].pk if ubicaciones else None),
            'stocks_by_ub': stocks_by_ub,
        }

    items_por_tipo = {'camiseta': [], 'pigmentos': [], 'lisa': [], 'otros': []}
    for item in all_items:
        items_por_tipo[_clasificar(item)].append(_build_item_dict(item))

    items_por_tipo_json = _json_safe(items_por_tipo)
    all_items_json = _json_safe([_build_item_dict(item) for item in all_items])
    ubicaciones_json = _json_safe([
        {'pk': u.pk, 'nombre': u.nombre, 'tipo': u.get_tipo_display()}
        for u in ubicaciones
    ])

    if request.method == 'POST':
        form = ConteoForm(request.POST)

        item_ids = request.POST.getlist('item[]')
        ubicacion_ids = request.POST.getlist('ubicacion[]')
        cantidades = request.POST.getlist('cantidad_contada[]')

        # Lista ordenada: [{item_id, ub_id, cant}] para todos los ítems con cantidad
        filas_previas = [
            {'item_id': iid, 'ub_id': uid, 'cant': cant}
            for iid, uid, cant in zip(item_ids, ubicacion_ids, cantidades)
            if cant.strip()
        ]
        filas_previas_json = _json_safe(filas_previas)
        tipo_conteo_previo = request.POST.get('tipo_conteo', 'camiseta')

        def _render_error(f):
            return render(request, 'conteos/form.html', {
                'form': f,
                'items_por_tipo_json': items_por_tipo_json,
                'all_items_json': all_items_json,
                'ubicaciones_json': ubicaciones_json,
                'hoy': hoy,
                'filas_previas_json': filas_previas_json,
                'tipo_conteo_inicial': tipo_conteo_previo,
                'tipos_conteo_fijos': ['camiseta', 'pigmentos', 'lisa'],
            })

        if not form.is_valid():
            return _render_error(form)

        fecha = form.cleaned_data['fecha']
        turno = form.cleaned_data['turno']
        tipo_conteo = form.cleaned_data['tipo_conteo']

        if Conteo.objects.filter(fecha=fecha, turno=turno, tipo_conteo=tipo_conteo, anulado=False).exists():
            label_tipo = dict(Conteo.TIPO_CONTEO_CHOICES).get(tipo_conteo, tipo_conteo)
            label_turno = dict(Conteo.TURNO_CHOICES).get(turno, turno)
            messages.error(
                request,
                f'Ya existe un conteo de {label_tipo} - {label_turno} para {fecha}.'
            )
            return _render_error(form)

        errores = []
        filas = []

        for i, (item_id, ub_id, cant_str) in enumerate(
            zip(item_ids, ubicacion_ids, cantidades), 1
        ):
            cant_str = cant_str.strip()
            if not cant_str:
                continue
            if not item_id:
                errores.append(f'Fila {i}: ítem inválido.')
                continue
            if not ub_id:
                errores.append(f'Fila {i}: selecciona una ubicación.')
                continue
            try:
                cantidad_contada = Decimal(cant_str)
            except Exception:
                errores.append(f'Fila {i}: cantidad inválida.')
                continue
            if cantidad_contada < 0:
                errores.append(f'Fila {i}: la cantidad no puede ser negativa.')
                continue
            try:
                item = Item.objects.get(pk=item_id, activo=True)
            except Item.DoesNotExist:
                errores.append(f'Fila {i}: ítem no encontrado.')
                continue
            try:
                ubicacion = Ubicacion.objects.get(pk=ub_id)
            except Ubicacion.DoesNotExist:
                errores.append(f'Fila {i}: ubicación no encontrada.')
                continue

            cantidad_sistema = stocks_map.get((item.pk, ubicacion.pk), Decimal('0'))
            filas.append((item, ubicacion, cantidad_contada, cantidad_sistema))

        if errores:
            for e in errores:
                messages.error(request, e)
            return _render_error(form)

        if not filas:
            messages.error(request, 'Ingresá al menos una cantidad en el conteo.')
            return _render_error(form)

        with transaction.atomic():
            conteo = form.save(commit=False)
            conteo.usuario = request.user
            conteo.save()
            for item, ubicacion, cantidad_contada, cantidad_sistema in filas:
                ConteoDetalle.objects.create(
                    conteo=conteo,
                    item=item,
                    ubicacion=ubicacion,
                    cantidad_contada=cantidad_contada,
                    cantidad_sistema_al_conteo=cantidad_sistema,
                )

        label_tipo = dict(Conteo.TIPO_CONTEO_CHOICES).get(conteo.tipo_conteo, conteo.tipo_conteo)
        messages.success(
            request,
            f'Conteo {label_tipo} - {conteo.get_turno_display()} registrado con {len(filas)} ítem(s). '
            f'Revisá la conciliación para calcular diferencias.'
        )
        return redirect('conteo_conciliar', pk=conteo.pk)

    form = ConteoForm(initial={
        'fecha': hoy,
        'tipo_conteo': 'camiseta',
        'fecha_hora_conteo': timezone.localtime(timezone.now()).strftime('%Y-%m-%dT%H:%M'),
    })
    return render(request, 'conteos/form.html', {
        'form': form,
        'items_por_tipo_json': items_por_tipo_json,
        'all_items_json': all_items_json,
        'ubicaciones_json': ubicaciones_json,
        'hoy': hoy,
        'filas_previas_json': '[]',
        'tipo_conteo_inicial': 'camiseta',
        'tipos_conteo_fijos': ['camiseta', 'pigmentos', 'lisa'],
    })


@login_required
@permission_required(_perm('registrar_conteo'), raise_exception=True)
def conteo_detalle(request, pk):
    conteo = get_object_or_404(Conteo, pk=pk)
    detalles = conteo.detalles.select_related('item', 'ubicacion').order_by('item__orden', 'item__nombre')
    total_contado = detalles.aggregate(t=Sum('cantidad_contada'))['t'] or 0
    total_dif_original = detalles.aggregate(t=Sum('diferencia_original'))['t'] or 0

    context = {
        'conteo': conteo,
        'detalles': detalles,
        'total_contado': total_contado,
        'total_dif_original': total_dif_original,
    }
    return render(request, 'conteos/detalle.html', context)


@login_required
@permission_required(_perm('aplicar_conciliacion'), raise_exception=True)
def conteo_conciliar(request, pk):
    conteo = get_object_or_404(Conteo, pk=pk)
    if conteo.anulado:
        messages.error(request, 'Este conteo está anulado y no puede conciliarse.')
        return redirect('conteo_detalle', pk=pk)
    detalles = conteo.detalles.select_related('item', 'ubicacion').order_by('item__orden', 'item__nombre')

    plan = []
    with transaction.atomic():
        for detalle in detalles:
            # Stock teórico al momento del conteo usando fecha_movimiento como
            # timestamp oficial. No depende de cuándo se registró el movimiento.
            stock_teorico = _stock_en_momento(
                detalle.item, detalle.ubicacion, conteo.fecha_hora_conteo
            )
            diferencia_final = detalle.cantidad_contada - stock_teorico

            # Persistir diferencia_final si cambió
            if detalle.diferencia_final != diferencia_final:
                detalle.diferencia_final = diferencia_final
                ConteoDetalle.objects.filter(pk=detalle.pk).update(
                    diferencia_final=diferencia_final
                )

            # Movimientos registrados DESPUÉS del conteo pero con fecha_movimiento
            # ANTES del conteo. Se muestran para transparencia: ya están
            # correctamente incluidos en stock_teorico (no son "atrasados" — son
            # movimientos reales anteriores al conteo, solo ingresados tarde).
            movs_tardios = (
                DetalleMovimiento.objects
                .filter(
                    item=detalle.item,
                    movimiento__anulado=False,
                    movimiento__eliminado=False,
                    movimiento__fecha__gt=conteo.fecha_hora_conteo,
                    movimiento__fecha_movimiento__lte=conteo.fecha_hora_conteo,
                )
                .filter(
                    Q(ubicacion_destino=detalle.ubicacion)
                    | Q(ubicacion_origen=detalle.ubicacion)
                )
                .select_related(
                    'movimiento', 'movimiento__usuario',
                    'ubicacion_origen', 'ubicacion_destino',
                )
                .order_by('movimiento__fecha_movimiento')
            )

            plan.append({
                'detalle': detalle,
                'stock_teorico': stock_teorico,
                'diferencia_final': diferencia_final,
                'movs_tardios': movs_tardios,   # solo informativos
            })

    return render(request, 'conteos/conciliar.html', {
        'conteo': conteo,
        'plan': plan,
    })


@login_required
@permission_required(_perm('aplicar_conciliacion'), raise_exception=True)
def conteo_ajustar_detalle(request, pk, det_pk):
    if request.method != 'POST':
        return redirect('conteo_conciliar', pk=pk)

    conteo = get_object_or_404(Conteo, pk=pk)
    if conteo.anulado:
        messages.error(request, 'Este conteo está anulado.')
        return redirect('conteo_detalle', pk=pk)
    estado_antes = conteo.estado
    detalle = get_object_or_404(ConteoDetalle, pk=det_pk, conteo=conteo)

    if detalle.ajuste_aplicado:
        messages.warning(request, 'Este ajuste ya fue aplicado.')
        return redirect('conteo_conciliar', pk=pk)

    if detalle.diferencia_final is None:
        messages.error(request, 'Primero calculá la diferencia final en la pantalla de conciliación.')
        return redirect('conteo_conciliar', pk=pk)

    if detalle.diferencia_final == 0:
        messages.info(request, f'{detalle.item.nombre}: no hay diferencia que ajustar.')
        return redirect('conteo_conciliar', pk=pk)

    with transaction.atomic():
        mov_ajuste = MovimientoInventario.objects.create(
            tipo_movimiento='ajuste',
            motivo=f'Ajuste por conciliación — Conteo #{conteo.pk} ({conteo.get_turno_display()} {conteo.fecha})',
            usuario=request.user,
        )
        det_ajuste = DetalleMovimiento.objects.create(
            movimiento=mov_ajuste,
            item=detalle.item,
            cantidad=detalle.diferencia_final,
            ubicacion_destino=detalle.ubicacion,
        )
        _aplicar_efecto_detalle(det_ajuste)
        _cerrar_pendientes_conciliacion(detalle.item, detalle.ubicacion)
        ConteoDetalle.objects.filter(pk=detalle.pk).update(ajuste_aplicado=True)
        conteo.refresh_from_db()
        conteo.actualizar_estado()
        # Si este ajuste cerró la conciliación, reenviar inventario camiseta (1 vez)
        _notificar_si_conciliacion_completa(conteo, estado_antes, request.user.username)

    send_event('count_difference', {
        'conteo_id': conteo.pk, 'item': detalle.item.nombre, 'codigo': detalle.item.codigo,
        'diferencia': str(detalle.diferencia_final), 'ubicacion': detalle.ubicacion.nombre,
        'usuario': request.user.username,
    })
    notify_stock(detalle.item, movimiento='ajuste', usuario=request.user.username)
    messages.success(request, f'Ajuste aplicado: {detalle.item.nombre} ({detalle.diferencia_final:+g} {detalle.item.unidad_medida}).')
    return redirect('conteo_conciliar', pk=pk)


@login_required
@permission_required(_perm('aplicar_conciliacion'), raise_exception=True)
def conteo_ajustar_todos(request, pk):
    if request.method != 'POST':
        return redirect('conteo_conciliar', pk=pk)

    conteo = get_object_or_404(Conteo, pk=pk)
    if conteo.anulado:
        messages.error(request, 'Este conteo está anulado.')
        return redirect('conteo_detalle', pk=pk)
    estado_antes = conteo.estado
    detalles = conteo.detalles.filter(
        ajuste_aplicado=False,
        diferencia_final__isnull=False,
    ).exclude(diferencia_final=0).select_related('item', 'ubicacion')

    if not detalles.exists():
        messages.info(request, 'No hay ajustes pendientes con diferencia.')
        return redirect('conteo_conciliar', pk=pk)

    count = 0
    with transaction.atomic():
        for detalle in detalles:
            mov_ajuste = MovimientoInventario.objects.create(
                tipo_movimiento='ajuste',
                motivo=f'Ajuste por conciliación — Conteo #{conteo.pk} ({conteo.get_turno_display()} {conteo.fecha})',
                usuario=request.user,
            )
            det_ajuste = DetalleMovimiento.objects.create(
                movimiento=mov_ajuste,
                item=detalle.item,
                cantidad=detalle.diferencia_final,
                ubicacion_destino=detalle.ubicacion,
            )
            _aplicar_efecto_detalle(det_ajuste)
            _cerrar_pendientes_conciliacion(detalle.item, detalle.ubicacion)
            count += 1
        ConteoDetalle.objects.filter(
            conteo=conteo, ajuste_aplicado=False,
            diferencia_final__isnull=False
        ).exclude(diferencia_final=0).update(ajuste_aplicado=True)
        conteo.refresh_from_db()
        conteo.actualizar_estado()
        # Si quedó conciliado, reenviar inventario camiseta (1 vez)
        _notificar_si_conciliacion_completa(conteo, estado_antes, request.user.username)

    send_event('count_difference', {
        'conteo_id': conteo.pk, 'ajustes_aplicados': count,
        'usuario': request.user.username,
    })
    messages.success(request, f'{count} ajuste(s) aplicado(s) exitosamente.')
    return redirect('conteo_conciliar', pk=pk)


@login_required
@permission_required(_perm('aplicar_conciliacion'), raise_exception=True)
def conteo_marcar_conciliado(request, pk):
    if request.method != 'POST':
        return redirect('conteo_conciliar', pk=pk)

    conteo = get_object_or_404(Conteo, pk=pk)
    if conteo.anulado:
        messages.error(request, 'Este conteo está anulado.')
        return redirect('conteo_detalle', pk=pk)
    estado_antes = conteo.estado

    # Verificar que no queden diferencias sin ajustar
    pendientes = conteo.detalles.filter(
        ajuste_aplicado=False,
        diferencia_final__isnull=False,
    ).exclude(diferencia_final=0).count()

    sin_calcular = conteo.detalles.filter(diferencia_final__isnull=True).count()

    if sin_calcular > 0:
        messages.warning(request, f'Hay {sin_calcular} línea(s) sin diferencia calculada. Abrí la conciliación primero.')
        return redirect('conteo_conciliar', pk=pk)

    if pendientes > 0:
        messages.warning(request, f'Hay {pendientes} ajuste(s) pendiente(s) con diferencia. Aplicalos o ignoralos antes de cerrar.')
        return redirect('conteo_conciliar', pk=pk)

    with transaction.atomic():
        conteo.estado = 'conciliado'
        conteo.save(update_fields=['estado'])
        # Reenviar inventario camiseta (1 vez) si recién ahora quedó conciliado
        _notificar_si_conciliacion_completa(conteo, estado_antes, request.user.username)

    messages.success(request, 'Conteo marcado como conciliado.')
    return redirect('conteo_detalle', pk=pk)


# ─── MÁQUINAS ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def maquina_lista(request):
    maquinas = Maquina.objects.order_by('nombre')
    return render(request, 'maquinas/lista.html', {'maquinas': maquinas})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_crear(request):
    if request.method == 'POST':
        form = MaquinaForm(request.POST)
        if form.is_valid():
            m = form.save()
            messages.success(request, f'Máquina "{m.nombre}" creada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm()
    return render(request, 'maquinas/form.html', {'form': form, 'titulo': 'Nueva Máquina'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def maquina_editar(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    if request.method == 'POST':
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, f'Máquina "{maquina.nombre}" actualizada.')
            return redirect('maquina_lista')
    else:
        form = MaquinaForm(instance=maquina)
    return render(request, 'maquinas/form.html', {
        'form': form, 'titulo': f'Editar: {maquina.nombre}', 'maquina': maquina
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def maquina_toggle_activo(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    maquina.activo = not maquina.activo
    maquina.save()
    messages.success(request, f'Máquina "{maquina.nombre}" {"activada" if maquina.activo else "desactivada"}.')
    return redirect('maquina_lista')


# ─── CLIENTES ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
def cliente_lista(request):
    q = request.GET.get('q', '')
    clientes = Cliente.objects.order_by('nombre')
    if q:
        clientes = clientes.filter(Q(nombre__icontains=q) | Q(rtn__icontains=q))
    return render(request, 'clientes/lista.html', {'clientes': clientes, 'q': q})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_crear(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            c = form.save()
            messages.success(request, f'Cliente "{c.nombre}" creado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm()
    return render(request, 'clientes/form.html', {'form': form, 'titulo': 'Nuevo Cliente'})


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
def cliente_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente "{cliente.nombre}" actualizado.')
            return redirect('cliente_lista')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/form.html', {
        'form': form, 'titulo': f'Editar: {cliente.nombre}', 'cliente': cliente
    })


@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def cliente_toggle_activo(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.activo = not cliente.activo
    cliente.save()
    messages.success(request, f'Cliente "{cliente.nombre}" {"activado" if cliente.activo else "desactivado"}.')
    return redirect('cliente_lista')


def _orden_operativo_producto(item):
    orden = [
        'bolsa camiseta grande',
        'bolsa camiseta mediana',
        'bolsa camiseta pequeña',
        'bolsa camiseta pequena',
        'bolsa camiseta grande negra',
        'bolsa camiseta mediana negra',
        'bolsa camiseta pequeña negra',
        'bolsa camiseta pequena negra',
    ]
    nombre = item.nombre.lower().strip()
    try:
        pos = orden.index(nombre)
    except ValueError:
        pos = 999
    return (pos, item.orden, item.nombre)


def _label_grupo_cliente(fecha, agrupar_por):
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if agrupar_por == 'dia':
        return fecha.strftime('%d/%m/%Y'), fecha
    if agrupar_por == 'semana':
        iso = fecha.isocalendar()
        return f'Semana {iso[1]}, {iso[0]}', (iso[0], iso[1])
    return f'{meses[fecha.month]} {fecha.year}', (fecha.year, fecha.month)


def _build_salidas_cliente_context(cliente, request):
    from datetime import datetime as _dt

    hoy = date.today()

    def _parse_date(key, fallback):
        try:
            return _dt.strptime(request.GET.get(key, ''), '%Y-%m-%d').date()
        except ValueError:
            return fallback

    fecha_fin = _parse_date('fecha_fin', hoy)
    fecha_inicio = _parse_date('fecha_inicio', fecha_fin - timedelta(days=30))
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio

    agrupar_por = request.GET.get('agrupar_por', 'dia')
    if agrupar_por not in ('dia', 'semana', 'mes'):
        agrupar_por = 'dia'

    producto_pk = request.GET.get('producto', '').strip()
    try:
        producto_id = int(producto_pk) if producto_pk else None
    except ValueError:
        producto_id = None

    productos = (
        Item.objects
        .filter(activo=True, tipo='producto')
        .select_related('categoria')
        .order_by('orden', 'nombre')
    )

    detalles = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__cliente=cliente,
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__date__gte=fecha_inicio,
            movimiento__fecha_movimiento__date__lte=fecha_fin,
        )
        .select_related('movimiento', 'movimiento__usuario', 'item', 'item__categoria')
        .order_by('movimiento__fecha_movimiento', 'movimiento_id', 'item__orden', 'item__nombre')
    )
    if producto_id:
        detalles = detalles.filter(item_id=producto_id)

    detalles = list(detalles)
    total_general = sum((d.cantidad for d in detalles), Decimal('0'))
    movimiento_ids = {d.movimiento_id for d in detalles}
    producto_ids = {d.item_id for d in detalles}

    por_producto_map = {}
    for det in detalles:
        data = por_producto_map.setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
            'unidad': det.item.unidad_medida,
        })
        data['cantidad'] += det.cantidad
    por_producto = sorted(por_producto_map.values(), key=lambda row: _orden_operativo_producto(row['item']))

    grupos_map = {}
    for det in detalles:
        fecha_local = timezone.localtime(det.movimiento.fecha_movimiento).date()
        label, key = _label_grupo_cliente(fecha_local, agrupar_por)
        grupo = grupos_map.setdefault(key, {
            'label': label,
            'fecha': fecha_local,
            'total': Decimal('0'),
            'items': {},
        })
        grupo['total'] += det.cantidad
        item_row = grupo['items'].setdefault(det.item_id, {
            'item': det.item,
            'cantidad': Decimal('0'),
        })
        item_row['cantidad'] += det.cantidad

    grupos = []
    for key in sorted(grupos_map):
        grupo = grupos_map[key]
        grupo['items'] = sorted(grupo['items'].values(), key=lambda row: _orden_operativo_producto(row['item']))
        grupos.append(grupo)

    movimientos_map = {}
    for det in detalles:
        mov = det.movimiento
        row = movimientos_map.setdefault(mov.pk, {
            'movimiento': mov,
            'total': Decimal('0'),
            'num_items': 0,
            'items': [],
        })
        row['total'] += det.cantidad
        row['num_items'] += 1
        row['items'].append(det)
    movimientos = sorted(
        movimientos_map.values(),
        key=lambda row: row['movimiento'].fecha_movimiento,
        reverse=True,
    )

    return {
        'cliente': cliente,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'agrupar_por': agrupar_por,
        'producto_id': producto_id,
        'productos': productos,
        'total_general': total_general,
        'num_movimientos': len(movimiento_ids),
        'num_productos': len(producto_ids),
        'por_producto': por_producto,
        'grupos': grupos,
        'movimientos': movimientos,
    }


def _exportar_salidas_cliente_excel(context):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Cliente', context['cliente'].nombre])
    ws.append(['Desde', context['fecha_inicio'].strftime('%d/%m/%Y')])
    ws.append(['Hasta', context['fecha_fin'].strftime('%d/%m/%Y')])
    ws.append(['Total fardos/items', float(context['total_general'])])
    ws.append(['Movimientos', context['num_movimientos']])
    ws.append(['Productos distintos', context['num_productos']])
    for cell in ws['A']:
        cell.font = Font(bold=True)

    ws_prod = wb.create_sheet('Por producto')
    ws_prod.append(['Producto', 'Código', 'Cantidad total', 'Unidad'])
    for row in ws_prod[1]:
        row.font = Font(bold=True)
    for r in context['por_producto']:
        ws_prod.append([r['item'].nombre, r['item'].codigo, float(r['cantidad']), r['unidad']])

    ws_mov = wb.create_sheet('Movimientos detallados')
    ws_mov.append(['Fecha movimiento', 'Movimiento ID', 'Usuario', 'Producto', 'Código', 'Cantidad', 'Unidad', 'Motivo'])
    for row in ws_mov[1]:
        row.font = Font(bold=True)
    for mov_row in context['movimientos']:
        mov = mov_row['movimiento']
        for det in mov_row['items']:
            ws_mov.append([
                timezone.localtime(mov.fecha_movimiento).strftime('%d/%m/%Y %H:%M'),
                mov.pk,
                mov.usuario.username,
                det.item.nombre,
                det.item.codigo,
                float(det.cantidad),
                det.item.unidad_medida,
                mov.motivo,
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'salidas_cliente_{context["cliente"].pk}_{context["fecha_inicio"]}_{context["fecha_fin"]}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required(_perm('ver_inventario'), raise_exception=True)
@_timed_view('cliente_salidas')
def cliente_salidas(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, activo=True)
    context = _build_salidas_cliente_context(cliente, request)
    if request.GET.get('export') == 'xlsx':
        response = _exportar_salidas_cliente_excel(context)
        if response is not None:
            return response
        messages.warning(request, 'openpyxl no está disponible para exportar Excel.')
    return render(request, 'clientes/salidas.html', context)


# ─── REPORTES ─────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_reportes'), raise_exception=True)
@_timed_view('reporte_stock_bajo')
def reporte_stock_bajo(request):
    items_bajo = [
        {'item': i, 'stock': i.stock_calc, 'deficit': i.stock_minimo - i.stock_calc}
        for i in (
            Item.objects
            .filter(activo=True)
            .select_related('categoria')
            .annotate(stock_calc=_STOCK_ANN)
            .filter(stock_calc__lte=F('stock_minimo'))
            .order_by('orden', 'nombre')
        )
    ]

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="stock_bajo.csv"'
        response.write('﻿')
        writer = csv.writer(response)
        writer.writerow(['Código', 'Nombre', 'Tipo', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Déficit', 'Unidad'])
        for r in items_bajo:
            writer.writerow([
                r['item'].codigo, r['item'].nombre,
                r['item'].get_tipo_display(),
                r['item'].categoria.nombre if r['item'].categoria else '',
                r['stock'], r['item'].stock_minimo, r['deficit'],
                r['item'].unidad_medida
            ])
        return response

    return render(request, 'reportes/stock_bajo.html', {'items_bajo': items_bajo})


@login_required
@permission_required(_perm('ver_reportes'), raise_exception=True)
@_timed_view('reporte_produccion')
def reporte_produccion(request):
    fecha_str = request.GET.get('fecha', '')
    if fecha_str:
        try:
            from datetime import datetime
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha = date.today()
    else:
        fecha = date.today()

    produccion = _calcular_produccion(fecha)

    # Detalles de los conteos camiseta usados para producción de día
    detalle_manana = []
    detalle_tarde  = []
    if produccion['tiene_manana']:
        conteo_m = (
            Conteo.objects
            .filter(fecha=fecha, turno='manana', tipo_conteo='camiseta', anulado=False)
            .order_by('fecha_hora_conteo')
            .first()
        )
        if conteo_m:
            detalle_manana = (
                ConteoDetalle.objects
                .filter(conteo=conteo_m, item__tipo='producto')
                .select_related('item')
            )

    if produccion['tiene_tarde']:
        conteo_t = (
            Conteo.objects
            .filter(fecha=fecha, turno='tarde', tipo_conteo='camiseta', anulado=False)
            .order_by('-fecha_hora_conteo')
            .first()
        )
        if conteo_t:
            detalle_tarde = (
                ConteoDetalle.objects
                .filter(conteo=conteo_t, item__tipo='producto')
                .select_related('item')
            )

    # Detalle del conteo mañana siguiente (para producción de noche)
    detalle_manana_sig = []
    if produccion['tiene_manana_sig']:
        from datetime import timedelta as _td
        fecha_sig = fecha + _td(days=1)
        conteo_ms = (
            Conteo.objects
            .filter(fecha=fecha_sig, turno='manana', tipo_conteo='camiseta', anulado=False)
            .order_by('fecha_hora_conteo')
            .first()
        )
        if conteo_ms:
            detalle_manana_sig = (
                ConteoDetalle.objects
                .filter(conteo=conteo_ms, item__tipo='producto')
                .select_related('item')
            )

    # Salidas del tramo día (entre conteo mañana y conteo tarde)
    if produccion['hora_manana'] and produccion['hora_tarde']:
        salidas_detalle_dia = (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='salida',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__gt=produccion['hora_manana'],
                movimiento__fecha_movimiento__lt=produccion['hora_tarde'],
                item__tipo='producto',
            )
            .select_related('item', 'cliente', 'movimiento')
        )
    else:
        salidas_detalle_dia = []

    # Salidas del tramo noche (entre conteo tarde y conteo mañana siguiente)
    if produccion['hora_tarde'] and produccion['hora_manana_sig']:
        salidas_detalle_noche = (
            DetalleMovimiento.objects
            .filter(
                movimiento__tipo_movimiento='salida',
                movimiento__anulado=False,
                movimiento__eliminado=False,
                movimiento__fecha_movimiento__gt=produccion['hora_tarde'],
                movimiento__fecha_movimiento__lt=produccion['hora_manana_sig'],
                item__tipo='producto',
            )
            .select_related('item', 'cliente', 'movimiento')
        )
    else:
        salidas_detalle_noche = []

    context = {
        'fecha': fecha,
        'produccion': produccion,
        'detalle_manana': detalle_manana,
        'detalle_tarde': detalle_tarde,
        'detalle_manana_sig': detalle_manana_sig,
        'salidas_detalle_dia': salidas_detalle_dia,
        'salidas_detalle_noche': salidas_detalle_noche,
        # compat con referencias antiguas al template
        'salidas_detalle': salidas_detalle_dia,
    }
    return render(request, 'reportes/produccion.html', context)


@login_required
@permission_required(_perm('ver_reportes'), raise_exception=True)
@_timed_view('reporte_produccion_avanzado')
def reporte_produccion_avanzado(request):
    """
    Reporte de producción + salidas PT usando lógica de tramos entre
    conteos consecutivos (soporta fines de semana y rangos sin conteos).

    Un tramo = par (conteo_ini, conteo_fin) consecutivos tipo Camiseta.
    Producción = total_fin − total_ini + salidas entre conteos.
    """
    from datetime import datetime as _dt

    hoy = date.today()
    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    # ── Filtros ─────────────────────────────────────────────────────────────────
    def _parse_date(key, fallback):
        try:
            return _dt.strptime(request.GET.get(key, ''), '%Y-%m-%d').date()
        except ValueError:
            return fallback

    fecha_fin    = _parse_date('fecha_fin',    hoy)
    fecha_inicio = _parse_date('fecha_inicio', fecha_fin - timedelta(days=6))
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio
    if (fecha_fin - fecha_inicio).days > 120:
        fecha_inicio = fecha_fin - timedelta(days=120)
        messages.warning(request, 'El reporte avanzado se limitó a 120 días para mantener tiempos de respuesta estables.')

    agrupar_por = request.GET.get('agrupar_por', 'dia')
    if agrupar_por not in ('dia', 'semana', 'mes'):
        agrupar_por = 'dia'

    export = request.GET.get('export', '')

    # ── Tramos de producción ────────────────────────────────────────────────────
    tramos = _calcular_tramos(fecha_inicio, fecha_fin)

    # ── Salidas de PT por fecha_movimiento.date() (KPI y desglose por producto) ─
    salidas_qs = (
        DetalleMovimiento.objects
        .filter(
            movimiento__tipo_movimiento='salida',
            movimiento__anulado=False,
            movimiento__eliminado=False,
            movimiento__fecha_movimiento__date__range=[fecha_inicio, fecha_fin],
            item__tipo='producto',
        )
        .select_related('movimiento', 'item')
        .only(
            'cantidad', 'item_id',
            'item__nombre', 'item__orden', 'item__unidad_medida', 'item__codigo',
            'movimiento__fecha_movimiento',
        )
    )
    sal_x_item: dict = {}       # item_pk → {'item': obj, 'total': qty}
    total_salidas_rango = Decimal('0')
    for det in salidas_qs:
        pk = det.item_id
        if pk not in sal_x_item:
            sal_x_item[pk] = {'item': det.item, 'total': Decimal('0')}
        sal_x_item[pk]['total'] += det.cantidad
        total_salidas_rango += det.cantidad

    # ── Totales globales ────────────────────────────────────────────────────────
    total_prod              = sum(t['produccion'] for t in tramos)
    total_salidas_formula   = sum(t['salidas']    for t in tramos)
    total_diferencia_formula= total_prod - total_salidas_formula   # consistente con tabla
    total_diferencia        = total_prod - total_salidas_rango     # para tarjeta resumen
    num_tramos           = len(tramos)
    num_dias_rango       = (fecha_fin - fecha_inicio).days + 1
    n_dia       = sum(1 for t in tramos if t['tipo'] == 'dia')
    n_noche     = sum(1 for t in tramos if t['tipo'] == 'noche')
    n_extendido = sum(1 for t in tramos if t['tipo'] == 'extendido')

    # ── Agrupación ──────────────────────────────────────────────────────────────
    def _gkey(tramo):
        f = tramo['fecha_asignada']
        if agrupar_por == 'dia':    return f
        if agrupar_por == 'semana': iso = f.isocalendar(); return (iso[0], iso[1])
        return (f.year, f.month)

    def _glabel(key):
        if agrupar_por == 'dia':
            return key.strftime('%d/%m/%Y')
        if agrupar_por == 'semana':
            y, w = key; return f'Sem. {w:02d} / {y}'
        y, m = key; return f'{MESES[m]} {y}'

    grupos = []
    cur_key = cur_g = None
    for t in tramos:
        k = _gkey(t)
        if k != cur_key:
            if cur_g:
                cur_g['diferencia'] = cur_g['prod_total'] - cur_g['salidas']
                grupos.append(cur_g)
            cur_key = k
            cur_g = {
                'key': k, 'label': _glabel(k),
                'prod_total':      Decimal('0'),
                'salidas':         Decimal('0'),
                'diferencia':      Decimal('0'),
                'num_tramos':      0,
                'tramos':          [],
            }
        cur_g['prod_total'] += t['produccion']
        cur_g['salidas']    += t['salidas']
        cur_g['num_tramos'] += 1
        cur_g['tramos'].append(t)
    if cur_g:
        cur_g['diferencia'] = cur_g['prod_total'] - cur_g['salidas']
        grupos.append(cur_g)

    # ── Por producto ────────────────────────────────────────────────────────────
    prod_x_item: dict = {}   # item_pk → Decimal
    for t in tramos:
        for pk, qty in t['por_item'].items():
            prod_x_item[pk] = prod_x_item.get(pk, Decimal('0')) + qty

    all_pks = set(prod_x_item.keys()) | set(sal_x_item.keys())
    items_pt = (
        Item.objects
        .filter(pk__in=all_pks, tipo='producto', activo=True)
        .order_by('orden', 'nombre')
    )
    por_producto = []
    for item in items_pt:
        prod_item = prod_x_item.get(item.pk, Decimal('0'))
        sal_item  = sal_x_item.get(item.pk, {}).get('total', Decimal('0'))
        por_producto.append({
            'item':       item,
            'prod_total': prod_item,
            'salidas':    sal_item,
            'diferencia': prod_item - sal_item,
        })

    # ── CSV export ──────────────────────────────────────────────────────────────
    if export == 'csv':
        resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        fname = f'produccion_{fecha_inicio}_{fecha_fin}_{agrupar_por}.csv'
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        w = csv.writer(resp)
        w.writerow(['Período', 'Tipo', 'Tramo', 'Duración (h)', 'Producción', 'Salidas (fórmula)', 'Diferencia'])
        tipo_labels = {'dia': 'Día', 'noche': 'Noche', 'extendido': 'Extendido'}
        for g in grupos:
            for t in g['tramos']:
                w.writerow([
                    g['label'],
                    tipo_labels.get(t['tipo'], t['tipo']),
                    t['label_rango'],
                    t['duracion_h'],
                    t['produccion'],
                    t['salidas'],
                    t['produccion'] - t['salidas'],
                ])
        w.writerow([])
        w.writerow(['Producto', 'Producción', 'Salidas (rango)', 'Diferencia'])
        for p in por_producto:
            w.writerow([p['item'].nombre, p['prod_total'], p['salidas'], p['diferencia']])
        return resp

    return render(request, 'reportes/produccion_avanzado.html', {
        'fecha_inicio':           fecha_inicio,
        'fecha_fin':              fecha_fin,
        'agrupar_por':            agrupar_por,
        'grupos':                 grupos,
        'tramos':                 tramos,
        'por_producto':           por_producto,
        'total_prod':             total_prod,
        'total_salidas':              total_salidas_rango,
        'total_salidas_formula':      total_salidas_formula,
        'total_diferencia':           total_diferencia,            # prod - salidas rango (tarjeta)
        'total_diferencia_formula':   total_diferencia_formula,    # prod - salidas fórmula (tabla)
        'num_tramos':             num_tramos,
        'num_dias_rango':         num_dias_rango,
        'n_dia':                  n_dia,
        'n_noche':                n_noche,
        'n_extendido':            n_extendido,
    })


# ─── API ──────────────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('editar_item'), raise_exception=True)
@require_POST
def api_categoria_nueva(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
        cat, created = Categoria.objects.get_or_create(nombre=nombre)
        return JsonResponse({'id': cat.pk, 'nombre': cat.nombre, 'created': created})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_item_info(request, pk):
    item = get_object_or_404(Item, pk=pk)
    stocks = list(Stock.objects.filter(item=item).values('ubicacion__id', 'ubicacion__nombre', 'cantidad_actual'))
    return JsonResponse({
        'tipo': item.tipo,
        'unidad_medida': item.unidad_medida,
        'stock_total': str(item.stock_total()),
        'stocks': [
            {'ubicacion_id': s['ubicacion__id'], 'ubicacion': s['ubicacion__nombre'],
             'cantidad': str(s['cantidad_actual'])}
            for s in stocks
        ]
    })


# ─── PRODUCCIÓN ───────────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('registrar_produccion'), raise_exception=True)
def produccion_nueva(request):
    ubicaciones = Ubicacion.objects.all()

    if request.method == 'POST':
        form = ProduccionForm(request.POST)
        if form.is_valid():
            item = form.cleaned_data['item']
            cantidad = form.cleaned_data['cantidad']
            ubicacion = form.cleaned_data['ubicacion_destino']
            fecha_movimiento = form.cleaned_data['fecha_movimiento']
            motivo = form.cleaned_data.get('motivo', '') or f'Producción registrada'

            if timezone.is_naive(fecha_movimiento):
                fecha_movimiento = timezone.make_aware(fecha_movimiento)

            with transaction.atomic():
                mov_prod = MovimientoInventario.objects.create(
                    tipo_movimiento='entrada',
                    motivo=motivo,
                    fecha_movimiento=fecha_movimiento,
                    usuario=request.user,
                )
                det_prod = DetalleMovimiento.objects.create(
                    movimiento=mov_prod,
                    item=item,
                    cantidad=cantidad,
                    ubicacion_destino=ubicacion,
                )
                _aplicar_efecto_detalle(det_prod)

            send_event('production_created', {
                'item': item.nombre, 'codigo': item.codigo,
                'cantidad': str(cantidad), 'ubicacion': ubicacion.nombre,
                'usuario': request.user.username,
                'fecha_movimiento': fecha_movimiento.isoformat(),
            })
            notify_stock(item, movimiento='produccion', usuario=request.user.username)

            messages.success(request, f'Producción registrada: {cantidad} {item.unidad_medida} de {item.nombre}.')
            return redirect('produccion_nueva')
    else:
        form = ProduccionForm(initial={
            'fecha_movimiento': timezone.localtime(timezone.now()).strftime('%Y-%m-%dT%H:%M'),
        })

    return render(request, 'produccion/form.html', {
        'form': form,
        'ubicaciones': ubicaciones,
    })


# ─── IMPORTAR EXCEL ───────────────────────────────────────────────────────────

@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def importar_items(request):
    resultados = None

    if request.method == 'POST':
        form = ImportarItemsForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            resultados = _procesar_excel_items(archivo)
    else:
        form = ImportarItemsForm()

    return render(request, 'importar/form.html', {
        'form': form,
        'resultados': resultados,
    })


def _procesar_excel_items(archivo):
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl no está instalado.', 'creados': 0, 'actualizados': 0, 'errores': []}

    creados = 0
    actualizados = 0
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {'error': f'No se pudo leer el archivo: {e}', 'creados': 0, 'actualizados': 0, 'errores': []}

    if not rows:
        return {'error': 'El archivo está vacío.', 'creados': 0, 'actualizados': 0, 'errores': []}

    # First row must be headers; find column indices
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    col = {name: headers.index(name) for name in ('codigo', 'nombre', 'tipo', 'unidad_medida') if name in headers}

    required = ['codigo', 'nombre', 'tipo', 'unidad_medida']
    missing = [r for r in required if r not in col]
    if missing:
        return {
            'error': f'Columnas requeridas faltantes: {", ".join(missing)}.',
            'creados': 0, 'actualizados': 0, 'errores': [],
        }

    col_opt = lambda name: headers.index(name) if name in headers else None
    idx_desc = col_opt('descripcion')
    idx_cat = col_opt('categoria')
    idx_min = col_opt('stock_minimo')

    TIPOS_VALIDOS = {'producto', 'repuesto', 'consumible'}

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            codigo = str(row[col['codigo']] or '').strip()
            nombre = str(row[col['nombre']] or '').strip()
            tipo = str(row[col['tipo']] or '').strip().lower()
            unidad = str(row[col['unidad_medida']] or '').strip()

            if not codigo or not nombre:
                errores.append(f'Fila {row_num}: código o nombre vacío, se omite.')
                continue
            if tipo not in TIPOS_VALIDOS:
                errores.append(f'Fila {row_num} ({codigo}): tipo "{tipo}" inválido (usar: producto/repuesto/consumible).')
                continue
            if not unidad:
                errores.append(f'Fila {row_num} ({codigo}): unidad de medida requerida.')
                continue

            descripcion = str(row[idx_desc] or '').strip() if idx_desc is not None else ''

            categoria = None
            if idx_cat is not None and row[idx_cat]:
                cat_nombre = str(row[idx_cat]).strip()
                if cat_nombre:
                    categoria, _ = Categoria.objects.get_or_create(nombre=cat_nombre)

            stock_minimo = Decimal('0')
            if idx_min is not None and row[idx_min] is not None:
                try:
                    stock_minimo = Decimal(str(row[idx_min]))
                except Exception:
                    pass

            item, created = Item.objects.update_or_create(
                codigo=codigo,
                defaults={
                    'nombre': nombre,
                    'tipo': tipo,
                    'unidad_medida': unidad,
                    'descripcion': descripcion,
                    'categoria': categoria,
                    'stock_minimo': stock_minimo,
                    'activo': True,
                },
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        except Exception as e:
            errores.append(f'Fila {row_num}: error inesperado — {e}')

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'error': None}


@login_required
@permission_required(_perm('importar_excel'), raise_exception=True)
def descargar_plantilla(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'openpyxl no está instalado.')
        return redirect('importar_items')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Items'

    headers = ['codigo', 'nombre', 'tipo', 'unidad_medida', 'stock_minimo', 'categoria', 'descripcion']
    header_fill = PatternFill('solid', fgColor='003087')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    ejemplos = [
        ['BOLSA-001', 'Bolsa de polietileno 10x15', 'producto', 'unidad', 100, 'Bolsas', ''],
        ['REP-001', 'Rodamiento 6205', 'repuesto', 'pieza', 5, 'Repuestos', 'Para máquina selladora'],
        ['CONS-001', 'Cinta adhesiva', 'consumible', 'rollo', 10, 'Consumibles', ''],
    ]
    for row_data in ejemplos:
        ws.append(row_data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_items.xlsx"'
    return response


# ─── USUARIOS ─────────────────────────────────────────────────────────────────

def _superuser_required(view_func):
    """Restrict view to superusers only; log denied attempts."""
    from functools import wraps
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not request.user.is_superuser:
            ip = _get_client_ip(request)
            send_security_event(
                'forbidden_403',
                title    = '🚫 Acceso denegado a ruta administrativa',
                user     = request.user.username,
                ip       = ip,
                path     = request.path,
                user_agent = request.META.get('HTTP_USER_AGENT', '')[:200],
                message  = f'Usuario "{request.user.username}" sin superuser intentó acceder a {request.path}',
            )
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return _wrapped


def _get_grupo(user):
    return user.groups.first().name if user.groups.exists() else ''


@login_required
@_superuser_required
def usuario_lista(request):
    usuarios = (
        User.objects.prefetch_related('groups')
        .order_by('username')
    )
    data = [
        {'user': u, 'grupo': _get_grupo(u)}
        for u in usuarios
    ]
    return render(request, 'usuarios/lista.html', {'data': data})


@login_required
@_superuser_required
def usuario_crear(request):
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = User.objects.create_user(
                username=cd['username'],
                password=cd['password'],
                first_name=cd.get('first_name', ''),
                last_name=cd.get('last_name', ''),
                email=cd.get('email', ''),
            )
            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    user.groups.set([grupo])
                except Group.DoesNotExist:
                    pass
            messages.success(request, f'Usuario "{user.username}" creado exitosamente.')
            return redirect('usuario_lista')
    else:
        form = UsuarioCrearForm()

    return render(request, 'usuarios/form.html', {'form': form, 'titulo': 'Nuevo Usuario'})


@login_required
@_superuser_required
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)

    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            usuario.first_name = cd.get('first_name', '')
            usuario.last_name = cd.get('last_name', '')
            usuario.email = cd.get('email', '')
            usuario.is_active = cd.get('is_active', True)

            nueva_pass = cd.get('nueva_password')
            if nueva_pass:
                usuario.set_password(nueva_pass)

            usuario.save()

            grupo_nombre = cd.get('grupo')
            if grupo_nombre:
                try:
                    grupo = Group.objects.get(name=grupo_nombre)
                    usuario.groups.set([grupo])
                except Group.DoesNotExist:
                    usuario.groups.clear()
            else:
                usuario.groups.clear()

            messages.success(request, f'Usuario "{usuario.username}" actualizado.')
            return redirect('usuario_lista')
    else:
        form = UsuarioEditarForm(initial={
            'first_name': usuario.first_name,
            'last_name': usuario.last_name,
            'email': usuario.email,
            'is_active': usuario.is_active,
            'grupo': _get_grupo(usuario),
        })

    return render(request, 'usuarios/form.html', {
        'form': form,
        'titulo': f'Editar: {usuario.username}',
        'usuario': usuario,
    })


# ─── REPORTE CONSUMO PIGMENTOS ────────────────────────────────────────────────

@login_required
@permission_required(_perm('ver_reportes'), raise_exception=True)
@_timed_view('reporte_consumo_pigmentos')
def reporte_consumo_pigmentos(request):
    """
    Reporte de consumo de pigmentos en un rango de fechas.

    Consumo = ajustes negativos sobre ítems de categoría Pigmentos.
    Muestra: consumo total, promedio diario, stock actual, días de cobertura,
    estado (ok / bajo / crítico) y pedido sugerido.
    """
    hoy = date.today()

    # ── Filtros ───────────────────────────────────────────────────────────────
    def _parse_date(s, fallback):
        try:
            return date.fromisoformat(s) if s else fallback
        except ValueError:
            return fallback

    fecha_inicio = _parse_date(request.GET.get('fecha_inicio', ''), hoy - timedelta(days=30))
    fecha_fin    = _parse_date(request.GET.get('fecha_fin', ''),    hoy)
    if fecha_fin < fecha_inicio:
        fecha_fin = fecha_inicio

    pigmento_pk  = request.GET.get('pigmento', '').strip()
    try:
        dias_objetivo = max(1, int(request.GET.get('dias_objetivo', 14)))
    except (ValueError, TypeError):
        dias_objetivo = 14

    dias_rango = max(1, (fecha_fin - fecha_inicio).days + 1)

    # ── Pigmentos activos ─────────────────────────────────────────────────────
    pigmentos_qs = (
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .select_related('categoria')
        .annotate(stock_calc=_STOCK_ANN)
        .order_by('orden', 'nombre')
    )
    if pigmento_pk:
        pigmentos_qs = pigmentos_qs.filter(pk=pigmento_pk)

    todos_pigmentos = (
        Item.objects
        .filter(activo=True, tipo='consumible', categoria__nombre__iexact='Pigmentos')
        .order_by('orden', 'nombre')
    )

    # ── Consumo por ítem: ajustes negativos en el rango ───────────────────────
    consumos_base = DetalleMovimiento.objects.filter(
        movimiento__tipo_movimiento='ajuste',
        movimiento__anulado=False,
        movimiento__eliminado=False,
        movimiento__fecha_movimiento__date__gte=fecha_inicio,
        movimiento__fecha_movimiento__date__lte=fecha_fin,
        item__tipo='consumible',
        item__categoria__nombre__iexact='Pigmentos',
        cantidad__lt=0,
    )
    if pigmento_pk:
        consumos_base = consumos_base.filter(item_id=pigmento_pk)

    consumo_por_item = {
        row['item_id']: abs(row['total'])
        for row in consumos_base.values('item_id').annotate(total=Sum('cantidad'))
    }

    # ── Construir tabla de resultados ─────────────────────────────────────────
    ESTADO_LABELS = {
        'ok':          'OK',
        'bajo':        'Bajo',
        'critico':     'Crítico',
        'sin_consumo': 'Sin consumo',
    }

    resultados = []
    total_consumo  = Decimal('0')
    total_criticos = 0
    total_pedido   = Decimal('0')

    for pig in pigmentos_qs:
        consumo = consumo_por_item.get(pig.pk, Decimal('0'))
        stock   = pig.stock_calc or Decimal('0')

        if consumo > 0:
            promedio_diario = consumo / Decimal(str(dias_rango))
            dias_cob = float(stock / promedio_diario) if promedio_diario else None
            pedido   = max(Decimal('0'), promedio_diario * Decimal(str(dias_objetivo)) - stock)
        else:
            promedio_diario = Decimal('0')
            dias_cob = None
            pedido   = Decimal('0')

        if dias_cob is None:
            estado = 'sin_consumo'
        elif dias_cob < 3:
            estado = 'critico'
            total_criticos += 1
        elif dias_cob <= 7:
            estado = 'bajo'
        else:
            estado = 'ok'

        total_consumo += consumo
        total_pedido  += pedido

        resultados.append({
            'item':           pig,
            'consumo':        consumo,
            'promedio_diario': round(promedio_diario, 2),
            'stock':          stock,
            'dias_cobertura': round(dias_cob, 1) if dias_cob is not None else None,
            'pedido':         round(pedido, 2),
            'estado':         estado,
            'estado_label':   ESTADO_LABELS[estado],
        })

    # ── Detalle de movimientos (solo cuando se filtra un pigmento) ─────────────
    detalle_movimientos = []
    if pigmento_pk:
        detalle_movimientos = list(
            consumos_base
            .select_related('movimiento', 'movimiento__usuario', 'item')
            .order_by('-movimiento__fecha_movimiento')
        )

    # ── Exportar CSV ──────────────────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        fname = f'consumo_pigmentos_{fecha_inicio}_{fecha_fin}.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        response.write('﻿')  # BOM para Excel
        writer = csv.writer(response)
        writer.writerow([
            'Pigmento', 'Código', 'Consumo total', 'Promedio diario',
            'Stock actual', 'Cobertura (días)', 'Pedido sugerido', 'Unidad', 'Estado',
        ])
        for r in resultados:
            writer.writerow([
                r['item'].nombre, r['item'].codigo,
                r['consumo'], r['promedio_diario'],
                r['stock'],
                r['dias_cobertura'] if r['dias_cobertura'] is not None else '',
                r['pedido'], r['item'].unidad_medida,
                r['estado_label'],
            ])
        return response

    return render(request, 'reportes/pigmentos.html', {
        'resultados':          resultados,
        'todos_pigmentos':     todos_pigmentos,
        'fecha_inicio':        fecha_inicio,
        'fecha_fin':           fecha_fin,
        'dias_rango':          dias_rango,
        'dias_objetivo':       dias_objetivo,
        'pigmento_pk':         pigmento_pk,
        'total_consumo':       total_consumo,
        'total_criticos':      total_criticos,
        'total_pedido':        total_pedido,
        'detalle_movimientos': detalle_movimientos,
    })
